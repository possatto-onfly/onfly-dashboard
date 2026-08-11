"""
Onfly Analytics Dashboard
"""

from __future__ import annotations
import re
import os
import numpy as np
import streamlit as st
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta
import pandas as pd
from google.cloud import bigquery
import plotly.express as px
import plotly.graph_objects as go

# ─── Anthropic (aba Análises) ─────────────────────────────────────────────────
def _get_anthropic_client():
    """Retorna cliente Anthropic, lendo a chave no momento da chamada."""
    try:
        import anthropic as _anthropic
        # Tenta secrets do projeto
        key = ""
        try:
            key = st.secrets["ANTHROPIC_API_KEY"]
        except Exception:
            pass
        # Fallback: lê direto do arquivo
        if not key:
            _secrets_path = os.path.join(os.path.dirname(__file__), ".streamlit", "secrets.toml")
            if os.path.exists(_secrets_path):
                import re as _re
                _txt = open(_secrets_path).read()
                _m = _re.search(r'ANTHROPIC_API_KEY\s*=\s*["\']([^"\']+)["\']', _txt)
                if _m:
                    key = _m.group(1)
        # Fallback: variável de ambiente
        if not key:
            key = os.environ.get("ANTHROPIC_API_KEY", "")
        return _anthropic.Anthropic(api_key=key) if key else None
    except Exception:
        return None

try:
    import anthropic as _anthropic
    _anthropic_client = None   # será inicializado na primeira chamada
except Exception:
    _anthropic = None
    _anthropic_client = None

# ─── Constantes ───────────────────────────────────────────────────────────────
PROJECT_ID  = "dw-onfly-prd"
TABLE       = "dw-onfly-prd.travel_core.silver_all_emissions"
TABLE_FARE  = "dw-onfly-prd.travel_core.silver_all_tickets"   # substituiu silver_onfly_old_fare_family_flight
TABLE_FLIGHT_ORDERS = "dw-onfly-prd.travel_core.silver_flight_orders"  # taxas de embarque (amount_taxes_v3_currency_brl)
TABLE_SEG   = "dw-onfly-prd.travel_core.gold_item_summaries_flight_by_protocol_traveler_segment_leg"
TABLE_COMP  = "dw-onfly-prd.onfly_dim_shared.silver_companies"

ONFLY_BLUE   = "#1890FF"
ONFLY_ORANGE = "#FF6B00"
ONFLY_GREEN  = "#22C55E"
IATA_NOMES = {
    "AD": "Azul",                    "G3": "GOL",                    "LA": "LATAM",
    "AA": "American Airlines",       "UA": "United Airlines",        "DL": "Delta Air Lines",
    "AF": "Air France",              "KL": "KLM",                    "BA": "British Airways",
    "LH": "Lufthansa",               "IB": "Iberia",                 "TP": "TAP Air Portugal",
    "EK": "Emirates",                "QR": "Qatar Airways",          "TK": "Turkish Airlines",
    "AC": "Air Canada",              "AV": "Avianca",                "CM": "Copa Airlines",
    "AR": "Aerolíneas Argentinas",   "AM": "Aeromexico",             "VB": "VivaAerobus",
    "AZ": "ITA Airways",             "LX": "Swiss",                  "UX": "Air Europa",
    "H2": "Sky Airline",             "ET": "Ethiopian Airlines",     "WN": "Southwest Airlines",
    "Y4": "Volaris",                 "JA": "JetSMART",               "JJ": "LATAM Brasil",
    "O6": "Avianca Brasil",          "2Z": "Passaredo",              "M3": "LATAM Cargo",
    "P3": "LATAM Cargo Brasil",      "T4": "LATAM Ecuador",          "XL": "LATAM Ecuador",
    "4M": "LATAM Argentina",         "MH": "Malaysia Airlines",      "CX": "Cathay Pacific",
    "JL": "Japan Airlines",          "NH": "ANA",                    "SQ": "Singapore Airlines",
    "OS": "Austrian Airlines",       "SN": "Brussels Airlines",      "LO": "LOT Polish Airlines",
    "SK": "SAS",                     "AY": "Finnair",                "VY": "Vueling",
    "FR": "Ryanair",                 "U2": "easyJet",                "W6": "Wizz Air",
    "FZ": "flydubai",                "G9": "Air Arabia",             "WY": "Oman Air",
    "MS": "EgyptAir",                "AT": "Royal Air Maroc",        "SA": "South African Airways",
    "KC": "Air Astana",              "SU": "Aeroflot",               "PS": "Ukraine International",
    "OU": "Croatia Airlines",        "RO": "TAROM",                  "BT": "airBaltic",
    "DT": "TAAG Angola Airlines",
}

# Códigos que devem ser agrupados sob um único código canônico
IATA_NORMALIZACAO = {
    "JJ": "LA",           # LATAM Brasil (ex-TAM)
    "XL": "LA",           # LATAM Ecuador
    "4M": "LA",           # LATAM Argentina
    "T4": "LA",           # LATAM Ecuador (antigo)
    "LP": "LA",           # LATAM Peru
    "P3": "LA",           # LATAM Cargo Brasil
    "M3": "LA",           # LATAM Cargo
    "TAP": "TP",                    # TAP Air Portugal (variante sem código)
    "TAP PORTUGAL": "TP",           # TAP Air Portugal (nome por extenso)
    "TAAG": "DT",                   # TAAG Angola Airlines (sigla)
    "TAAG ANGOLANA AIRLINES": "DT", # TAAG Angola Airlines (nome por extenso)
    "TAAG ANGOLA AIRLINES": "DT",   # TAAG Angola Airlines (variante em inglês)
    "COPA": "CM",                   # Copa Airlines (nome por extenso)
    "COPA AIRLINES": "CM",          # Copa Airlines (variante em inglês)
}

# Mapeamento de nomes usados em silver_flight_orders.standard_airline → código IATA canônico
# Cobre variantes tipadas incorretamente e formas regionais não presentes em IATA_NOMES
_SFO_NAME_TO_IATA: dict[str, str] = {
    "AMERICA AIRLINES":       "AA",   # typo (American)
    "UINITED AIRLINES":       "UA",   # typo (United)
    "AEROVÍAS DE MÉXICO":     "AM",
    "AEROVIAS DE MEXICO":     "AM",
    "DELTA":                  "DL",   # forma curta
    "KLM CIA":                "KL",
    "LUFHANSA":               "LH",   # typo (Lufthansa)
    "BRISTISH AIRWAYS":       "BA",   # typo (British)
    "IBÉRIA LINEAS":          "IB",
    "IBERIA LINEAS":          "IB",
    "SWISS AIRLINE":          "LX",
    "LAN":                    "LA",   # marca antiga LATAM
    "TAM":                    "LA",   # marca antiga LATAM
    "AZUL CONECTA":           "AD",
    "AEROLINEAS ARGENTINAS":  "AR",   # sem acento
    "AIR CHINA":              "CA",
}

# Canais considerados emissão Manual (agências / GDS offline)
_CANAIS_MANUAL = frozenset([
    "AMADEUS", "CONSOLIDAMADEUS", "ONFLYAMADEUS",
    "FLYTOURLAG3", "FLYTOURTO", "FLYTOURWITHOUTDU", "FLYTOURADTOCREDITCARD",
    "REXTUR", "REXTUR_IATA", "CONSOLID", "AERTICKET", "CONFIANCA", "BRT",
])

AEROPORTO_NOMES = {
    # Brasil
    "GRU": "São Paulo / Guarulhos",    "CGH": "São Paulo / Congonhas",
    "VCP": "Campinas / Viracopos",     "GIG": "Rio de Janeiro / Galeão",
    "SDU": "Rio de Janeiro / Santos Dumont",
    "BSB": "Brasília",                 "CNF": "Belo Horizonte / Confins",
    "PLU": "Belo Horizonte / Pampulha","SSA": "Salvador",
    "REC": "Recife",                   "FOR": "Fortaleza",
    "MAO": "Manaus",                   "BEL": "Belém",
    "POA": "Porto Alegre",             "FLN": "Florianópolis",
    "CWB": "Curitiba",                 "VIX": "Vitória",
    "MCZ": "Maceió",                   "NAT": "Natal",
    "THE": "Teresina",                 "SLZ": "São Luís",
    "JPA": "João Pessoa",              "AJU": "Aracaju",
    "PMW": "Palmas",                   "CGB": "Cuiabá",
    "CGR": "Campo Grande",             "PVH": "Porto Velho",
    "RBR": "Rio Branco",               "MCP": "Macapá",
    "BVB": "Boa Vista",                "STM": "Santarém",
    "IMP": "Imperatriz",               "IOS": "Ilhéus",
    "BPS": "Porto Seguro",             "CXJ": "Caxias do Sul",
    "LDB": "Londrina",                 "MGF": "Maringá",
    "JOI": "Joinville",                "NVT": "Navegantes",
    "XAP": "Chapecó",                  "UDI": "Uberlândia",
    "GYN": "Goiânia",                  "AQA": "Araraquara",
    "RAO": "Ribeirão Preto",           "JDO": "Juazeiro do Norte",
    "MNX": "Manicoré",                 "PPB": "Presidente Prudente",
    "SJP": "São José do Rio Preto",    "BAU": "Bauru",
    "CFB": "Cabo Frio",                "PMG": "Ponta Porã",
    # América do Norte
    "JFK": "Nova York / JFK",          "EWR": "Newark",
    "LGA": "Nova York / LaGuardia",    "MIA": "Miami",
    "LAX": "Los Angeles",              "ORD": "Chicago / O'Hare",
    "ATL": "Atlanta",                  "DFW": "Dallas / Fort Worth",
    "IAH": "Houston",                  "BOS": "Boston",
    "SFO": "San Francisco",            "SEA": "Seattle",
    "DCA": "Washington / Reagan",      "IAD": "Washington / Dulles",
    "MCO": "Orlando",                  "FLL": "Fort Lauderdale",
    "LAS": "Las Vegas",                "PHX": "Phoenix",
    "YYZ": "Toronto",                  "YUL": "Montreal",
    "YVR": "Vancouver",
    # América Latina
    "EZE": "Buenos Aires / Ezeiza",    "AEP": "Buenos Aires / Aeroparque",
    "SCL": "Santiago",                 "BOG": "Bogotá",
    "LIM": "Lima",                     "UIO": "Quito",
    "GYE": "Guayaquil",               "MVD": "Montevidéu",
    "ASU": "Assunção",                 "MEX": "Cidade do México",
    "CUN": "Cancún",                   "PTY": "Cidade do Panamá",
    "SJO": "San José (CR)",           "GUA": "Cidade da Guatemala",
    "HAV": "Havana",                   "SDQ": "Santo Domingo",
    "SJU": "San Juan",                 "CCS": "Caracas",
    "LPB": "La Paz",                   "VVI": "Santa Cruz (BO)",
    "MDE": "Medellín",                 "CTG": "Cartagena",
    # Europa
    "LHR": "Londres / Heathrow",       "LGW": "Londres / Gatwick",
    "CDG": "Paris / Charles de Gaulle","ORY": "Paris / Orly",
    "AMS": "Amsterdã",                 "FRA": "Frankfurt",
    "MAD": "Madri",                    "BCN": "Barcelona",
    "FCO": "Roma / Fiumicino",         "MXP": "Milão / Malpensa",
    "LIS": "Lisboa",                   "OPO": "Porto",
    "MUC": "Munique",                  "VIE": "Viena",
    "ZRH": "Zurique",                  "GVA": "Genebra",
    "BRU": "Bruxelas",                 "CPH": "Copenhague",
    "ARN": "Estocolmo",                "HEL": "Helsinki",
    "OSL": "Oslo",                     "DUB": "Dublin",
    "IST": "Istambul",                 "ATH": "Atenas",
    # África / Médio Oriente / Ásia
    "DXB": "Dubai",                    "DOH": "Doha",
    "AUH": "Abu Dhabi",                "LUX": "Luxemburgo",
    "NBO": "Nairóbi",                  "JNB": "Joanesburgo",
    "LAD": "Luanda",                   "DKR": "Dakar",
    "HKG": "Hong Kong",               "SIN": "Singapura",
    "NRT": "Tóquio / Narita",          "ICN": "Seul",
    "PEK": "Pequim",                   "PVG": "Xangai",
    "SYD": "Sydney",                   "MEL": "Melbourne",
}

COR_LATAM    = "#D10B2F"
COR_AZUL     = "#003DA5"
COR_GOL      = "#FF6600"
COR_AMADEUS  = "#00A1DE"
AMADEUS_CONSOLIDATORS = ("AMADEUS", "ONFLYAMADEUS", "CONSOLIDAMADEUS")

# ─── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Sourcing Aéreo",
    page_icon="✈️",
    layout="wide",
    initial_sidebar_state="expanded",
)

if "dark_mode" not in st.session_state:
    st.session_state.dark_mode = False

# ─── CSS global ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
  /* ── Global ── */
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
  html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
  .stApp { background: #F0F2F5; }
  #MainMenu, footer { visibility: hidden; }

  /* ── Sidebar ── */
  [data-testid="stSidebar"] {
    background: #FFFFFF;
    border-right: 1px solid #E8ECF0;
  }
  [data-testid="stSidebar"] label,
  [data-testid="stSidebar"] p,
  [data-testid="stSidebar"] span,
  [data-testid="stSidebar"] div { color: #4A5568; }

  [data-testid="collapsedControl"] {
    display: flex !important;
    visibility: visible !important;
    background: #1890FF;
    border-radius: 0 8px 8px 0;
    color: white;
  }

  /* ── KPI Cards ── */
  .kpi-card {
    background: #FFFFFF;
    border-radius: 12px;
    padding: 20px 24px 16px 24px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06), 0 4px 16px rgba(0,0,0,0.04);
    border-top: 3px solid #1890FF;
    border: 1px solid #EEF2F7;
    border-top: 3px solid #1890FF;
  }
  .kpi-card.orange { border-top-color: #FF6B00; }
  .kpi-card.green  { border-top-color: #52C41A; }
  .kpi-card.red    { border-top-color: #FF4D4F; }
  .kpi-card.azul   { border-top-color: #003DA5; }
  .kpi-card.gol    { border-top-color: #FF6600; }
  .kpi-card.amadeus{ border-top-color: #13C2C2; }

  .kpi-label {
    font-size: 0.7rem;
    font-weight: 600;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: #8C9BAB;
    margin: 0 0 8px 0;
  }
  .kpi-value {
    font-size: 1.65rem;
    font-weight: 700;
    color: #1A202C;
    line-height: 1.2;
    margin: 0 0 6px 0;
  }
  .kpi-delta {
    font-size: 0.76rem;
    font-weight: 600;
    margin: 0;
  }
  .kpi-delta.up   { color: #52C41A; }
  .kpi-delta.down { color: #FF4D4F; }
  .kpi-delta.neu  { color: #8C9BAB; }

  /* ── Section headers ── */
  .sec-header-wrap {
    margin-top: 0 !important;
    padding-top: 48px;
  }
  .sec-header {
    font-size: 0.82rem;
    font-weight: 700;
    color: #5A6475;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    padding-bottom: 10px;
    border-bottom: 1px solid #E8ECF0;
    margin: 0 0 20px 0;
  }

  /* ── Period badge ── */
  .period-badge {
    background: #F7FAFF;
    border: 1px solid #D6E4FF;
    border-radius: 10px;
    padding: 8px 12px;
    margin-top: 10px;
    display: flex;
    gap: 12px;
  }
  .period-badge .period-col { flex: 1; }
  .period-badge .label {
    font-size: 0.58rem;
    color: #8C9BAB;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin: 0;
  }
  .period-badge .value {
    font-size: 0.72rem;
    font-weight: 600;
    color: #1890FF;
    margin: 1px 0 0 0;
  }

  /* ── Abas nativas (st.tabs) ── */
  button[data-baseweb="tab"] {
    font-weight: 600;
    color: #8C9BAB;
    font-size: 0.87rem;
  }
  button[data-baseweb="tab"][aria-selected="true"] {
    color: #1890FF !important;
    font-weight: 700 !important;
  }
  button[data-baseweb="tab"][aria-selected="true"]::after {
    background-color: #1890FF !important;
    height: 2px !important;
    border-radius: 2px 2px 0 0;
  }

  /* ── Comp badge ── */
  .comp-badge {
    background: #FFF7E6;
    border: 1px solid #FFD591;
    border-radius: 8px;
    padding: 8px 12px;
    margin-top: 8px;
    font-size: 0.72rem;
    color: #D46B08;
    font-weight: 500;
  }

  /* ── Botão flutuante de compartilhar ── */
  .share-fab {
    position: fixed;
    bottom: 32px;
    right: 32px;
    width: 52px;
    height: 52px;
    border-radius: 50%;
    background: #1890FF;
    color: white;
    font-size: 1.4rem;
    display: flex;
    align-items: center;
    justify-content: center;
    cursor: pointer;
    box-shadow: 0 4px 16px rgba(24,144,255,0.45);
    z-index: 9999;
    transition: transform 0.15s, box-shadow 0.15s;
    user-select: none;
  }
  .share-fab:hover {
    transform: scale(1.08);
    box-shadow: 0 6px 24px rgba(24,144,255,0.55);
  }
  .share-menu {
    position: fixed;
    bottom: 94px;
    right: 32px;
    background: #FFFFFF;
    border: 1px solid #E8ECF0;
    border-radius: 12px;
    box-shadow: 0 8px 32px rgba(0,0,0,0.12);
    z-index: 9998;
    overflow: hidden;
    display: none;
    flex-direction: column;
    min-width: 200px;
  }
  .share-menu.open { display: flex; }
  .share-menu button {
    background: none;
    border: none;
    padding: 14px 20px;
    text-align: left;
    font-size: 0.9rem;
    font-weight: 500;
    color: #1A202C;
    cursor: pointer;
    display: flex;
    align-items: center;
    gap: 10px;
    font-family: 'Inter', sans-serif;
    transition: background 0.12s;
  }
  .share-menu button:hover { background: #F0F7FF; color: #1890FF; }
  .share-menu button + button { border-top: 1px solid #F0F4F8; }

  /* ── Impressão ── */
  @media print {
    [data-testid="stSidebar"],
    [data-testid="collapsedControl"],
    div[data-testid="stRadio"],
    .share-fab, .share-menu { display: none !important; }
    .block-container { padding: 1rem !important; max-width: 100% !important; }
  }
</style>
""", unsafe_allow_html=True)

# ── Tema escuro ───────────────────────────────────────────────────────────────
if st.session_state.dark_mode:
    st.markdown("""
<style>
  [data-testid="stAppViewContainer"], .main, [data-testid="stMain"] {
    background-color: #0E1117 !important;
  }
  [data-testid="stSidebar"] > div:first-child {
    background-color: #161B22 !important;
    border-right: 1px solid #30363D !important;
  }
  .kpi-card { background: #161B22 !important; border-color: #30363D !important; }
  .kpi-card .kpi-value, .kpi-card .kpi-label { color: #E6EDF3 !important; }
  p, span, div, label { color: #E6EDF3; }
  .sec-header { color: #8B949E !important; border-bottom-color: #30363D !important; }
  hr { border-color: #30363D !important; }
  [data-testid="stRadio"] > div > label:has(input:checked) {
    background: #161B22 !important;
    color: #1890FF !important;
    border-color: #30363D !important;
  }
</style>
""", unsafe_allow_html=True)

# ── Print header (injetado na sidebar via JS) ────────────────────────────────
import streamlit.components.v1 as _components
import base64 as _b64, pathlib as _pl
_logo_svg_b64 = _b64.b64encode(_pl.Path(__file__).parent.joinpath("assets/logo.svg").read_bytes()).decode()
_logo_data_url = f"data:image/svg+xml;base64,{_logo_svg_b64}"
_components.html("""
<script>
(function() {
  const p = window.parent;
  const pd = p.document;
  if (!pd.getElementById('print-header')) {
    const ph = pd.createElement('div');
    ph.id = 'print-header';
    ph.style.cssText = 'display:none;align-items:center;gap:16px;padding:16px 24px 12px;border-bottom:2px solid #1890FF;margin-bottom:16px;';
    ph.innerHTML = '<img src="__LOGO_DATA_URL__" style="height:36px;width:auto;" alt="Onfly"><span style="font-size:1.25rem;font-weight:700;color:#1A202C;font-family:Inter,sans-serif;">Sourcing A\u00e9reo</span>';
    pd.body.insertBefore(ph, pd.body.firstChild);
    const phStyle = pd.createElement('style');
    phStyle.textContent = '@page { margin: 32mm 12mm; } @media print { #print-header { display:flex!important; margin-bottom:48px; } [data-testid="stSidebar"],[data-testid="collapsedControl"],div[data-testid="stRadio"] { display:none!important; } * { -webkit-print-color-adjust:exact; color-adjust:exact; } *:focus, *:active, *::selection { background:transparent!important; outline:none!important; box-shadow:none!important; } }';
    pd.head.appendChild(phStyle);
  }
})();
</script>
""".replace("__LOGO_DATA_URL__", _logo_data_url), height=0)



# ─── Helpers ──────────────────────────────────────────────────────────────────
def brl(valor: float) -> str:
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def _gerar_excel_top15(df: pd.DataFrame, cia_nome: str, periodo: str) -> bytes:
    """Gera Excel formatado com top 15 clientes por GMV, pronto para e-mail."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    top15 = df.head(15).reset_index(drop=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Top 15 Clientes"

    BLUE_DARK  = "1D4ED8"
    BLUE_MID   = "1890FF"
    BLUE_LIGHT = "EFF6FF"
    WHITE      = "FFFFFF"
    GRAY_TEXT  = "374151"

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Linha 1: Título ──────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = f"Top 15 Clientes · {cia_nome} · {periodo}"
    title_cell.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=BLUE_DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Linha 2: Subtítulo Onfly ──────────────────────────────────────────────
    ws.merge_cells("A2:D2")
    sub_cell = ws["A2"]
    sub_cell.value = "Onfly · Sourcing Aéreo"
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color="93C5FD")
    sub_cell.fill = PatternFill("solid", fgColor=BLUE_DARK)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Linha 3: Cabeçalho ───────────────────────────────────────────────────
    headers = ["Nº", "Razão Social", "Nome Fantasia", "CNPJ"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE_MID)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[3].height = 22

    # ── Linhas de dados ──────────────────────────────────────────────────────
    for i, row in top15.iterrows():
        row_num = i + 4
        fill_color = BLUE_LIGHT if i % 2 == 0 else WHITE
        row_fill = PatternFill("solid", fgColor=fill_color)

        values = [
            i + 1,
            row["Razão Social"],
            row["Nome Fantasia"],
            row["CNPJ"],
        ]
        aligns = ["center", "left", "left", "left"]
        for col_idx, (val, align) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=10, color=GRAY_TEXT)
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = border
        ws.row_dimensions[row_num].height = 18

    # ── Linha de total ───────────────────────────────────────────────────────
    total_row = len(top15) + 4
    ws.merge_cells(f"A{total_row}:D{total_row}")
    total_label = ws[f"A{total_row}"]
    total_label.value = f"Total: {len(top15)} clientes"
    total_label.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    total_label.fill = PatternFill("solid", fgColor=BLUE_MID)
    total_label.alignment = Alignment(horizontal="center", vertical="center")
    total_label.border = border
    ws.row_dimensions[total_row].height = 20

    # ── Larguras de coluna ───────────────────────────────────────────────────
    col_widths = [5, 42, 32, 18]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
def _gerar_excel_voos(df: pd.DataFrame, filtros_desc: str, periodo: str) -> bytes:
    """Gera Excel formatado com todos os voos filtrados do bloco Consolidadores."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Voos Emitidos"

    BLUE_DARK  = "1D4ED8"
    BLUE_MID   = "1890FF"
    BLUE_LIGHT = "EFF6FF"
    WHITE      = "FFFFFF"
    GRAY_TEXT  = "374151"

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLS = ["Data Emissão", "Protocolo", "Localizador", "Consolidadora",
            "Cia", "Rota", "Data Voo", "Cliente", "Emissor", "GMV"]
    n_cols = len(COLS)
    last_col = get_column_letter(n_cols)

    # ── Linha 1: Título ──────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    tc = ws["A1"]
    tc.value = f"Voos Emitidos · {filtros_desc} · {periodo}"
    tc.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    tc.fill = PatternFill("solid", fgColor=BLUE_DARK)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Linha 2: Subtítulo ───────────────────────────────────────────────────
    ws.merge_cells(f"A2:{last_col}2")
    sc = ws["A2"]
    sc.value = f"Onfly · Sourcing Aéreo · {len(df)} registro{'s' if len(df) != 1 else ''}"
    sc.font = Font(name="Calibri", size=10, italic=True, color="93C5FD")
    sc.fill = PatternFill("solid", fgColor=BLUE_DARK)
    sc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Linha 3: Cabeçalho ───────────────────────────────────────────────────
    for ci, header in enumerate(COLS, 1):
        cell = ws.cell(row=3, column=ci, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE_MID)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[3].height = 22

    # ── Linhas de dados ──────────────────────────────────────────────────────
    aligns = ["center", "center", "center", "left",
              "left", "left", "center", "left", "left", "right"]
    for i, (_, row) in enumerate(df[COLS].iterrows()):
        row_num = i + 4
        fill_color = BLUE_LIGHT if i % 2 == 0 else WHITE
        row_fill = PatternFill("solid", fgColor=fill_color)
        for ci, (col, align) in enumerate(zip(COLS, aligns), 1):
            val = row[col]
            if col == "GMV" and isinstance(val, (int, float)):
                cell = ws.cell(row=row_num, column=ci, value=float(val))
                cell.number_format = 'R$ #,##0.00'
            else:
                cell = ws.cell(row=row_num, column=ci, value=str(val) if val is not None else "")
            cell.font = Font(name="Calibri", size=10, color=GRAY_TEXT)
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = border
        ws.row_dimensions[row_num].height = 18

    # ── Linha de total GMV ───────────────────────────────────────────────────
    total_row = len(df) + 4
    gmv_total = df["GMV"].sum() if "GMV" in df.columns else 0
    ws.merge_cells(f"A{total_row}:{get_column_letter(n_cols - 1)}{total_row}")
    label_cell = ws[f"A{total_row}"]
    label_cell.value = f"Total — {len(df)} registro{'s' if len(df) != 1 else ''}"
    label_cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    label_cell.fill = PatternFill("solid", fgColor=BLUE_MID)
    label_cell.alignment = Alignment(horizontal="right", vertical="center")
    label_cell.border = border

    gmv_cell = ws.cell(row=total_row, column=n_cols, value=float(gmv_total))
    gmv_cell.number_format = 'R$ #,##0.00'
    gmv_cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    gmv_cell.fill = PatternFill("solid", fgColor=BLUE_MID)
    gmv_cell.alignment = Alignment(horizontal="right", vertical="center")
    gmv_cell.border = border
    ws.row_dimensions[total_row].height = 20

    # ── Larguras de coluna ───────────────────────────────────────────────────
    col_widths = [14, 14, 14, 20, 28, 18, 14, 28, 24, 16]
    for ci, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _brl_df(df):
    """Retorna cópia do df com colunas de moeda formatadas em BRL (R$ 1.234,56)."""
    df = df.copy()
    for _col in ["GMV", "Ticket Médio", "Gross Revenue", "Incentivo (R$)", "Incentivo"]:
        if _col in df.columns:
            try:
                df[_col] = df[_col].apply(brl)
            except Exception:
                pass
    return df



def hex_to_rgba(hex_color: str, alpha: float) -> str:
    h = hex_color.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f"rgba({r},{g},{b},{alpha})"

def delta_html(atual: float, anterior: float) -> str:
    if not anterior:
        return ""
    pct = (atual - anterior) / anterior * 100
    if pct > 0:
        return f'<p class="kpi-delta up">▲ +{pct:.1f}% vs período anterior</p>'
    elif pct < 0:
        return f'<p class="kpi-delta down">▼ {pct:.1f}% vs período anterior</p>'
    return '<p class="kpi-delta neu">= sem variação</p>'

def plotly_layout(fig, height=340):
    fig.update_layout(
        height=height,
        margin=dict(l=8, r=8, t=16, b=8),
        paper_bgcolor="#E8ECF0",
        plot_bgcolor="#E8ECF0",
        font=dict(family="Inter, sans-serif", size=12, color="#4A5568"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    )
    fig.update_xaxes(gridcolor="#D4DAE2", zeroline=False, tickfont=dict(weight="bold"))
    fig.update_yaxes(gridcolor="#D4DAE2", zeroline=False, tickfont=dict(weight="bold"))
    return fig


# ─── BigQuery ─────────────────────────────────────────────────────────────────
@st.cache_resource
def bq_client():
    if "GCP_SERVICE_ACCOUNT" in st.secrets:
        import json
        from google.oauth2 import service_account
        info = json.loads(st.secrets["GCP_SERVICE_ACCOUNT"])
        creds = service_account.Credentials.from_service_account_info(
            info,
            scopes=["https://www.googleapis.com/auth/bigquery"]
        )
        return bigquery.Client(project=PROJECT_ID, credentials=creds)
    return bigquery.Client(project=PROJECT_ID)

def _where(inicio, fim, cia=None):
    filtro = f"""
        WHERE type = 'flight'
          AND status = 2

          AND created_at >= '{inicio}'
          AND created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
    """
    if cia:
        if isinstance(cia, (list, tuple)):
            vals = ", ".join(f"'{c}'" for c in cia)
            filtro += f"  AND consolidator_unified IN ({vals})\n"
        else:
            filtro += f"  AND consolidator_unified = '{cia}'\n"
    return filtro

@st.cache_data(ttl=300, show_spinner=False)
def q_resumo(inicio: str, fim: str, cia: str = None) -> dict:
    q = f"""
        SELECT
            COUNT(*)                                   AS qtd_reservas,
            ROUND(SUM(total_amount_currency_brl), 2)   AS gmv_total,
            ROUND(AVG(total_amount_currency_brl), 2)   AS ticket_medio
        FROM `{TABLE}`
        {_where(inicio, fim, cia)}
    """
    row = list(bq_client().query(q).result())[0]
    return {
        "qtd_reservas": int(row.qtd_reservas),
        "gmv_total":    float(row.gmv_total or 0),
        "ticket_medio": float(row.ticket_medio or 0),
    }

@st.cache_data(ttl=300, show_spinner=False)
def q_resumo_seg(inicio: str, fim: str, variantes: tuple) -> dict:
    """q_resumo filtrado por airline no segmento (para cias com consolidador misto)."""
    q = f"""
        SELECT
            COUNT(DISTINCT uuid)                       AS qtd_reservas,
            ROUND(SUM(total_amount_currency_brl), 2)   AS gmv_total,
            ROUND(SUM(total_amount_currency_brl) /
                  NULLIF(COUNT(DISTINCT uuid), 0), 2)  AS ticket_medio
        FROM `{TABLE}`
        WHERE type = 'flight' AND status = 2

          AND created_at >= '{inicio}'
          AND created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
          AND uuid IN (
              SELECT DISTINCT RTRIM(s.uuid, '_') FROM `{TABLE_SEG}` s
              WHERE s.segment = 0 AND s.step = 1 AND {_in_cias(variantes)}
          )
    """
    row = list(bq_client().query(q).result())[0]
    return {
        "qtd_reservas": int(row.qtd_reservas or 0),
        "gmv_total":    float(row.gmv_total or 0),
        "ticket_medio": float(row.ticket_medio or 0),
    }

@st.cache_data(ttl=300, show_spinner=False)
def q_diario_seg(inicio: str, fim: str, variantes: tuple) -> pd.DataFrame:
    """q_diario filtrado por airline no segmento."""
    q = f"""
        SELECT
            DATE(created_at)                           AS dia,
            COUNT(DISTINCT uuid)                       AS qtd_reservas,
            ROUND(SUM(total_amount_currency_brl), 2)   AS gmv_dia
        FROM `{TABLE}`
        WHERE type = 'flight' AND status = 2

          AND created_at >= '{inicio}'
          AND created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
          AND uuid IN (
              SELECT DISTINCT RTRIM(s.uuid, '_') FROM `{TABLE_SEG}` s
              WHERE s.segment = 0 AND s.step = 1 AND {_in_cias(variantes)}
          )
        GROUP BY dia ORDER BY dia
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Data": str(r.dia), "Reservas": int(r.qtd_reservas), "GMV": float(r.gmv_dia or 0)}
        for r in rows
    ])

@st.cache_data(ttl=3600, show_spinner=False)
def q_gmv_mensal_seg(variantes: tuple, ano: int = None) -> pd.DataFrame:
    """q_gmv_mensal filtrado por airline no segmento."""
    hoje = date.today()
    ano = ano or hoje.year
    mes_limite = hoje.month if ano == hoje.year else 12
    q = f"""
        SELECT
            EXTRACT(MONTH FROM created_at)             AS mes,
            ROUND(SUM(total_amount_currency_brl), 2)   AS gmv,
            COUNT(DISTINCT uuid)                       AS reservas
        FROM `{TABLE}`
        WHERE type = 'flight' AND status = 2

          AND EXTRACT(YEAR FROM created_at) = {ano}
          AND uuid IN (
              SELECT DISTINCT RTRIM(s.uuid, '_') FROM `{TABLE_SEG}` s
              WHERE s.segment = 0 AND s.step = 1 AND {_in_cias(variantes)}
          )
        GROUP BY mes ORDER BY mes
    """
    rows = list(bq_client().query(q).result())
    meses_com_dados = {int(r.mes): (float(r.gmv or 0), int(r.reservas or 0)) for r in rows}
    return pd.DataFrame([
        {"Mês": m, "GMV": meses_com_dados.get(m, (0, 0))[0],
         "Reservas": meses_com_dados.get(m, (0, 0))[1]}
        for m in range(1, mes_limite + 1)
    ])

@st.cache_data(ttl=300, show_spinner=False)
def q_por_cia(inicio: str, fim: str) -> pd.DataFrame:
    q = f"""
        SELECT
            consolidator_unified                       AS cia,
            COUNT(*)                                   AS qtd_reservas,
            ROUND(SUM(total_amount_currency_brl), 2)   AS gmv_total,
            ROUND(AVG(total_amount_currency_brl), 2)   AS ticket_medio
        FROM `{TABLE}`
        {_where(inicio, fim)}
        GROUP BY cia
        ORDER BY gmv_total DESC
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Cia": r.cia, "Reservas": int(r.qtd_reservas),
         "GMV": float(r.gmv_total or 0), "Ticket Médio": float(r.ticket_medio or 0)}
        for r in rows
    ])

@st.cache_data(ttl=300, show_spinner=False)
def q_segmentos(inicio: str, fim: str, cia=None) -> int:
    if isinstance(cia, (list, tuple)):
        vals = ", ".join(f"'{c}'" for c in cia)
        filtro_cia = f"AND e.consolidator_unified IN ({vals})"
    else:
        filtro_cia = f"AND e.consolidator_unified = '{cia}'" if cia else ""
    q = f"""
        SELECT COUNT(DISTINCT CONCAT(
            s.uuid, '-', s.departure_airport_code, '-', s.arrival_airport_code, '-',
            COALESCE(FORMAT_DATETIME('%H:%M', s.departure_date_hour), ''), '-',
            CAST(s.segment AS STRING)
        )) AS segmentos
        FROM `{TABLE_SEG}` s
        JOIN `{TABLE}` e ON e.uuid = RTRIM(s.uuid, '_')
        WHERE e.type = 'flight' AND e.status = 2
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
          {filtro_cia}
    """
    row = list(bq_client().query(q).result())[0]
    return int(row.segmentos or 0)

@st.cache_data(ttl=300, show_spinner=False)
def q_diario(inicio: str, fim: str, cia: str = None) -> pd.DataFrame:
    q = f"""
        SELECT
            DATE(created_at)                           AS dia,
            COUNT(*)                                   AS qtd_reservas,
            ROUND(SUM(total_amount_currency_brl), 2)   AS gmv_dia
        FROM `{TABLE}`
        {_where(inicio, fim, cia)}
        GROUP BY dia
        ORDER BY dia
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Data": str(r.dia), "Reservas": int(r.qtd_reservas), "GMV": float(r.gmv_dia or 0)}
        for r in rows
    ])

@st.cache_data(ttl=300, show_spinner=False)
def q_onhappy_diario(inicio: str, fim: str) -> pd.DataFrame:
    """GMV diário da ONHAPPY (company_id=4609) para sobreposição no gráfico."""
    q = f"""
        SELECT
            DATE(created_at)                         AS dia,
            ROUND(SUM(total_amount_currency_brl), 2) AS gmv_dia
        FROM `{TABLE}`
        WHERE type = 'flight' AND status = 2
          AND company_id = 4609
          AND created_at >= '{inicio}'
          AND created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        GROUP BY dia ORDER BY dia
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Data": str(r.dia), "GMV": float(r.gmv_dia or 0)}
        for r in rows
    ])

@st.cache_data(ttl=300, show_spinner=False)
def q_cias_amadeus(inicio: str, fim: str) -> pd.DataFrame:
    q = f"""
        SELECT
            COALESCE(NULLIF(TRIM(s.company_operator), ''), 'Não informado') AS Cia,
            COUNT(DISTINCT e.uuid)                     AS Reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2) AS GMV,
            ROUND(AVG(e.total_amount_currency_brl), 2) AS ticket_medio
        FROM `{TABLE_SEG}` s
        JOIN `{TABLE}` e ON e.uuid = RTRIM(s.uuid, '_')
        WHERE e.type = 'flight'
          AND e.status = 2
          AND e.consolidator_unified = 'AMADEUS'
          AND s.segment = 0 AND s.step = 1
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        GROUP BY 1
        ORDER BY GMV DESC
    """
    rows = list(bq_client().query(q).result())
    df = pd.DataFrame([
        {"Sigla": r.Cia, "Reservas": int(r.Reservas),
         "GMV": float(r.GMV or 0), "Ticket Médio": float(r.ticket_medio or 0)}
        for r in rows
    ])
    df.insert(1, "Companhia", df["Sigla"].map(lambda x: IATA_NOMES.get(x, "—")))
    return df

@st.cache_data(ttl=300, show_spinner=False)
def q_rotas(inicio: str, fim: str, cia=None) -> pd.DataFrame:
    if isinstance(cia, (list, tuple)):
        vals = ", ".join(f"'{c}'" for c in cia)
        filtro_cia = f"AND e.consolidator_unified IN ({vals})"
    else:
        filtro_cia = f"AND e.consolidator_unified = '{cia}'" if cia else ""
    q = f"""
        SELECT
            s.departure_airport_code      AS Origem,
            s.arrival_airport_code AS Destino,
            COUNT(DISTINCT s.uuid)                     AS Reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2) AS GMV
        FROM `{TABLE_SEG}` s
        JOIN `{TABLE}` e ON e.uuid = RTRIM(s.uuid, '_')
        WHERE e.type = 'flight'
          AND e.status = 2
          AND s.segment = 0 AND s.step = 1
          AND s.departure_airport_code != s.arrival_airport_code
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
          {filtro_cia}
        GROUP BY 1, 2
        ORDER BY GMV DESC
        LIMIT 10
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Origem": r.Origem, "Destino": r.Destino,
         "Reservas": int(r.Reservas), "GMV": float(r.GMV or 0)}
        for r in rows
    ])

@st.cache_data(ttl=3600, show_spinner=False)
def q_gmv_anual(cia=None) -> pd.DataFrame:
    if isinstance(cia, (list, tuple)):
        vals = ", ".join(f"'{c}'" for c in cia)
        filtro_cia = f"AND consolidator_unified IN ({vals})"
    elif cia:
        filtro_cia = f"AND consolidator_unified = '{cia}'"
    else:
        filtro_cia = ""
    q = f"""
        SELECT
            CAST(EXTRACT(YEAR FROM created_at) AS INT64) AS ano,
            ROUND(SUM(total_amount_currency_brl), 2)     AS gmv,
            COUNT(*)                                     AS reservas
        FROM `{TABLE}`
        WHERE type = 'flight' AND status = 2

          {filtro_cia}
        GROUP BY ano ORDER BY ano
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Ano": int(r.ano), "GMV": float(r.gmv or 0), "Reservas": int(r.reservas or 0)}
        for r in rows
    ])

@st.cache_data(ttl=3600, show_spinner=False)
def q_gmv_mensal(cia=None, ano: int = None) -> pd.DataFrame:
    hoje = date.today()
    ano = ano or hoje.year
    if isinstance(cia, (list, tuple)):
        vals = ", ".join(f"'{c}'" for c in cia)
        filtro_cia = f"AND consolidator_unified IN ({vals})"
    elif cia:
        filtro_cia = f"AND consolidator_unified = '{cia}'"
    else:
        filtro_cia = ""
    q = f"""
        SELECT
            EXTRACT(MONTH FROM created_at)               AS mes,
            ROUND(SUM(total_amount_currency_brl), 2)     AS gmv,
            COUNT(*)                                     AS reservas
        FROM `{TABLE}`
        WHERE type = 'flight' AND status = 2

          AND EXTRACT(YEAR FROM created_at) = {ano}
          {filtro_cia}
        GROUP BY mes ORDER BY mes
    """
    rows = list(bq_client().query(q).result())
    meses_com_dados = {int(r.mes): (float(r.gmv or 0), int(r.reservas or 0)) for r in rows}
    # Se é o ano atual, mostra só até o mês corrente; senão, mostra todos os 12 meses
    mes_limite = hoje.month if ano == hoje.year else 12
    return pd.DataFrame([
        {"Mês": m, "GMV": meses_com_dados.get(m, (0, 0))[0],
         "Reservas": meses_com_dados.get(m, (0, 0))[1]}
        for m in range(1, mes_limite + 1)
    ])

@st.cache_data(ttl=3600, show_spinner=False)
def q_gmv_mensal_ano(ano: int) -> dict:
    """Retorna {mes: gmv} para todos os 12 meses de um ano."""
    q = f"""
        SELECT
            CAST(EXTRACT(MONTH FROM created_at) AS INT64) AS mes,
            ROUND(SUM(total_amount_currency_brl), 2)      AS gmv
        FROM `{TABLE}`
        WHERE type = 'flight' AND status = 2

          AND EXTRACT(YEAR FROM created_at) = {ano}
        GROUP BY mes
    """
    rows = list(bq_client().query(q).result())
    return {int(r.mes): float(r.gmv or 0) for r in rows}

@st.cache_data(ttl=300, show_spinner=False)
def q_cabine(inicio: str, fim: str, cia) -> pd.DataFrame:
    if isinstance(cia, (list, tuple)):
        vals = ", ".join(f"'{c}'" for c in cia)
        filtro_cia = f"AND e.consolidator_unified IN ({vals})"
    else:
        filtro_cia = f"AND e.consolidator_unified = '{cia}'"
    q = f"""
        SELECT
            COALESCE(NULLIF(TRIM(f.fare_family), ''), 'Não informado') AS cabine,
            COUNT(*)                                    AS qtd_reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)  AS gmv
        FROM `{TABLE}` e
        JOIN `{TABLE_FARE}` f
          ON CONCAT(SPLIT(f.uuid, '_flight_')[OFFSET(0)], '_flight') = e.uuid
        WHERE e.type = 'flight'
          AND e.status = 2
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
          {filtro_cia}
          AND f.type = 'flight'
          AND f.is_origin = 1
        GROUP BY cabine
        ORDER BY gmv DESC
        LIMIT 15
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Cabine": r.cabine, "Reservas": int(r.qtd_reservas), "GMV": float(r.gmv or 0)}
        for r in rows
    ])


# ─── Queries: Clientes ────────────────────────────────────────────────────────
@st.cache_data(ttl=300, show_spinner=False)
@st.cache_data(ttl=3600, show_spinner=False)
def q_cliente_info(company_id: int) -> dict:
    q = f"""
        SELECT company_id, name, social_name, cnpj
        FROM `{TABLE_COMP}`
        WHERE company_id = {company_id}
        LIMIT 1
    """
    try:
        rows = list(bq_client().query(q).result())
        if rows:
            r = rows[0]
            return {
                "company_id":  int(r.company_id),
                "name":        r.name or "",
                "social_name": r.social_name or "",
                "cnpj":        r.cnpj or "",
            }
    except Exception:
        pass
    return {}


def q_clientes_ranking(inicio: str, fim: str) -> pd.DataFrame:
    q = f"""
        SELECT
            c.company_id                                      AS company_id,
            c.name                                            AS cliente,
            COUNT(DISTINCT e.uuid)                            AS qtd_reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)        AS gmv_total,
            ROUND(AVG(e.total_amount_currency_brl), 2)        AS ticket_medio
        FROM `{TABLE}` e
        JOIN `{TABLE_COMP}` c ON e.company_id = c.company_id
        WHERE e.type = 'flight'
          AND e.status = 2
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        GROUP BY c.company_id, c.name
        ORDER BY gmv_total DESC
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"company_id": int(r.company_id), "Cliente": r.cliente,
         "Reservas": int(r.qtd_reservas), "GMV": float(r.gmv_total or 0),
         "Ticket Médio": float(r.ticket_medio or 0)}
        for r in rows
    ])

@st.cache_data(ttl=300, show_spinner=False)
def q_cliente_diario(company_id: int, inicio: str, fim: str) -> pd.DataFrame:
    q = f"""
        SELECT
            DATE(e.created_at)                               AS dia,
            COUNT(DISTINCT e.uuid)                           AS qtd_reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)       AS gmv
        FROM `{TABLE}` e
        WHERE e.type = 'flight'
          AND e.status = 2
          AND e.company_id = {company_id}
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        GROUP BY dia
        ORDER BY dia
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Data": str(r.dia), "Reservas": int(r.qtd_reservas), "GMV": float(r.gmv or 0)}
        for r in rows
    ])

@st.cache_data(ttl=300, show_spinner=False)
def q_cliente_rotas(company_id: int, inicio: str, fim: str) -> pd.DataFrame:
    q = f"""
        SELECT
            s.departure_airport_code                                         AS origem,
            s.arrival_airport_code                                    AS destino,
            COUNT(DISTINCT e.uuid)                           AS qtd_reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)       AS gmv
        FROM `{TABLE}` e
        JOIN `{TABLE_SEG}` s ON RTRIM(s.uuid, '_') = e.uuid
        WHERE e.type = 'flight'
          AND e.status = 2
          AND e.company_id = {company_id}
          AND s.segment = 0 AND s.step = 1
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        GROUP BY origem, destino
        ORDER BY gmv DESC
        LIMIT 15
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Origem": r.origem, "Destino": r.destino,
         "Reservas": int(r.qtd_reservas), "GMV": float(r.gmv or 0)}
        for r in rows
    ])

@st.cache_data(ttl=300, show_spinner=False)
def q_cliente_cias(company_id: int, inicio: str, fim: str) -> pd.DataFrame:
    q = f"""
        SELECT
            e.consolidator_unified                           AS cia,
            COUNT(DISTINCT e.uuid)                           AS qtd_reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)       AS gmv
        FROM `{TABLE}` e
        WHERE e.type = 'flight'
          AND e.status = 2
          AND e.company_id = {company_id}
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        GROUP BY cia
        ORDER BY gmv DESC
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Cia": r.cia, "Reservas": int(r.qtd_reservas), "GMV": float(r.gmv or 0)}
        for r in rows
    ])


# ─── Queries: Consolidadores ──────────────────────────────────────────────────
@st.cache_data(ttl=300, show_spinner=False)
def q_clientes_emissoes(inicio: str, fim: str) -> pd.DataFrame:
    """Lista todas as emissões do período — 1 linha por uuid, com GROUP BY para respeitar row-level security."""
    q = f"""
        SELECT
            FORMAT_DATETIME('%d/%m/%Y %H:%M', MAX(e.created_at)) AS data_emissao,
            e.uuid                                                 AS uuid,
            MAX(e.id)                                              AS protocolo,
            COALESCE(MIN(c.name), 'Não identificado')             AS cliente,
            TRIM(COALESCE(MAX(e.emitter_name), '—'))              AS emissor,
            COALESCE(MAX(e.consolidator_unified), 'N/I')          AS consolidadora,
            COALESCE(MAX(e.flight_locator_outbound), '—')         AS localizador,
            ROUND(MAX(e.total_amount_currency_brl), 2)            AS gmv
        FROM `{TABLE}` e
        LEFT JOIN `{TABLE_COMP}` c ON c.company_id = e.company_id
        WHERE e.type = 'flight'
          AND e.status = 2
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        GROUP BY e.uuid
        ORDER BY MAX(e.created_at) DESC
        LIMIT 5000
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {
            "Data Emissão": r.data_emissao or "—",
            "Protocolo":    r.protocolo or "—",
            "Cliente":      r.cliente,
            "Emissor":      r.emissor.title() if r.emissor and r.emissor != "—" else "—",
            "Consolidadora":r.consolidadora,
            "Localizador":  r.localizador or "—",
            "GMV":          float(r.gmv or 0),
        }
        for r in rows
    ])


@st.cache_data(ttl=300, show_spinner=False)
def q_consolidadores_lista(inicio: str, fim: str) -> pd.DataFrame:
    """Ranking de consolidadores com GMV, reservas e ticket médio."""
    q = f"""
        SELECT
            COALESCE(e.consolidator_unified, 'Não identificado')  AS consolidador,
            COUNT(DISTINCT e.uuid)                                  AS qtd_reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)             AS gmv_total,
            ROUND(SUM(e.total_amount_currency_brl) /
                  NULLIF(COUNT(DISTINCT e.uuid), 0), 2)            AS ticket_medio
        FROM `{TABLE}` e
        WHERE e.type = 'flight'
          AND e.status = 2
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        GROUP BY consolidador
        ORDER BY gmv_total DESC
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Consolidador": r.consolidador, "Reservas": int(r.qtd_reservas),
         "GMV": float(r.gmv_total or 0), "Ticket Médio": float(r.ticket_medio or 0)}
        for r in rows
    ])

@st.cache_data(ttl=300, show_spinner=False)
def q_consolidador_cias(consolidador, inicio: str, fim: str) -> pd.DataFrame:
    """Cias aéreas emitidas por um consolidador ou lista deles (None = todos)."""
    if isinstance(consolidador, (list, tuple)):
        _vals = ", ".join(f"'{c}'" for c in consolidador)
        _filtro = f"AND COALESCE(e.consolidator_unified, 'Não identificado') IN ({_vals})"
    elif consolidador:
        _filtro = f"AND COALESCE(e.consolidator_unified, 'Não identificado') = '{consolidador}'"
    else:
        _filtro = ""
    q = f"""
        SELECT
            e.consolidator_unified                                 AS cia,
            COUNT(DISTINCT e.uuid)                                 AS qtd_reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)             AS gmv,
            ROUND(SUM(e.total_amount_currency_brl) /
                  NULLIF(COUNT(DISTINCT e.uuid), 0), 2)            AS ticket_medio
        FROM `{TABLE}` e
        WHERE e.type = 'flight'
          AND e.status = 2
          {_filtro}
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        GROUP BY cia
        ORDER BY gmv DESC
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Cia": r.cia, "Reservas": int(r.qtd_reservas),
         "GMV": float(r.gmv or 0), "Ticket Médio": float(r.ticket_medio or 0)}
        for r in rows
    ])

@st.cache_data(ttl=300, show_spinner=False)
def q_consolidador_clientes(consolidador, inicio: str, fim: str) -> pd.DataFrame:
    """Clientes que emitiram via um consolidador ou lista deles (None = todos)."""
    if isinstance(consolidador, (list, tuple)):
        _vals = ", ".join(f"'{c}'" for c in consolidador)
        _filtro = f"AND COALESCE(e.consolidator_unified, 'Não identificado') IN ({_vals})"
    elif consolidador:
        _filtro = f"AND COALESCE(e.consolidator_unified, 'Não identificado') = '{consolidador}'"
    else:
        _filtro = ""
    q = f"""
        SELECT
            COALESCE(c.name, 'Não identificado')                   AS cliente,
            COUNT(DISTINCT e.uuid)                                 AS qtd_reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)             AS gmv
        FROM `{TABLE}` e
        LEFT JOIN `{TABLE_COMP}` c ON c.company_id = e.company_id
        WHERE e.type = 'flight'
          AND e.status = 2
          {_filtro}
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        GROUP BY cliente
        ORDER BY gmv DESC
        LIMIT 30
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Cliente": r.cliente, "Reservas": int(r.qtd_reservas), "GMV": float(r.gmv or 0)}
        for r in rows
    ])

@st.cache_data(ttl=300, show_spinner=False)
def q_consolidador_rotas(consolidador, inicio: str, fim: str) -> pd.DataFrame:
    """Top rotas emitidas via um consolidador ou lista deles (None = todos)."""
    if isinstance(consolidador, (list, tuple)):
        _vals = ", ".join(f"'{c}'" for c in consolidador)
        _filtro = f"AND COALESCE(e.consolidator_unified, 'Não identificado') IN ({_vals})"
    elif consolidador:
        _filtro = f"AND COALESCE(e.consolidator_unified, 'Não identificado') = '{consolidador}'"
    else:
        _filtro = ""
    q = f"""
        SELECT
            s.departure_airport_code                                               AS origem,
            s.arrival_airport_code                                          AS destino,
            COUNT(DISTINCT e.uuid)                                 AS qtd_reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)             AS gmv
        FROM `{TABLE}` e
        JOIN `{TABLE_SEG}` s ON RTRIM(s.uuid, '_') = e.uuid
        WHERE e.type = 'flight'
          AND e.status = 2
          AND s.segment = 0 AND s.step = 1
          {_filtro}
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        GROUP BY origem, destino
        ORDER BY gmv DESC
        LIMIT 20
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Rota": f"{r.origem} → {r.destino}", "Reservas": int(r.qtd_reservas), "GMV": float(r.gmv or 0)}
        for r in rows
    ])

@st.cache_data(ttl=300, show_spinner=False)
def q_consolidador_voos(consolidador, inicio: str, fim: str) -> pd.DataFrame:
    """Lista de emissões — 1 linha por uuid. consolidador pode ser str, lista ou None (todos)."""
    if isinstance(consolidador, (list, tuple)):
        _vals = ", ".join(f"'{c}'" for c in consolidador)
        _filtro = f"AND COALESCE(e.consolidator_unified, 'Não identificado') IN ({_vals})"
    elif consolidador:
        _filtro = f"AND COALESCE(e.consolidator_unified, 'Não identificado') = '{consolidador}'"
    else:
        _filtro = ""
    q = f"""
        SELECT
            FORMAT_DATETIME('%d/%m/%Y %H:%M', MAX(e.created_at)) AS data_emissao,
            MAX(e.id)                                               AS protocolo,
            MAX(e.flight_locator_outbound)                          AS localizador,
            MAX(e.consolidator_unified)                             AS consolidador,
            MIN(s.company_operator)                                          AS cia,
            CONCAT(MIN(s.departure_airport_code), ' → ', MIN(s.arrival_airport_code))       AS rota,
            CAST(MIN(DATE(s.departure_date_hour)) AS STRING)          AS data_voo,
            COALESCE(MIN(c.name), 'Não identificado')               AS cliente,
            MAX(e.emitter_name)                                     AS emissor,
            ROUND(MAX(e.total_amount_currency_brl), 2)              AS gmv
        FROM `{TABLE}` e
        JOIN `{TABLE_SEG}` s ON RTRIM(s.uuid, '_') = e.uuid AND s.segment = 0 AND s.step = 1
        LEFT JOIN `{TABLE_COMP}` c ON c.company_id = e.company_id
        WHERE e.type = 'flight'
          AND e.status = 2
          {_filtro}
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        GROUP BY e.uuid
        ORDER BY MAX(e.created_at) DESC
        LIMIT 1500
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {
            "Data Emissão":  r.data_emissao or "—",
            "Protocolo":     r.protocolo or "—",
            "Localizador":   r.localizador or "—",
            "Consolidadora": r.consolidador or "—",
            "Cia":           (f"{r.cia} — {IATA_NOMES[r.cia]}" if r.cia and r.cia in IATA_NOMES else r.cia or "—"),
            "Rota":          r.rota,
            "Data Voo":      r.data_voo or "—",
            "Cliente":       r.cliente,
            "Emissor":       r.emissor or "—",
            "GMV":           float(r.gmv or 0),
        }
        for r in rows
    ])

@st.cache_data(ttl=60, show_spinner=False)
def q_busca_protocolo(protocolo: str) -> pd.DataFrame:
    """Busca direta por protocolo (e.id) na base inteira, sem filtro de data."""
    _prot = protocolo.strip().upper()
    q = f"""
        SELECT
            FORMAT_DATETIME('%d/%m/%Y %H:%M', MAX(e.created_at)) AS data_emissao,
            MAX(e.id)                                               AS protocolo,
            MAX(e.flight_locator_outbound)                          AS localizador,
            MAX(e.consolidator_unified)                             AS consolidador,
            MIN(s.company_operator)                                          AS cia,
            CONCAT(MIN(s.departure_airport_code), ' → ', MIN(s.arrival_airport_code))       AS rota,
            CAST(MIN(DATE(s.departure_date_hour)) AS STRING)          AS data_voo,
            COALESCE(MIN(c.name), 'Não identificado')               AS cliente,
            MAX(e.emitter_name)                                     AS emissor,
            ROUND(MAX(e.total_amount_currency_brl), 2)              AS gmv
        FROM `{TABLE}` e
        JOIN `{TABLE_SEG}` s ON RTRIM(s.uuid, '_') = e.uuid AND s.segment = 0 AND s.step = 1
        LEFT JOIN `{TABLE_COMP}` c ON c.company_id = e.company_id
        WHERE e.type = 'flight'
          AND e.status = 2
          AND UPPER(e.id) LIKE '%{_prot}%'
        GROUP BY e.uuid
        ORDER BY MAX(e.created_at) DESC
        LIMIT 100
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {
            "Data Emissão":  str(r.data_emissao),
            "Protocolo":     r.protocolo or "—",
            "Localizador":   r.localizador or "—",
            "Consolidadora": r.consolidador or "—",
            "Cia":           (f"{r.cia} — {IATA_NOMES[r.cia]}" if r.cia and r.cia in IATA_NOMES else r.cia or "—"),
            "Rota":          r.rota,
            "Data Voo":      r.data_voo or "—",
            "Cliente":       r.cliente,
            "Emissor":       r.emissor or "—",
            "GMV":           float(r.gmv or 0),
        }
        for r in rows
    ])


# ─── Queries: Aeroportos ──────────────────────────────────────────────────────
@st.cache_data(ttl=300, show_spinner=False)
def q_aeroportos_ranking(inicio: str, fim: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Retorna (df_destino, df_origem) com ranking por GMV e qtd bilhetes."""
    q = f"""
        SELECT
            tipo,
            aeroporto,
            COUNT(DISTINCT uuid)                     AS bilhetes,
            ROUND(SUM(total_amount_currency_brl), 2) AS gmv
        FROM (
            SELECT 'destino' AS tipo, UPPER(TRIM(s.arrival_airport_code)) AS aeroporto,
                   e.uuid, e.total_amount_currency_brl
            FROM `{TABLE}` e
            JOIN `{TABLE_SEG}` s ON RTRIM(s.uuid, '_') = e.uuid AND s.segment = 0 AND s.step = 1
            WHERE e.type = 'flight' AND e.status = 2
              AND e.created_at >= '{inicio}'
              AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
            UNION ALL
            SELECT 'origem' AS tipo, UPPER(TRIM(s.departure_airport_code)) AS aeroporto,
                   e.uuid, e.total_amount_currency_brl
            FROM `{TABLE}` e
            JOIN `{TABLE_SEG}` s ON RTRIM(s.uuid, '_') = e.uuid AND s.segment = 0 AND s.step = 1
            WHERE e.type = 'flight' AND e.status = 2
              AND e.created_at >= '{inicio}'
              AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        )
        WHERE aeroporto IS NOT NULL AND LENGTH(aeroporto) = 3
        GROUP BY tipo, aeroporto
        ORDER BY tipo, gmv DESC
    """
    rows = list(bq_client().query(q).result())

    def _build(tipo_filtro):
        data = []
        rank = 1
        for r in rows:
            if r.tipo != tipo_filtro:
                continue
            data.append({
                "Nº":        rank,
                "IATA":      r.aeroporto,
                "Aeroporto": AEROPORTO_NOMES.get(r.aeroporto, r.aeroporto),
                "Bilhetes":  int(r.bilhetes or 0),
                "GMV":       float(r.gmv or 0),
            })
            rank += 1
        return pd.DataFrame(data)

    return _build("destino"), _build("origem")

# ─── Queries: Consolidador Evolução Mensal ────────────────────────────────────
@st.cache_data(ttl=300, show_spinner=False)
def q_consolidador_evolucao_mensal(consolidadores, inicio: str, fim: str) -> pd.DataFrame:
    """GMV mensal para os consolidadores selecionados (None ou [] = todos)."""
    if isinstance(consolidadores, (list, tuple)) and len(consolidadores) > 0:
        _vals = ", ".join(f"'{c}'" for c in consolidadores)
        _filtro = f"AND COALESCE(e.consolidator_unified, 'Não identificado') IN ({_vals})"
    else:
        _filtro = ""
    q = f"""
        SELECT
            FORMAT_DATE('%Y-%m', DATE(e.created_at)) AS mes,
            ROUND(SUM(e.total_amount_currency_brl), 2) AS gmv
        FROM `{TABLE}` e
        WHERE e.type = 'flight'
          AND e.status = 2
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
          {_filtro}
        GROUP BY mes
        ORDER BY mes
    """
    rows = list(bq_client().query(q).result())
    MESES_PT = {"01":"Jan","02":"Fev","03":"Mar","04":"Abr","05":"Mai","06":"Jun",
                "07":"Jul","08":"Ago","09":"Set","10":"Out","11":"Nov","12":"Dez"}
    data = []
    for r in rows:
        ano, mes_num = r.mes.split("-")
        data.append({
            "mes_key":  r.mes,
            "Mês":      f"{MESES_PT[mes_num]}/{ano}",
            "GMV":      float(r.gmv or 0),
        })
    df = pd.DataFrame(data)
    if not df.empty:
        df["GMV Acumulado"] = df["GMV"].cumsum()
    return df

# ─── Queries: Quem voa o que? ─────────────────────────────────────────────────
@st.cache_data(ttl=300, show_spinner=False)
def q_quem_voa_o_que(variantes: tuple, inicio: str, fim: str) -> pd.DataFrame:
    """Clientes que voaram em uma cia aérea no período, com GMV, trechos e rota top.
    Filtra por silver_flight_orders.standard_airline — cobre cias internacionais sem TABLE_SEG."""
    _sfo_in = ", ".join(f"'{v}'" for v in variantes)
    q = f"""
        WITH protos_cia AS (
            SELECT DISTINCT protocol
            FROM `{TABLE_FLIGHT_ORDERS}`
            WHERE status = 2
              AND UPPER(TRIM(standard_airline)) IN ({_sfo_in})
        ),
        uuids_cia AS (
            SELECT e.uuid, e.company_id, e.total_amount_currency_brl AS gmv
            FROM `{TABLE}` e
            WHERE e.type = 'flight'
              AND e.status = 2
              AND e.created_at >= '{inicio}'
              AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
              AND REPLACE(e.uuid, '_flight', '') IN (SELECT protocol FROM protos_cia)
        ),
        gmv_empresa AS (
            SELECT company_id, ROUND(SUM(gmv), 2) AS gmv
            FROM uuids_cia
            GROUP BY company_id
        ),
        segmentos AS (
            SELECT u.company_id, CONCAT(s.departure_airport_code, ' → ', s.arrival_airport_code) AS rota
            FROM uuids_cia u
            JOIN `{TABLE_SEG}` s ON RTRIM(s.uuid, '_') = u.uuid
            WHERE s.segment = 0 AND s.step = 1
        ),
        trechos_empresa AS (
            SELECT company_id, COUNT(*) AS trechos
            FROM segmentos
            GROUP BY company_id
        ),
        rotas_freq AS (
            SELECT company_id, rota, COUNT(*) AS freq
            FROM segmentos
            GROUP BY company_id, rota
        ),
        top_rota AS (
            SELECT company_id,
                   ARRAY_AGG(rota ORDER BY freq DESC LIMIT 1)[OFFSET(0)] AS trecho_top
            FROM rotas_freq
            GROUP BY company_id
        )
        SELECT
            COALESCE(NULLIF(TRIM(c.social_name), ''), c.name) AS razao_social,
            c.name          AS nome_fantasia,
            c.cnpj          AS cnpj,
            g.gmv           AS gmv,
            COALESCE(t.trechos, 0)          AS trechos,
            COALESCE(r.trecho_top, '—')     AS trecho_top
        FROM gmv_empresa g
        JOIN `{TABLE_COMP}` c ON c.company_id = g.company_id
        LEFT JOIN trechos_empresa t ON t.company_id = g.company_id
        LEFT JOIN top_rota r ON r.company_id = g.company_id
        ORDER BY g.gmv DESC
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {
            "Razão Social":        r.razao_social or "—",
            "Nome Fantasia":       r.nome_fantasia or "—",
            "CNPJ":                r.cnpj or "—",
            "GMV":                 float(r.gmv or 0),
            "Trechos":             int(r.trechos or 0),
            "Trecho Mais Voado":   r.trecho_top or "—",
        }
        for r in rows
    ])

# ─── Queries: Pricing ─────────────────────────────────────────────────────────
@st.cache_data(ttl=300, show_spinner=False)
def q_pricing(origem: str, destino: str, inicio: str, fim: str) -> pd.DataFrame:
    q = f"""
        SELECT
            DATE(e.created_at)                               AS data_emissao,
            DATE(s.departure_date_hour)                      AS flight_date,
            UPPER(TRIM(s.company_operator))                  AS cia_raw,
            e.total_amount_currency_brl                      AS preco,
            FORMAT_DATETIME('%H:%M', s.departure_date_hour) AS departure_time,
            FORMAT_DATETIME('%H:%M', s.arrival_date_hour)   AS arrival_time,
            CAST(s.segment + 1 AS INT64)                    AS trecho
        FROM `{TABLE}` e
        JOIN `{TABLE_SEG}` s ON RTRIM(s.uuid, '_') = e.uuid
        WHERE e.type = 'flight'
          AND e.status = 2
          AND s.segment = 0 AND s.step = 1
          AND UPPER(TRIM(s.departure_airport_code)) = '{origem.upper().strip()}'
          AND UPPER(TRIM(s.arrival_airport_code))   = '{destino.upper().strip()}'
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        ORDER BY data_emissao
    """
    rows = list(bq_client().query(q).result())
    df = pd.DataFrame([
        {
            "Data Emissão": str(r.data_emissao),
            "Data Voo":     str(r.flight_date),
            "Cia":          IATA_NOMES.get(r.cia_raw, r.cia_raw),
            "Preço (R$)":   float(r.preco or 0),
            "Saída":        str(r.departure_time or ""),
            "Chegada":      str(r.arrival_time or ""),
            "Trecho":       "Ida" if int(r.trecho or 1) == 1 else "Volta",
        }
        for r in rows
    ])
    return df


# ─── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    # Logo
    svg_path = os.path.join(os.path.dirname(__file__), "assets", "logo.svg")
    png_path = os.path.join(os.path.dirname(__file__), "assets", "logo.png")
    if os.path.exists(svg_path):
        svg_content = open(svg_path).read()
        svg_content = re.sub(r'\bwidth="[^"]*"',  'width="160"', svg_content, count=1)
        svg_content = re.sub(r'\bheight="[^"]*"', 'height="47"', svg_content, count=1)
        st.markdown(f'<div style="padding:16px 0 8px 0;">{svg_content}</div>', unsafe_allow_html=True)
    elif os.path.exists(png_path):
        st.image(png_path, width=140)
    else:
        st.markdown("""
            <div style="padding:18px 0 4px 0;">
                <span style="font-size:1.7rem;font-weight:900;letter-spacing:-0.04em;color:#1890FF;">on</span>
                <span style="font-size:1.7rem;font-weight:900;letter-spacing:-0.04em;color:#1890FF;">fly</span>
                <span style="font-size:1.1rem;">🌐</span>
            </div>
        """, unsafe_allow_html=True)

    st.markdown("<hr style='border-color:#E8ECF0;margin:16px 0;'>", unsafe_allow_html=True)

    # ── Categoria ──
    GRUPOS = {
        "📊  Resultados": ["📊  GMV", "🎫  GMV de Incentivo", "💰  Take Rate", "💎  Take Rate c/ Incentivo", "📈  Tendência"],
        "✈️  Visitas":    ["✈️  Cia Aérea", "🌍  Destino", "🗺️  Distribuição", "🌐  Potencial de Voo", "🏭  Consolidadores", "📋  Cia Aérea Legado", "⚖️  Balanceamento", "🔍  Buscas & Conversão"],
        "👥  Clientes":   ["🏢  Clientes", "💲  Pricing", "🛫  Quem voa o que?"],
        "💼  Incentivos": ["🔵  Azul", "🔴  LATAM"],
        "🔍  Análises":   [],
        "📝  Anotações":  [],
        "🤝  CRM Aéreo":  [],
    }
    st.markdown("<p style='font-size:0.65rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#8C9BAB;margin-bottom:6px;'>Grupo</p>", unsafe_allow_html=True)
    grupo = st.selectbox("", list(GRUPOS.keys()), index=None, placeholder="Selecione...", label_visibility="collapsed", key="sel_grupo")
    if grupo is not None:
        itens = GRUPOS[grupo]
        if itens:
            st.markdown("<p style='font-size:0.65rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#8C9BAB;margin:10px 0 6px 0;'>Categoria</p>", unsafe_allow_html=True)
            secao = st.selectbox("", itens, label_visibility="collapsed", key="sel_secao")
        else:
            secao = grupo  # Análises é standalone
    else:
        secao = None
    st.markdown("<hr style='border-color:#E8ECF0;margin:16px 0;'>", unsafe_allow_html=True)

    # ── Período ──
    st.markdown("<p style='font-size:0.65rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#8C9BAB;margin-bottom:6px;'>Período</p>", unsafe_allow_html=True)
    OPCOES = ["Esta semana", "Este mês", "Mês anterior", "Trimestre", "Semestre", "Este ano", "Ano anterior", "Desde o início", "Personalizado"]
    periodo = st.selectbox("", OPCOES, index=1, label_visibility="collapsed")

    hoje = date.today()
    if periodo == "Esta semana":
        inicio = hoje - timedelta(days=hoje.weekday())
        fim = hoje
    elif periodo == "Este mês":
        inicio = hoje.replace(day=1)
        fim = hoje
    elif periodo == "Mês anterior":
        fim = hoje.replace(day=1) - timedelta(days=1)
        inicio = fim.replace(day=1)
    elif periodo == "Trimestre":
        mes_ini = ((hoje.month - 1) // 3) * 3 + 1
        inicio = hoje.replace(month=mes_ini, day=1)
        fim = hoje
    elif periodo == "Semestre":
        mes_ini = 1 if hoje.month <= 6 else 7
        inicio = hoje.replace(month=mes_ini, day=1)
        fim = hoje
    elif periodo == "Este ano":
        inicio = hoje.replace(month=1, day=1)
        fim = hoje
    elif periodo == "Ano anterior":
        inicio = hoje.replace(year=hoje.year - 1, month=1, day=1)
        fim    = hoje.replace(year=hoje.year - 1, month=12, day=31)
    elif periodo == "Desde o início":
        inicio = date(2020, 1, 1)
        fim    = hoje
    else:
        col_a, col_b = st.columns(2)
        with col_a:
            inicio = st.date_input("De", hoje.replace(day=1))
        with col_b:
            fim = st.date_input("Até", hoje)

    if periodo != "Personalizado":
        st.markdown(f"""
            <div class="period-badge">
                <div class="period-col">
                    <p class="label">De</p>
                    <p class="value">{inicio.strftime("%d/%m/%Y")}</p>
                </div>
                <div class="period-col">
                    <p class="label">Até</p>
                    <p class="value">{fim.strftime("%d/%m/%Y")}</p>
                </div>
            </div>
        """, unsafe_allow_html=True)

    st.markdown("<hr style='border-color:#E8ECF0;margin:20px 0 10px 0;'>", unsafe_allow_html=True)

    # ── Comparativo ──
    st.markdown("<p style='font-size:0.65rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#8C9BAB;margin-bottom:6px;'>Comparativo</p>", unsafe_allow_html=True)
    comp_tipo = st.selectbox(
        "",
        ["Nenhum", "WoW — Semana anterior", "MoM — Mês anterior", "YoY — Ano anterior"],
        index=0,
        label_visibility="collapsed",
    )
    comp_tipo = comp_tipo.split(" ")[0]  # extrai só "Nenhum", "WoW", "MoM" ou "YoY"

    # Calcula datas do período de comparação
    if comp_tipo == "WoW":
        c_inicio = inicio - timedelta(weeks=1)
        c_fim    = fim    - timedelta(weeks=1)
        comp_label = "semana anterior"
    elif comp_tipo == "MoM":
        c_inicio = inicio - relativedelta(months=1)
        c_fim    = fim    - relativedelta(months=1)
        comp_label = "mês anterior"
    elif comp_tipo == "YoY":
        c_inicio = inicio - relativedelta(years=1)
        c_fim    = fim    - relativedelta(years=1)
        comp_label = "ano anterior"
    else:
        c_inicio = c_fim = None
        comp_label = ""

    if c_inicio:
        st.markdown(f"""
            <div class="comp-badge">
                vs {c_inicio.strftime("%d/%m/%Y")} → {c_fim.strftime("%d/%m/%Y")}
            </div>
        """, unsafe_allow_html=True)

    # ── Filtro por Dia (apenas Resultados) ──
    dia_filtro = None
    if grupo == "📊  Resultados":
        st.markdown("<hr style='border-color:#E8ECF0;margin:16px 0 10px 0;'>", unsafe_allow_html=True)
        st.markdown("<p style='font-size:0.65rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#8C9BAB;margin-bottom:6px;'>Filtrar por Dia</p>", unsafe_allow_html=True)
        dia_filtro = st.date_input("", value=None, min_value=inicio, max_value=fim, label_visibility="collapsed", key="dia_filtro")
        if dia_filtro:
            st.markdown(f"<p style='font-size:0.72rem;color:#1890FF;font-weight:600;margin:4px 0 0 2px;'>📅 {dia_filtro.strftime('%d/%m/%Y')}</p>", unsafe_allow_html=True)

    st.markdown("<p style='font-size:0.65rem;color:#8C9BAB;text-align:center;margin-top:8px;'>Onfly Analytics © 2026</p>", unsafe_allow_html=True)


# ─── Datas formatadas ─────────────────────────────────────────────────────────
if dia_filtro:
    i_str = f_str = dia_filtro.strftime("%Y-%m-%d")
    ci_str = cf_str = None  # desativa comparativo ao filtrar por dia
else:
    i_str  = inicio.strftime("%Y-%m-%d")
    f_str  = fim.strftime("%Y-%m-%d")
    ci_str = c_inicio.strftime("%Y-%m-%d") if c_inicio else None
    cf_str = c_fim.strftime("%Y-%m-%d")    if c_fim    else None


# ─── Renderizadores ───────────────────────────────────────────────────────────
def render_kpis(resumo: dict, comp: dict = None, cor1="", cor2="orange", cor3="green",
                segmentos: int = None, segmentos_comp: int = None):
    st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
    if segmentos is not None:
        c1, c2, c3, c4 = st.columns(4)
    else:
        c1, c2, c3 = st.columns(3)
        c4 = None

    cards = [
        (c1, cor1,    "GMV Total",    brl(resumo["gmv_total"]),      comp["gmv_total"]    if comp else None, resumo["gmv_total"]),
        (c2, cor2,    "Reservas",     f'{resumo["qtd_reservas"]:,}', comp["qtd_reservas"] if comp else None, resumo["qtd_reservas"]),
        (c3, cor3,    "Ticket Médio", brl(resumo["ticket_medio"]),   comp["ticket_medio"] if comp else None, resumo["ticket_medio"]),
    ]
    for col, cor, label, valor, anterior, atual in cards:
        with col:
            st.markdown(f"""
                <div class="kpi-card {cor}">
                    <p class="kpi-label">{label}</p>
                    <p class="kpi-value">{valor}</p>
                    {delta_html(atual, anterior) if anterior is not None else ""}
                </div>
            """, unsafe_allow_html=True)

    if c4 is not None:
        with c4:
            st.markdown(f"""
                <div class="kpi-card" style="border-top-color:#8B5CF6;">
                    <p class="kpi-label">Segmentos</p>
                    <p class="kpi-value">{segmentos:,}</p>
                    {delta_html(segmentos, segmentos_comp) if segmentos_comp is not None else ""}
                </div>
            """, unsafe_allow_html=True)


def render_diario(df: pd.DataFrame, cor: str, df_comp: pd.DataFrame = None, comp_label: str = "",
                  df_extra: pd.DataFrame = None, extra_label: str = "", extra_cor: str = "#F59E0B"):
    st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">Evolução</p></div>', unsafe_allow_html=True)
    if df.empty:
        st.info("Nenhum dado para o período.")
        return

    fill = hex_to_rgba(cor, 0.08)
    fig = go.Figure()

    # Linha de comparação (tracejada, mais clara)
    if df_comp is not None and not df_comp.empty:
        df_c = df_comp.reset_index(drop=True)
        df_c["Dia"] = range(1, len(df_c) + 1)
        df_a = df.reset_index(drop=True)
        df_a["Dia"] = range(1, len(df_a) + 1)

        fig.add_trace(go.Scatter(
            x=df_a["Dia"], y=df_c["GMV"].reindex(df_a.index),
            mode="lines",
            name=comp_label,
            line=dict(color=hex_to_rgba(cor, 0.4), width=1.8, dash="dash"),
            hovertemplate="<b>" + comp_label + "</b><br>Dia %{x}<br>GMV: R$ %{y:,.2f}<extra></extra>",
        ))
        fig.add_trace(go.Scatter(
            x=df_a["Dia"], y=df_a["GMV"],
            mode="lines+markers",
            name="Período atual",
            line=dict(color=cor, width=2.5),
            marker=dict(size=5, color=cor),
            fill="tozeroy",
            fillcolor=fill,
            hovertemplate="<b>Período atual</b><br>Dia %{x}<br>GMV: R$ %{y:,.2f}<extra></extra>",
        ))
        fig.update_xaxes(title_text="Dia do período")
    else:
        fig.add_trace(go.Scatter(
            x=df["Data"], y=df["GMV"],
            mode="lines+markers",
            name="GMV Total",
            line=dict(color=cor, width=2.5),
            marker=dict(size=5, color=cor),
            fill="tozeroy",
            fillcolor=fill,
            hovertemplate="<b>%{x}</b><br>GMV: R$ %{y:,.2f}<extra></extra>",
        ))

    # Linha extra (ex: ONHAPPY) — sobreposta sem fill
    if df_extra is not None and not df_extra.empty:
        _lbl = extra_label or "Destaque"
        if df_comp is not None and not df_comp.empty:
            # No modo comparação o eixo X é "Dia do período"
            _df_e = df_extra.reset_index(drop=True)
            _df_e["Dia"] = range(1, len(_df_e) + 1)
            fig.add_trace(go.Scatter(
                x=_df_e["Dia"], y=_df_e["GMV"],
                mode="lines+markers",
                name=_lbl,
                line=dict(color=extra_cor, width=2, dash="dot"),
                marker=dict(size=4, color=extra_cor),
                hovertemplate=f"<b>{_lbl}</b><br>Dia %{{x}}<br>GMV: R$ %{{y:,.2f}}<extra></extra>",
            ))
        else:
            fig.add_trace(go.Scatter(
                x=df_extra["Data"], y=df_extra["GMV"],
                mode="lines+markers",
                name=_lbl,
                line=dict(color=extra_cor, width=2, dash="dot"),
                marker=dict(size=4, color=extra_cor),
                hovertemplate=f"<b>{_lbl}</b><br><b>%{{x}}</b><br>GMV: R$ %{{y:,.2f}}<extra></extra>",
            ))

    fig.update_yaxes(tickprefix="R$ ")
    st.plotly_chart(plotly_layout(fig, 300), use_container_width=True)


def render_cabine(df: pd.DataFrame, cor: str):
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">Distribuição por Cabine / Família de Tarifa</p></div>', unsafe_allow_html=True)
    if df.empty:
        st.info("Nenhum dado de cabine disponível.")
        return

    col_pie, col_bar = st.columns([1, 1], gap="large")
    with col_pie:
        top = df.head(8)
        fig_pie = go.Figure(go.Pie(
            labels=top["Cabine"], values=top["GMV"],
            hole=0.52,
            marker_colors=px.colors.sequential.Blues_r[:len(top)],
            textinfo="label+percent",
            hovertemplate="<b>%{label}</b><br>GMV: R$ %{value:,.2f}<br>%{percent}<extra></extra>",
        ))
        fig_pie.update_layout(
            height=320, margin=dict(l=0, r=0, t=10, b=0),
            paper_bgcolor="#E8ECF0", showlegend=False,
            font=dict(family="Inter, sans-serif", size=11),
        )
        st.plotly_chart(fig_pie, use_container_width=True)

    with col_bar:
        gmv_total = df["GMV"].sum()
        df_show = df.copy()
        df_show["% GMV"] = (df_show["GMV"] / gmv_total * 100).round(1)
        st.dataframe(
            _brl_df(df_show[["Cabine", "Reservas", "GMV", "% GMV"]]),
            use_container_width=True, hide_index=True, height=320,
            column_config={
                "GMV":      st.column_config.TextColumn("GMV"),
                "Reservas": st.column_config.NumberColumn("Reservas", format="%d"),
                "% GMV":    st.column_config.NumberColumn("% GMV",    format="%.1f%%"),
            },
        )


def render_rotas(df: pd.DataFrame):
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">Top 10 Rotas mais Vendidas</p></div>', unsafe_allow_html=True)
    if df.empty:
        st.info("Nenhuma rota encontrada para o período.")
        return
    gmv_total = df["GMV"].sum()
    df_show = df.copy()
    df_show.insert(2, "Rota", df_show["Origem"] + " → " + df_show["Destino"])
    df_show["% GMV"] = (df_show["GMV"] / gmv_total * 100).round(1)
    st.dataframe(
        _brl_df(df_show[["Rota", "Reservas", "GMV", "% GMV"]]),
        use_container_width=True, hide_index=True,
        column_config={
            "GMV":      st.column_config.TextColumn("GMV"),
            "Reservas": st.column_config.NumberColumn("Reservas", format="%d"),
            "% GMV":    st.column_config.NumberColumn("% GMV",    format="%.1f%%"),
        },
    )


def render_gmv_anual(df: pd.DataFrame, cor: str, proj_gmv_anual: float = None):
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">Evolução do GMV Ano a Ano</p></div>', unsafe_allow_html=True)
    if df.empty:
        st.info("Nenhum dado anual disponível.")
        return
    hoje      = date.today()
    ano_atual = df["Ano"].max()

    # Projeção anual
    linha_ano = df[df["Ano"] == ano_atual]
    gmv_ano   = float(linha_ano["GMV"].iloc[0]) if not linha_ano.empty else 0.0
    proj_resto_anual = max(0.0, (proj_gmv_anual or 0.0) - gmv_ano)

    col_chart, col_table = st.columns([3, 2], gap="large")
    with col_chart:
        colors = [cor if a == ano_atual else hex_to_rgba(cor, 0.45) for a in df["Ano"]]
        anos_str = df["Ano"].astype(str).tolist()

        fig = go.Figure()

        # Barra realizado
        text_real = [
            "" if (a == ano_atual and proj_resto_anual > 0) else brl(v)
            for a, v in zip(df["Ano"], df["GMV"])
        ]
        fig.add_trace(go.Bar(
            name="Realizado",
            x=anos_str, y=df["GMV"],
            marker_color=colors,
            text=text_real,
            textposition="outside",
            textfont=dict(size=11, color="#334155"),
            hovertemplate="<b>%{x}</b><br>Realizado: R$ %{y:,.2f}<extra></extra>",
        ))

        # Barra projeção (só no ano atual)
        if proj_resto_anual > 0:
            proj_vals = [proj_resto_anual if a == ano_atual else 0 for a in df["Ano"]]
            text_proj = [brl(gmv_ano + proj_resto_anual) if a == ano_atual else "" for a in df["Ano"]]
            fig.add_trace(go.Bar(
                name="Projeção",
                x=anos_str, y=proj_vals,
                marker_color=hex_to_rgba("#F59E0B", 0.55),
                text=text_proj,
                textposition="outside",
                textfont=dict(size=11, color="#334155"),
                hovertemplate="<b>%{x}</b><br>Projeção restante: R$ %{y:,.2f}<extra></extra>",
            ))

        fig.update_layout(
            barmode="stack",
            showlegend=proj_resto_anual > 0,
            legend=dict(
                orientation="h",
                yanchor="bottom", y=1.02,
                xanchor="left",   x=0,
                font=dict(size=12),
            ),
        )
        fig.update_yaxes(tickprefix="R$ ")
        fig.update_xaxes(type="category")
        st.plotly_chart(plotly_layout(fig, 300), use_container_width=True)

    with col_table:
        df_show = df.copy()
        # Normaliza nome da coluna de contagem (Reservas ou Trechos)
        if "Trechos" in df_show.columns and "Reservas" not in df_show.columns:
            df_show = df_show.rename(columns={"Trechos": "Reservas"})
        # Para o ano atual, usa o projetado na coluna GMV da tabela
        if proj_gmv_anual and proj_gmv_anual > 0:
            df_show.loc[df_show["Ano"] == ano_atual, "GMV"] = proj_gmv_anual
        df_show["Var. YoY"] = df_show["GMV"].pct_change() * 100
        df_show["GMV Label"] = df_show.apply(
            lambda r: f"{brl(r['GMV'])} *" if (r["Ano"] == ano_atual and proj_gmv_anual) else brl(r["GMV"]),
            axis=1
        )
        st.dataframe(
            df_show[["Ano", "Reservas", "GMV Label", "Var. YoY"]].rename(columns={"GMV Label": "GMV"}),
            use_container_width=True, hide_index=True, height=300,
            column_config={
                "Ano":      st.column_config.NumberColumn("Ano",      format="%d"),
                "GMV":      st.column_config.TextColumn("GMV"),
                "Reservas": st.column_config.NumberColumn("Reservas", format="%d"),
                "Var. YoY": st.column_config.NumberColumn("Var. YoY", format="%.1f%%"),
            },
        )
        if proj_gmv_anual:
            st.caption("* Projeção baseada na sazonalidade do ano anterior")


MESES_PT = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
            "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

def render_gmv_mensal(df: pd.DataFrame, cor: str, proj_gmv: float = None):
    from calendar import monthrange as _mr
    hoje      = date.today()
    ano_atual = hoje.year
    mes_atual = hoje.month
    st.markdown(f'<div class="sec-header-wrap"><p class="sec-header">GMV Mês a Mês — {ano_atual}</p></div>', unsafe_allow_html=True)
    if df.empty or df["GMV"].sum() == 0:
        st.info("Nenhum dado mensal disponível.")
        return
    labels = [MESES_PT[int(m) - 1] for m in df["Mês"]]
    colors = [cor if int(m) == mes_atual else hex_to_rgba(cor, 0.45) for m in df["Mês"]]

    # Calcula GMV atual do mês corrente e o restante projetado
    linha_mes = df[df["Mês"].astype(int) == mes_atual]
    gmv_atual = float(linha_mes["GMV"].iloc[0]) if not linha_mes.empty else 0.0
    proj_resto = max(0.0, (proj_gmv or 0.0) - gmv_atual)

    fig = go.Figure()

    # Barra base — realizado (sem label no mês atual quando há projeção)
    text_realizado = [
        "" if (int(m) == mes_atual and proj_resto > 0) else (brl(v) if v > 0 else "")
        for m, v in zip(df["Mês"], df["GMV"])
    ]
    fig.add_trace(go.Bar(
        name="Realizado",
        x=labels,
        y=df["GMV"],
        marker_color=colors,
        text=text_realizado,
        textposition="outside",
        textfont=dict(size=11, color="#334155"),
        hovertemplate="<b>%{x}</b><br>Realizado: R$ %{y:,.2f}<extra></extra>",
    ))

    # Barra de projeção (empilhada no mês corrente, 0 nos demais)
    if proj_resto > 0:
        proj_vals = [proj_resto if int(m) == mes_atual else 0 for m in df["Mês"]]
        # Label no topo da barra empilhada = total projetado
        text_proj = [brl(gmv_atual + proj_resto) if int(m) == mes_atual else "" for m in df["Mês"]]
        fig.add_trace(go.Bar(
            name="Projeção",
            x=labels,
            y=proj_vals,
            marker_color=hex_to_rgba("#F59E0B", 0.55),
            text=text_proj,
            textposition="outside",
            textfont=dict(size=11, color="#334155"),
            hovertemplate="<b>%{x}</b><br>Projeção restante: R$ %{y:,.2f}<extra></extra>",
        ))

    fig.update_layout(
        barmode="stack",
        showlegend=proj_resto > 0,
        legend=dict(
            orientation="h",
            yanchor="bottom", y=1.02,
            xanchor="left",   x=0,
            font=dict(size=12),
        ),
    )
    fig.update_yaxes(tickprefix="R$ ")
    fig.update_xaxes(type="category")
    st.plotly_chart(plotly_layout(fig, 340), use_container_width=True)


def _proj_mensal_de_df(df_mensal: "pd.DataFrame") -> "float | None":
    """Projeta o GMV do mês corrente com base na média diária (df_mensal já carregado)."""
    from calendar import monthrange as _mr
    hoje = date.today()
    linha = df_mensal[df_mensal["Mês"].astype(int) == hoje.month]
    if linha.empty or hoje.day == 0:
        return None
    gmv_so_far = float(linha["GMV"].iloc[0])
    return (gmv_so_far / hoje.day) * _mr(hoje.year, hoje.month)[1]


def render_tendencia_cia(cia: str, cor: str):
    """Bloco de tendência GMV para uma cia: mês atual vs mês anterior + projeção."""
    from calendar import monthrange as _mr

    _hoje_tc   = date.today()
    _mes_at    = _hoje_tc.month
    _ano_at    = _hoje_tc.year
    _mes_ant   = _mes_at - 1 if _mes_at > 1 else 12
    _ano_ant   = _ano_at if _mes_at > 1 else _ano_at - 1

    _MESES_PT  = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                  "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
    _label_at  = _MESES_PT[_mes_at - 1]
    _label_ant = _MESES_PT[_mes_ant - 1]
    _tot_at    = _mr(_ano_at,  _mes_at)[1]
    _tot_ant   = _mr(_ano_ant, _mes_ant)[1]

    with st.spinner("Carregando tendência..."):
        _df_at  = q_tendencia_diario_cia(_mes_at,  _ano_at,  cia)
        _df_ant = q_tendencia_diario_cia(_mes_ant, _ano_ant, cia)

    # ── Cálculos ──────────────────────────────────────────────────────────────
    if not _df_at.empty:
        _dias_c   = len(_df_at)
        _gmv_now  = float(_df_at["gmv"].sum())
        _proj     = (_gmv_now / _dias_c) * _tot_at if _dias_c > 0 else 0.0
    else:
        _gmv_now = _proj = 0.0

    _gmv_ant = float(_df_ant["gmv"].sum()) if not _df_ant.empty else 0.0
    _var     = ((_proj - _gmv_ant) / _gmv_ant * 100) if _gmv_ant > 0 else 0.0

    # ── Badge ──────────────────────────────────────────────────────────────────
    _bc  = "#27AE60" if _var >= 0 else "#C0392B"
    _em  = "🟢" if _var >= 0 else "🔴"
    _txt = f"{'Acima' if _var >= 0 else 'Abaixo'} de {_label_ant}  ({'%+.1f' % _var}%)"
    st.markdown(
        f"<div style='background:{_bc}18;border-left:4px solid {_bc};"
        f"padding:10px 16px;border-radius:6px;margin-bottom:16px;"
        f"font-weight:600;color:{_bc};'>{_em} {_txt}</div>",
        unsafe_allow_html=True,
    )

    # ── KPI cards ──────────────────────────────────────────────────────────────
    _k1, _k2, _k3 = st.columns(3)
    with _k1:
        st.markdown(f'<div class="kpi-card"><p class="kpi-label">GMV Projetado ({_label_at})</p><p class="kpi-value">{brl(_proj)}</p></div>', unsafe_allow_html=True)
    with _k2:
        st.markdown(f'<div class="kpi-card"><p class="kpi-label">Realizado até hoje</p><p class="kpi-value">{brl(_gmv_now)}</p></div>', unsafe_allow_html=True)
    with _k3:
        _vc = "#27AE60" if _var >= 0 else "#C0392B"
        _vs = "▲" if _var >= 0 else "▼"
        st.markdown(f'<div class="kpi-card"><p class="kpi-label">vs {_label_ant}</p><p class="kpi-value" style="color:{_vc};">{_vs} {abs(_var):.1f}%</p><p style="font-size:0.8rem;color:#888;">{brl(_gmv_ant)}</p></div>', unsafe_allow_html=True)

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

    # ── Gráfico cumulativo ─────────────────────────────────────────────────────
    def _cum(df, total_dias):
        if df.empty:
            return [], []
        s = df.groupby("dia_num")["gmv"].sum().reindex(range(1, total_dias + 1), fill_value=0)
        return list(range(1, total_dias + 1)), s.cumsum().tolist()

    _fig = go.Figure()
    _dant, _cant = _cum(_df_ant, _tot_ant)
    if _dant:
        _fig.add_trace(go.Scatter(
            x=_dant, y=_cant, mode="lines", name=_label_ant,
            line=dict(color="#94A3B8", width=2, dash="dash"),
            hovertemplate="Dia %{x}<br>Acum.: R$ %{y:,.0f}<extra></extra>",
        ))
    if not _df_at.empty:
        _last = int(_df_at["dia_num"].max())
        _s_at = _df_at.groupby("dia_num")["gmv"].sum().reindex(range(1, _last + 1), fill_value=0).cumsum()
        _fig.add_trace(go.Scatter(
            x=list(range(1, _last + 1)), y=_s_at.tolist(), mode="lines",
            name=f"{_label_at} (real)", line=dict(color=cor, width=3),
            hovertemplate="Dia %{x}<br>Acum.: R$ %{y:,.0f}<extra></extra>",
        ))
        _dp = list(range(_last, _tot_at + 1))
        _cp = list(np.linspace(float(_s_at.iloc[-1]), _proj, len(_dp)))
        _fig.add_trace(go.Scatter(
            x=_dp, y=_cp, mode="lines", name="Projeção",
            line=dict(color=cor, width=2, dash="dot"),
            hovertemplate="Dia %{x}<br>Projeção: R$ %{y:,.0f}<extra></extra>",
        ))

    plotly_layout(_fig, height=300)
    _fig.update_layout(
        xaxis_title="Dia do mês", yaxis_title="GMV Acumulado (R$)",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    )
    st.plotly_chart(_fig, use_container_width=True)


def render_airline_tab(cia: str, cor: str, cor_kpi1: str, variantes: tuple = ()):
    _ano_periodo = date.fromisoformat(i_str).year
    with st.spinner(f"Carregando dados {cia}..."):
        try:
            if variantes:
                # Usa filtro por segmento (airline code) — captura histórico mesmo com consolidador misto
                resumo    = q_resumo_seg(i_str, f_str, variantes)
                df_day    = q_diario_seg(i_str, f_str, variantes)
                df_mensal = q_gmv_mensal_seg(variantes, ano=_ano_periodo)
                df_ano    = q_gmv_anual_cia(variantes)
                comp      = q_resumo_seg(ci_str, cf_str, variantes) if ci_str else None
                df_comp   = q_diario_seg(ci_str, cf_str, variantes) if ci_str else None
            else:
                resumo    = q_resumo(i_str, f_str, cia)
                df_day    = q_diario(i_str, f_str, cia)
                df_mensal = q_gmv_mensal(cia, ano=_ano_periodo)
                df_ano    = q_gmv_anual(cia)
                comp      = q_resumo(ci_str, cf_str, cia) if ci_str else None
                df_comp   = q_diario(ci_str, cf_str, cia) if ci_str else None
            df_cab   = q_cabine(i_str, f_str, cia)
            df_rot   = q_rotas(i_str, f_str, cia)
            segs     = q_segmentos(i_str, f_str, cia)
            segs_c   = q_segmentos(ci_str, cf_str, cia) if ci_str else None
        except Exception as e:
            st.error(f"Erro ao conectar ao BigQuery: {e}")
            return

    render_kpis(resumo, comp, cor_kpi1, "orange", "green", segmentos=segs, segmentos_comp=segs_c)
    render_diario(df_day, cor, df_comp, comp_label)
    render_gmv_mensal(df_mensal, cor, proj_gmv=_proj_mensal_de_df(df_mensal))
    render_cabine(df_cab, cor)
    render_rotas(df_rot)
    render_gmv_anual(df_ano, cor)


# ─── Distribuição — persistência local ───────────────────────────────────────
DISTRIBUICAO_PATH = os.path.join(os.path.dirname(__file__), "distribuicao.json")

CONSOLIDADORAS = [
    "IATA-AZUL-LATAM-BTC-ACORDO", "LATAM", "GOL", "REXTUR_IATA", "AZUL",
    "FLYTOURADTOCREDITCARD", "FLYTOURLAG3", "IATAAZULLATANBTCACORDO",
    "CONFIANÇA", "FLIGHT", "REXTUR", "SAKURAWITHDU", "FLYTOURTO",
    "AMADEUS", "ONFLYAMADEUS", "GRUPOS-EVENTOS", "FLYTOUR",
    "LUNA", "MILHASFACIL", "CONSOLIDAMADEUS", "BRT", "COMPANHIA",
    "AGREEMENTCUSTUMER", "SKYTEAM", "PRIMUS", "VOETUR", "UNIGLOBE",
]

def load_distribuicao() -> dict:
    import json
    if os.path.exists(DISTRIBUICAO_PATH):
        with open(DISTRIBUICAO_PATH) as f:
            return json.load(f)
    return {}

def save_distribuicao(data: dict):
    import json
    with open(DISTRIBUICAO_PATH, "w") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

ANOTACOES_PATH = os.path.join(os.path.dirname(__file__), "anotacoes.json")

def load_anotacoes() -> list:
    import json
    if os.path.exists(ANOTACOES_PATH):
        with open(ANOTACOES_PATH) as f:
            return json.load(f)
    return []

def save_anotacoes(data: list):
    import json
    with open(ANOTACOES_PATH, "w") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def _get_anexos(cfg: dict) -> list:
    """Retorna lista de {nome, dados (base64), mime}.
    Faz migração automática do campo legado comissao_bsp_imagem."""
    if cfg.get("anexos"):
        return list(cfg["anexos"])
    b64 = cfg.get("comissao_bsp_imagem")
    if b64:
        nome = cfg.get("comissao_bsp_imagem_nome") or "comprovante"
        ext  = nome.rsplit(".", 1)[-1].lower() if "." in nome else "png"
        mime = "application/pdf" if ext == "pdf" else ("image/jpeg" if ext in ("jpg","jpeg") else "image/png")
        return [{"nome": nome, "dados": b64, "mime": mime}]
    return []

# ─── Incentivo — persistência local ──────────────────────────────────────────
INCENTIVO_DIST_PATH = os.path.join(os.path.dirname(__file__), "incentivo_dist.json")

TIPOS_INCENTIVO = ["Up Front", "Backend", "Comissão", "OBT"]

def load_incentivo_dist() -> dict:
    import json
    if os.path.exists(INCENTIVO_DIST_PATH):
        with open(INCENTIVO_DIST_PATH) as f:
            return json.load(f)
    return {}

def save_incentivo_dist(data: dict):
    import json
    with open(INCENTIVO_DIST_PATH, "w") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ─── Incentivo Azul — persistência local ─────────────────────────────────────
INCENTIVO_AZUL_PATH   = os.path.join(os.path.dirname(__file__), "incentivo_azul.json")
MESES_ABREV           = ["JAN","FEV","MAR","ABR","MAI","JUN","JUL","AGO","SET","OUT","NOV","DEZ"]
AZUL_NAC_PCTS         = [0.015, 0.020, 0.025, 0.030, 0.035, 0.040]   # fixos 2026
AZUL_INT_PCTS         = [0.005, 0.010, 0.015, 0.020, 0.025, 0.030]   # fixos 2026
AZUL_NAC_LABELS       = ["1º(1,5%)", "2º(2,0%)", "3º(2,5%)", "4º(3,0%)", "5º(3,5%)", "6º(4,0%)"]
AZUL_INT_LABELS       = ["1º(0,5%)", "2º(1,0%)", "3º(1,5%)", "4º(2,0%)", "5º(2,5%)", "6º(3,0%)"]

# ─── Incentivo LATAM — persistência local ────────────────────────────────────
INCENTIVO_LATAM_PATH = os.path.join(os.path.dirname(__file__), "incentivo_latam.json")
LATAM_TAXA_REF       = 0.065   # 6,50% — fixo contrato
LATAM_FIXO_CORP      = 0.030   # 3,00% — fixo contrato
LATAM_PESO_REC       = 0.80    # 80%
LATAM_PESO_PRM       = 0.20    # 20%

def load_incentivo_latam() -> dict:
    import json
    if os.path.exists(INCENTIVO_LATAM_PATH):
        with open(INCENTIVO_LATAM_PATH) as f:
            return json.load(f)
    return {}

def save_incentivo_latam(data: dict):
    import json
    with open(INCENTIVO_LATAM_PATH, "w") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

LATAM_FIXO_INTER = 0.035   # 3,5% fixo — Internacional

def _latam_inter_calcular(realizado, ranges_data: list) -> dict | None:
    """
    Interpolação linear dentro da faixa + fixo 3,5%.
    ranges_data: list of {min, max, taxa_min, taxa_max} com taxas em % (ex: 0.5 para 0,50%).
    Retorna dict com variavel_pct, pct_incent, rs_incent; ou None se incompleto.
    """
    if not realizado or (isinstance(realizado, float) and pd.isna(realizado)) or realizado <= 0:
        return None
    valid = [r for r in (ranges_data or [])
             if r.get("min") is not None and r.get("max") is not None]
    if not valid:
        return None
    variavel_pct = 0.0
    for r in valid:
        r_min, r_max = r["min"], r["max"]
        t_min = (r.get("taxa_min") or 0) / 100
        t_max = (r.get("taxa_max") or 0) / 100
        if r_min <= realizado <= r_max:
            ratio = (realizado - r_min) / (r_max - r_min) if r_max != r_min else 1.0
            variavel_pct = (t_min + ratio * (t_max - t_min)) * 100
            break
    else:
        if realizado < valid[0]["min"]:
            variavel_pct = 0.0
        else:  # acima do último range → taxa_max do último
            variavel_pct = valid[-1].get("taxa_max", 0)
    pct_incent = LATAM_FIXO_INTER * 100 + variavel_pct
    return {
        "variavel_pct": variavel_pct,
        "pct_incent":   pct_incent,
        "rs_incent":    realizado * pct_incent / 100,
    }

def _latam_calcular(meta_receita, realizado_proj, baseline_prm, realizado_prm, taxa_real) -> dict | None:
    """Retorna dict com todos os KPIs calculados, ou None se dados incompletos."""
    vals = [meta_receita, realizado_proj, baseline_prm, realizado_prm, taxa_real]
    if any(v is None or (isinstance(v, float) and pd.isna(v)) or v == 0 for v in vals):
        return None
    pct_rec    = realizado_proj / meta_receita
    pct_prm    = realizado_prm  / baseline_prm
    pct_final  = LATAM_PESO_REC * pct_rec + LATAM_PESO_PRM * pct_prm
    variavel   = (taxa_real - LATAM_FIXO_CORP) * pct_final
    pct_incent = LATAM_FIXO_CORP + variavel
    rs_incent  = realizado_proj * pct_incent
    return {
        "pct_rec":    pct_rec,
        "pct_prm":    pct_prm,
        "pct_final":  pct_final,
        "variavel":   variavel,
        "pct_incent": pct_incent,
        "rs_incent":  rs_incent,
    }

def load_incentivo_azul() -> dict:
    import json
    if os.path.exists(INCENTIVO_AZUL_PATH):
        with open(INCENTIVO_AZUL_PATH) as f:
            return json.load(f)
    return {}

def save_incentivo_azul(data: dict):
    import json
    with open(INCENTIVO_AZUL_PATH, "w") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def _azul_nivel(realizado, thresholds: list, pcts: list) -> tuple:
    """Retorna (pct_float, pago_float). None se realizado ausente ou abaixo do 1º nível."""
    if realizado is None or (isinstance(realizado, float) and pd.isna(realizado)) or realizado <= 0:
        return None, None
    pct = 0.0
    for thr, p in zip(thresholds, pcts):
        if thr is not None and not (isinstance(thr, float) and pd.isna(thr)) and realizado >= thr:
            pct = p
    if pct == 0:
        return 0.0, 0.0
    return pct, realizado * pct

def _azul_build_df(dados_bloco: dict, labels: list) -> pd.DataFrame:
    """DataFrame somente com colunas de entrada (sem colunas calculadas)."""
    rows = []
    for mes in MESES_ABREV:
        d = dados_bloco.get(mes, {})
        real = d.get("realizado")
        nivs = d.get("niveis", [None]*6)
        while len(nivs) < 6:
            nivs.append(None)
        row: dict = {"MÊS": mes, "Realizado": real}
        for lbl, val in zip(labels, nivs):
            row[lbl] = val
        rows.append(row)
    return pd.DataFrame(rows)

def _azul_compute_results(edited_df: pd.DataFrame, pcts: list, labels: list) -> pd.DataFrame:
    """Recalcula % Incentivo e Pago a partir do DataFrame editado."""
    rows = []
    for _, r in edited_df.iterrows():
        real = r["Realizado"]
        nivs = [r[lbl] for lbl in labels]
        pct, pago = _azul_nivel(real, nivs, pcts)
        nivel_nome = "—"
        if pct and pct > 0:
            for i, p in enumerate(pcts):
                if abs(p - pct) < 1e-9:
                    nivel_nome = labels[i]
                    break
        rows.append({
            "MÊS":        r["MÊS"],
            "Realizado":  real or 0.0,
            "Nível":      nivel_nome,
            "% Incentivo": f"{pct*100:.1f}%" if pct else "—",
            "Pago":       round(pago, 2) if pago else 0.0,
        })
    return pd.DataFrame(rows)


# ─── Incentivos — persistência local ─────────────────────────────────────────
INCENTIVOS_PATH = os.path.join(os.path.dirname(__file__), "incentivos.json")

def load_incentivos() -> dict:
    if os.path.exists(INCENTIVOS_PATH):
        import json
        with open(INCENTIVOS_PATH) as f:
            return json.load(f)
    return {}

def save_incentivos(data: dict):
    import json
    with open(INCENTIVOS_PATH, "w") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_incentivo_periodo(inicio: str, fim: str) -> dict:
    """Soma incentivos de todos os meses dentro do período."""
    import json
    from datetime import date
    dados = load_incentivos()
    d_ini = date.fromisoformat(inicio)
    d_fim = date.fromisoformat(fim)
    totais: dict[str, float] = {}
    cur = d_ini.replace(day=1)
    while cur <= d_fim:
        chave = cur.strftime("%Y-%m")
        mes = dados.get(chave, {})
        for cia, val in mes.items():
            totais[cia] = totais.get(cia, 0.0) + val
        if cur.month == 12:
            cur = cur.replace(year=cur.year + 1, month=1)
        else:
            cur = cur.replace(month=cur.month + 1)
    return totais


# ─── Queries — Take Rate ──────────────────────────────────────────────────────
LABEL_TIPO = {"flight": "Aéreo", "hotel": "Hotel", "auto": "Carro", "bus": "Ônibus"}

def _tr_where(inicio, fim):
    return f"""
        WHERE type = 'flight' AND status = 2

          AND created_at >= '{inicio}'
          AND created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
    """

@st.cache_data(ttl=300, show_spinner=False)
def q_tr_resumo(inicio: str, fim: str) -> dict:
    q = f"""
        SELECT
            ROUND(SUM(total_amount_currency_brl), 2)  AS gmv,
            ROUND(SUM(profit_currency_brl), 2)        AS gross_revenue,
            ROUND(SUM(profit_currency_brl) / NULLIF(SUM(total_amount_currency_brl), 0) * 100, 2) AS take_rate
        FROM `{TABLE}` {_tr_where(inicio, fim)}
    """
    row = list(bq_client().query(q).result())[0]
    return {"gmv": float(row.gmv or 0), "gross_revenue": float(row.gross_revenue or 0), "take_rate": float(row.take_rate or 0)}

@st.cache_data(ttl=300, show_spinner=False)
def q_tr_diario(inicio: str, fim: str) -> pd.DataFrame:
    q = f"""
        SELECT
            DATE(created_at)                         AS dia,
            ROUND(SUM(total_amount_currency_brl), 2) AS gmv,
            ROUND(SUM(profit_currency_brl), 2)       AS gross_revenue,
            ROUND(SUM(profit_currency_brl) / NULLIF(SUM(total_amount_currency_brl), 0) * 100, 2) AS take_rate
        FROM `{TABLE}` {_tr_where(inicio, fim)}
        GROUP BY dia ORDER BY dia
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Data": str(r.dia), "GMV": float(r.gmv or 0),
         "Gross Revenue": float(r.gross_revenue or 0), "Take Rate (%)": float(r.take_rate or 0)}
        for r in rows
    ])

@st.cache_data(ttl=300, show_spinner=False)
def q_tr_por_cia(inicio: str, fim: str) -> pd.DataFrame:
    q = f"""
        SELECT
            consolidator_unified                          AS cia,
            ROUND(SUM(total_amount_currency_brl), 2)      AS gmv,
            ROUND(SUM(onfly_amount_currency_brl), 2)      AS onfly_amount,
            ROUND(SUM(profit_currency_brl), 2)            AS gross_revenue,
            ROUND(SUM(profit_currency_brl) / NULLIF(SUM(total_amount_currency_brl), 0) * 100, 2) AS take_rate
        FROM `{TABLE}` {_tr_where(inicio, fim)}
        GROUP BY cia ORDER BY gmv DESC LIMIT 15
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Cia": r.cia, "GMV": float(r.gmv or 0),
         "Onfly Amount": float(r.onfly_amount or 0),
         "Gross Revenue": float(r.gross_revenue or 0), "Take Rate (%)": float(r.take_rate or 0)}
        for r in rows
    ])

@st.cache_data(ttl=300, show_spinner=False)
def q_tr_negativos(inicio: str, fim: str) -> pd.DataFrame:
    """Todas as cias com take rate negativo no período — sem LIMIT."""
    q = f"""
        SELECT
            consolidator_unified                          AS cia,
            ROUND(SUM(total_amount_currency_brl), 2)      AS gmv,
            ROUND(SUM(profit_currency_brl), 2)            AS gross_revenue,
            ROUND(SUM(profit_currency_brl) / NULLIF(SUM(total_amount_currency_brl), 0) * 100, 2) AS take_rate
        FROM `{TABLE}` {_tr_where(inicio, fim)}
        GROUP BY cia
        HAVING take_rate < 0
        ORDER BY take_rate ASC
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Cia": r.cia, "GMV": float(r.gmv or 0),
         "Gross Revenue": float(r.gross_revenue or 0), "Take Rate (%)": float(r.take_rate or 0)}
        for r in rows
    ])


@st.cache_data(ttl=300, show_spinner=False)
def q_tr_drill_transacoes(inicio: str, fim: str) -> pd.DataFrame:
    """Transações individuais de Take Rate para drill-down por dia."""
    q = f"""
        SELECT
            e.uuid                                                              AS Reserva,
            COALESCE(c.name, 'Não identificado')                                AS Empresa,
            e.consolidator_unified                                              AS Cia,
            ROUND(e.total_amount_currency_brl, 2)                              AS GMV,
            ROUND(e.profit_currency_brl, 2)                                    AS gross_revenue,
            ROUND(e.profit_currency_brl / NULLIF(e.total_amount_currency_brl,0) * 100, 2) AS take_rate,
            CAST(e.created_at AS STRING)                                        AS Horario
        FROM `{TABLE}` e
        LEFT JOIN `{TABLE_COMP}` c ON c.company_id = e.company_id
        WHERE e.type = 'flight' AND e.status = 2
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        ORDER BY e.created_at DESC
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {
            "Horário":       str(r.Horario)[11:19],
            "Reserva":       r.Reserva,
            "Empresa":       r.Empresa,
            "Cia":           r.Cia,
            "GMV":           float(r.GMV or 0),
            "Gross Revenue": float(r.gross_revenue or 0),
            "Take Rate (%)": float(r.take_rate or 0),
        }
        for r in rows
    ])

@st.cache_data(ttl=300, show_spinner=False)
def q_tr_drill_clientes(inicio: str, fim: str) -> pd.DataFrame:
    """Agrupado por cliente para drill-down de Take Rate."""
    q = f"""
        SELECT
            COALESCE(c.name, 'Não identificado')                                AS Empresa,
            COUNT(*)                                                            AS Reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)                         AS GMV,
            ROUND(SUM(e.profit_currency_brl), 2)                               AS gross_revenue,
            ROUND(SUM(e.profit_currency_brl) / NULLIF(SUM(e.total_amount_currency_brl),0) * 100, 2) AS take_rate
        FROM `{TABLE}` e
        LEFT JOIN `{TABLE_COMP}` c ON c.company_id = e.company_id
        WHERE e.type = 'flight' AND e.status = 2
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        GROUP BY 1
        ORDER BY GMV DESC
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {
            "Empresa":       r.Empresa,
            "Reservas":      int(r.Reservas or 0),
            "GMV":           float(r.GMV or 0),
            "Gross Revenue": float(r.gross_revenue or 0),
            "Take Rate (%)": float(r.take_rate or 0),
        }
        for r in rows
    ])


@st.cache_data(ttl=300, show_spinner=False)
@st.cache_data(ttl=300, show_spinner=False)
def q_cias_balanceamento(inicio: str, fim: str) -> list:
    """Lista de cias aéreas disponíveis no período para o filtro de Balanceamento."""
    q = f"""
        SELECT DISTINCT UPPER(TRIM(s.company_operator)) AS cia
        FROM `{TABLE_SEG}` s
        JOIN `{TABLE}` e ON RTRIM(s.uuid, '_') = e.uuid
        WHERE e.type = 'flight' AND e.status = 2
          AND s.segment = 0 AND s.step = 1
          AND s.company_operator IS NOT NULL AND TRIM(s.company_operator) != ''
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        ORDER BY cia
    """
    rows = list(bq_client().query(q).result())
    return [r.cia for r in rows]


@st.cache_data(ttl=300, show_spinner=False)
def q_balanceamento(inicio: str, fim: str, cias_sel: tuple) -> pd.DataFrame:
    """Nacional/Internacional × Manual/Automático.
    Nacional = LATAM, TAM, AZUL, AZUL CONECTA, GOL (independente do destino).
    Internacional = todas as demais cias operadoras.
    Filtro por cia aplicado quando cias_sel não está vazio.
    """
    _manual_sql  = ", ".join(f"'{c}'" for c in sorted(_CANAIS_MANUAL))
    _nacionais   = "'LATAM', 'TAM', 'AZUL', 'AZUL CONECTA', 'GOL'"
    _filtro_cia  = (f"AND s.cia IN ({', '.join(f'{chr(39)}{c}{chr(39)}' for c in cias_sel)})"
                    if cias_sel else "")
    q = f"""
        WITH segs AS (
            SELECT
                REPLACE(RTRIM(uuid, '_'), '_flight', '') AS seg_protocol,
                UPPER(TRIM(company_operator))            AS cia
            FROM `{TABLE_SEG}`
            WHERE segment = 0 AND step = 1
              AND company_operator IS NOT NULL AND TRIM(company_operator) != ''
            QUALIFY ROW_NUMBER() OVER (PARTITION BY uuid ORDER BY company_operator) = 1
        )
        SELECT
            CASE WHEN s.cia IN ({_nacionais}) THEN 'Nacional' ELSE 'Internacional' END AS escopo,
            CASE WHEN UPPER(TRIM(e.consolidator_unified)) IN ({_manual_sql})
                 THEN 'Manual' ELSE 'Automático' END                                   AS canal,
            COUNT(DISTINCT e.uuid)                                                     AS reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)                                AS gmv
        FROM `{TABLE}` e
        JOIN segs s ON s.seg_protocol = REPLACE(e.uuid, '_flight', '')
        WHERE e.type = 'flight' AND e.status = 2
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
          {_filtro_cia}
        GROUP BY 1, 2
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Escopo": r.escopo, "Canal": r.canal,
         "Reservas": int(r.reservas or 0), "GMV": float(r.gmv or 0)}
        for r in rows
    ])


@st.cache_data(ttl=300, show_spinner=False)
def q_balanceamento_canais(inicio: str, fim: str, cias_sel: tuple) -> pd.DataFrame:
    """Detalha onde foram feitas as emissões MANUAIS, por canal e escopo."""
    _manual_sql = ", ".join(f"'{c}'" for c in sorted(_CANAIS_MANUAL))
    _nacionais  = "'LATAM', 'TAM', 'AZUL', 'AZUL CONECTA', 'GOL'"
    _filtro_cia = (f"AND s.cia IN ({', '.join(f'{chr(39)}{c}{chr(39)}' for c in cias_sel)})"
                   if cias_sel else "")
    q = f"""
        WITH segs AS (
            SELECT
                REPLACE(RTRIM(uuid, '_'), '_flight', '') AS seg_protocol,
                UPPER(TRIM(company_operator))            AS cia
            FROM `{TABLE_SEG}`
            WHERE segment = 0 AND step = 1
              AND company_operator IS NOT NULL AND TRIM(company_operator) != ''
            QUALIFY ROW_NUMBER() OVER (PARTITION BY uuid ORDER BY company_operator) = 1
        )
        SELECT
            CASE WHEN s.cia IN ({_nacionais}) THEN 'Nacional' ELSE 'Internacional' END AS escopo,
            COALESCE(NULLIF(TRIM(e.consolidator_unified), ''), 'Não informado')        AS canal,
            COUNT(DISTINCT e.uuid)                                                     AS reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)                                AS gmv
        FROM `{TABLE}` e
        JOIN segs s ON s.seg_protocol = REPLACE(e.uuid, '_flight', '')
        WHERE e.type = 'flight' AND e.status = 2
          AND UPPER(TRIM(e.consolidator_unified)) IN ({_manual_sql})
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
          {_filtro_cia}
        GROUP BY 1, 2
        ORDER BY escopo, gmv DESC
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Escopo": r.escopo, "Canal": r.canal,
         "Reservas": int(r.reservas or 0), "GMV": float(r.gmv or 0)}
        for r in rows
    ])


@st.cache_data(ttl=300, show_spinner=False)
def q_balanceamento_emissores(inicio: str, fim: str, cias_sel: tuple) -> pd.DataFrame:
    """Ranking de emissores manuais por escopo (Nacional/Internacional)."""
    _manual_sql = ", ".join(f"'{c}'" for c in sorted(_CANAIS_MANUAL))
    _nacionais  = "'LATAM', 'TAM', 'AZUL', 'AZUL CONECTA', 'GOL'"
    _filtro_cia = (f"AND s.cia IN ({', '.join(f'{chr(39)}{c}{chr(39)}' for c in cias_sel)})"
                   if cias_sel else "")
    q = f"""
        WITH segs AS (
            SELECT
                REPLACE(RTRIM(uuid, '_'), '_flight', '') AS seg_protocol,
                UPPER(TRIM(company_operator))            AS cia
            FROM `{TABLE_SEG}`
            WHERE segment = 0 AND step = 1
              AND company_operator IS NOT NULL AND TRIM(company_operator) != ''
            QUALIFY ROW_NUMBER() OVER (PARTITION BY uuid ORDER BY company_operator) = 1
        )
        SELECT
            CASE WHEN s.cia IN ({_nacionais}) THEN 'Nacional' ELSE 'Internacional' END AS escopo,
            COALESCE(NULLIF(TRIM(e.emitter_name), ''), 'Não informado')                AS emissor,
            COALESCE(NULLIF(TRIM(e.consolidator_unified), ''), 'Não informado')        AS canal,
            COUNT(DISTINCT e.uuid)                                                     AS reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)                                AS gmv
        FROM `{TABLE}` e
        JOIN segs s ON s.seg_protocol = REPLACE(e.uuid, '_flight', '')
        WHERE e.type = 'flight' AND e.status = 2
          AND UPPER(TRIM(e.consolidator_unified)) IN ({_manual_sql})
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
          {_filtro_cia}
        GROUP BY 1, 2, 3
        ORDER BY escopo, gmv DESC
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Escopo": r.escopo, "Emissor": r.emissor, "Canal": r.canal,
         "Reservas": int(r.reservas or 0), "GMV": float(r.gmv or 0)}
        for r in rows
    ])


def q_cia_legado(inicio: str, fim: str) -> pd.DataFrame:
    """GMV e GMV Incentivo por cia aérea, normalizado por código IATA.
    ROW_NUMBER garante 1 segmento por uuid (primeira saída), evitando
    double-counting em voos com conexão operados por cias diferentes."""
    q = f"""
        WITH ranked AS (
            SELECT
                s.uuid,
                UPPER(TRIM(s.company_operator))                    AS cia_raw,
                ROW_NUMBER() OVER (
                    PARTITION BY s.uuid
                    ORDER BY s.departure_date_hour ASC NULLS LAST
                )                                         AS rn
            FROM `{TABLE_SEG}` s
            JOIN `{TABLE}` e ON RTRIM(s.uuid, '_') = e.uuid
            WHERE e.type = 'flight' AND e.status = 2
              AND s.segment = 0 AND s.step = 1
              AND e.created_at >= '{inicio}'
              AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
              AND s.company_operator IS NOT NULL AND TRIM(s.company_operator) != ''
        )
        SELECT
            r.cia_raw,
            COUNT(DISTINCT r.uuid)                        AS reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)    AS gmv,
            ROUND(SUM(e.onfly_amount_currency_brl), 2)    AS gmv_bilhete
        FROM ranked r
        JOIN `{TABLE}` e ON RTRIM(r.uuid, '_') = e.uuid
        WHERE r.rn = 1
        GROUP BY 1
        ORDER BY gmv DESC
    """
    rows = list(bq_client().query(q).result())
    nomes_iata = {v.upper(): k for k, v in IATA_NOMES.items()}
    agg: dict = {}
    for r in rows:
        raw = r.cia_raw or ""
        codigo = raw if len(raw) <= 3 else nomes_iata.get(raw, raw)
        codigo = IATA_NORMALIZACAO.get(codigo, codigo)
        if codigo not in agg:
            agg[codigo] = {"reservas": 0, "gmv": 0.0, "gmv_bilhete": 0.0}
        agg[codigo]["reservas"]    += int(r.reservas or 0)
        agg[codigo]["gmv"]         += float(r.gmv or 0)
        agg[codigo]["gmv_bilhete"] += float(r.gmv_bilhete or 0)
    records = [
        {
            "Sigla":          cod,
            "Companhia":      IATA_NOMES.get(cod, cod),
            "Reservas":       v["reservas"],
            "GMV":            round(v["gmv"], 2),
            "GMV Incentivo":  round(v["gmv_bilhete"], 2),
        }
        for cod, v in agg.items()
    ]
    df = pd.DataFrame(records) if records else pd.DataFrame(columns=["Sigla","Companhia","Reservas","GMV","GMV Incentivo"])
    if not df.empty:
        df = df.sort_values("GMV", ascending=False).reset_index(drop=True)
    return df


@st.cache_data(ttl=300, show_spinner=False)
def q_origem_emissoes(inicio: str, fim: str, variantes: tuple = ()) -> pd.DataFrame:
    """Distribuição de GMV por origem de emissão (consolidator_unified).
    Se variantes fornecido, filtra pela cia via JOIN com segmentos (s.company_operator)."""
    if variantes:
        vals = ", ".join(f"'{v}'" for v in variantes)
        q = f"""
            SELECT
                COALESCE(NULLIF(TRIM(consolidator_unified), ''), 'Não informado') AS Origem,
                COUNT(DISTINCT uuid)                                               AS Reservas,
                ROUND(SUM(total_amount_currency_brl), 2)                          AS GMV,
                ROUND(SUM(profit_currency_brl), 2)                                AS gross_revenue,
                ROUND(SUM(profit_currency_brl) /
                      NULLIF(SUM(total_amount_currency_brl), 0) * 100, 2)         AS take_rate
            FROM `{TABLE}`
            WHERE type = 'flight' AND status = 2
              AND created_at >= '{inicio}'
              AND created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
              AND uuid IN (
                  SELECT DISTINCT RTRIM(s.uuid, '_') FROM `{TABLE_SEG}` s
                  WHERE s.segment = 0 AND s.step = 1 AND UPPER(TRIM(s.company_operator)) IN ({vals})
              )
            GROUP BY 1
            ORDER BY GMV DESC
        """
    else:
        q = f"""
            SELECT
                COALESCE(NULLIF(TRIM(consolidator_unified), ''), 'Não informado') AS Origem,
                COUNT(*)                                                           AS Reservas,
                ROUND(SUM(total_amount_currency_brl), 2)                          AS GMV,
                ROUND(SUM(profit_currency_brl), 2)                                AS gross_revenue,
                ROUND(SUM(profit_currency_brl) /
                      NULLIF(SUM(total_amount_currency_brl), 0) * 100, 2)         AS take_rate
            FROM `{TABLE}`
            WHERE type = 'flight' AND status = 2
              AND created_at >= '{inicio}'
              AND created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
            GROUP BY 1
            ORDER BY GMV DESC
        """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {
            "Origem":        r.Origem,
            "Reservas":      int(r.Reservas or 0),
            "GMV":           float(r.GMV or 0),
            "Gross Revenue": float(r.gross_revenue or 0),
            "Take Rate (%)": float(r.take_rate or 0),
        }
        for r in rows
    ])


@st.cache_data(ttl=300, show_spinner=False)
def q_lista_cias() -> dict:
    """Retorna {codigo_iata: (variante1, variante2, ...)} usando silver_flight_orders.standard_airline.
    Cobre cias internacionais que não possuem entradas em TABLE_SEG (ex: Qatar via AMADEUS)."""
    q = f"""
        SELECT UPPER(TRIM(f.standard_airline)) AS cia
        FROM `{TABLE_FLIGHT_ORDERS}` f
        WHERE f.status = 2
          AND f.standard_airline IS NOT NULL AND TRIM(f.standard_airline) != ''
        GROUP BY 1
        HAVING SUM(f.onfly_amount_currency_brl) > 0
    """
    rows = list(bq_client().query(q).result())
    nomes_iata = {v.upper(): k for k, v in IATA_NOMES.items()}
    agrupado: dict[str, list[str]] = {}
    for r in rows:
        raw = r.cia
        # _SFO_NAME_TO_IATA cobre typos e variantes exclusivas do SFO
        codigo = _SFO_NAME_TO_IATA.get(raw)
        if codigo is None:
            codigo = raw if len(raw) == 2 else nomes_iata.get(raw, raw)
            codigo = IATA_NORMALIZACAO.get(codigo, codigo)
        agrupado.setdefault(codigo, []).append(raw)
    return {k: tuple(v) for k, v in sorted(agrupado.items())}

def _in_cias(variantes: tuple) -> str:
    vals = ", ".join(f"'{v}'" for v in variantes)
    return f"UPPER(TRIM(s.company_operator)) IN ({vals})"

def _in_cias_sfo(variantes: tuple, seg_mode: bool = False) -> str:
    """Filtro de CIA baseado em silver_flight_orders.standard_airline.
    seg_mode=True quando s é o alias de TABLE_SEG (queries de rotas/cabine)."""
    vals = ", ".join(f"'{v}'" for v in variantes)
    uuid_expr = (
        "REPLACE(RTRIM(s.uuid, '_'), '_flight', '')"
        if seg_mode else
        "REPLACE(e.uuid, '_flight', '')"
    )
    return (
        f"{uuid_expr} IN (\n"
        f"              SELECT DISTINCT f2.protocol FROM `{TABLE_FLIGHT_ORDERS}` f2\n"
        f"              WHERE UPPER(TRIM(f2.standard_airline)) IN ({vals}) AND f2.status = 2\n"
        f"          )"
    )

@st.cache_data(ttl=300, show_spinner=False)
def q_clientes_por_cia(variantes: tuple, inicio: str, fim: str) -> pd.DataFrame:
    """Clientes que emitiram na cia, com CNPJ e valores."""
    vals = ", ".join(f"'{v}'" for v in variantes)
    q = f"""
        SELECT
          c.company_id,
          c.name                                           AS cliente,
          c.social_name                                    AS razao_social,
          c.cnpj,
          COUNT(DISTINCT e.uuid)                           AS reservas,
          ROUND(SUM(e.total_amount_currency_brl), 2)       AS gmv,
          ROUND(AVG(e.total_amount_currency_brl), 2)       AS ticket_medio
        FROM `{TABLE}` e
        JOIN `{TABLE_COMP}` c ON e.company_id = c.company_id
        WHERE e.type = 'flight'
          AND e.status = 2
          AND DATE(e.created_at) BETWEEN '{inicio}' AND '{fim}'
          AND REPLACE(e.uuid, '_flight', '') IN (
              SELECT DISTINCT f2.protocol
              FROM `{TABLE_FLIGHT_ORDERS}` f2
              WHERE UPPER(TRIM(f2.standard_airline)) IN ({vals})
                AND f2.status = 2
          )
        GROUP BY 1, 2, 3, 4
        ORDER BY gmv DESC
    """
    try:
        rows = list(bq_client().query(q).result())
        return pd.DataFrame([{
            "company_id":   int(r.company_id),
            "Cliente":      r.cliente or "",
            "Razão Social": r.razao_social or "",
            "CNPJ":         r.cnpj or "",
            "Reservas":     int(r.reservas),
            "GMV":          float(r.gmv or 0),
            "Ticket Médio": float(r.ticket_medio or 0),
        } for r in rows])
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=3600, show_spinner=False)
def q_buscas_conversao_mensal(inicio: str, fim: str) -> pd.DataFrame:
    q = f"""
        WITH buscas AS (
          SELECT FORMAT_DATE('%Y-%m', DATE(created_at)) AS mes, COUNT(*) AS total_buscas
          FROM `dw-onfly-prd.travel_core.silver_flight_quotes`
          WHERE created_at >= '{inicio}' AND created_at <= '{fim}'
          GROUP BY 1
        ),
        emissoes AS (
          SELECT FORMAT_DATE('%Y-%m', partition_date) AS mes, COUNT(*) AS total_emissoes
          FROM `{TABLE_FLIGHT_ORDERS}`
          WHERE partition_date BETWEEN '{inicio}' AND '{fim}' AND status = 2
          GROUP BY 1
        )
        SELECT b.mes, b.total_buscas, IFNULL(e.total_emissoes,0) AS total_emissoes,
          ROUND(IFNULL(e.total_emissoes,0)*100.0/NULLIF(b.total_buscas,0),2) AS conversao_pct
        FROM buscas b LEFT JOIN emissoes e ON b.mes = e.mes
        ORDER BY 1
    """
    try:
        rows = list(bq_client().query(q).result())
        return pd.DataFrame([{
            "Mês": r.mes, "Buscas": int(r.total_buscas),
            "Emissões": int(r.total_emissoes), "Conversão (%)": float(r.conversao_pct or 0),
        } for r in rows])
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=3600, show_spinner=False)
def q_buscas_destinos(inicio: str, fim: str, internacional: bool, limit: int = 20) -> pd.DataFrame:
    flag = 1 if internacional else 0
    q = f"""
        SELECT IFNULL(destiny_city, destiny) AS destino,
               IFNULL(destiny_city_country, '') AS pais,
               COUNT(*) AS buscas,
               ROUND(COUNT(*)*100.0/SUM(COUNT(*)) OVER(),1) AS pct
        FROM `dw-onfly-prd.travel_core.silver_flight_quotes`
        WHERE created_at >= '{inicio}' AND created_at <= '{fim}'
          AND is_international = {flag}
          AND destiny IS NOT NULL
          {'AND LOWER(IFNULL(destiny_city_country,"")) NOT IN ("brazil","brasil","br") AND destiny NOT IN ("Guarulhos International Airport","Tancredo Neves International Airport","São Paulo","Sao Paulo","Rio de Janeiro","Belo Horizonte","Brasilia","Curitiba","Porto Alegre","Salvador","Fortaleza","Recife","Vitoria","Goiania","Manaus","Natal","Maceio","Belem","Florianopolis","Campinas","Navegantes")' if internacional else ''}
        GROUP BY 1, 2
        ORDER BY buscas DESC
        LIMIT {limit}
    """
    try:
        rows = list(bq_client().query(q).result())
        return pd.DataFrame([{
            "Destino": r.destino or "", "País": r.pais or "",
            "Buscas": int(r.buscas), "% do Total": float(r.pct or 0),
        } for r in rows])
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=3600, show_spinner=False)
def q_buscas_conversao_empresas(inicio: str, fim: str, min_buscas: int = 300) -> pd.DataFrame:
    q = f"""
        WITH buscas AS (
          SELECT company_id, COUNT(*) AS total_buscas
          FROM `dw-onfly-prd.travel_core.silver_flight_quotes`
          WHERE created_at >= '{inicio}' AND created_at <= '{fim}'
          GROUP BY 1 HAVING COUNT(*) >= {min_buscas}
        ),
        emissoes AS (
          SELECT company_id, MIN(company) AS empresa,
                 COUNT(*) AS total_emissoes,
                 ROUND(SUM(total_amount_currency_brl),2) AS gmv
          FROM `{TABLE_FLIGHT_ORDERS}`
          WHERE partition_date BETWEEN '{inicio}' AND '{fim}' AND status = 2
          GROUP BY 1
        )
        SELECT b.company_id,
               IFNULL(e.empresa, CAST(b.company_id AS STRING)) AS empresa,
               b.total_buscas,
               IFNULL(e.total_emissoes, 0) AS total_emissoes,
               ROUND(IFNULL(e.total_emissoes,0)*100.0/NULLIF(b.total_buscas,0),1) AS conversao_pct,
               IFNULL(e.gmv, 0) AS gmv
        FROM buscas b LEFT JOIN emissoes e ON b.company_id = e.company_id
        ORDER BY conversao_pct ASC
        LIMIT 30
    """
    try:
        rows = list(bq_client().query(q).result())
        return pd.DataFrame([{
            "ID": int(r.company_id), "Empresa": r.empresa or "",
            "Buscas": int(r.total_buscas), "Emissões": int(r.total_emissoes),
            "Conversão (%)": float(r.conversao_pct or 0), "GMV": float(r.gmv or 0),
        } for r in rows])
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=3600, show_spinner=False)
def q_variantes_cia_like(*patterns: str) -> tuple:
    """Retorna tuple de airline codes cujo nome contém qualquer dos padrões (case-insensitive)."""
    conditions = " OR ".join(f"UPPER(TRIM(s.company_operator)) LIKE '%{p.upper()}%'" for p in patterns)
    q = f"""
        SELECT DISTINCT UPPER(TRIM(s.company_operator)) AS cia
        FROM `{TABLE_SEG}` s
        JOIN `{TABLE}` e ON e.uuid = RTRIM(s.uuid, '_')
        WHERE e.type='flight' AND e.status=2
          AND s.company_operator IS NOT NULL AND TRIM(s.company_operator) != ''
          AND ({conditions})
    """
    rows = list(bq_client().query(q).result())
    return tuple(r.cia for r in rows) or patterns

@st.cache_data(ttl=300, show_spinner=False)
def q_tr_cia(inicio: str, fim: str, variantes: tuple) -> dict:
    q = f"""
        SELECT
            ROUND(SUM(total_amount_currency_brl), 2)  AS gmv,
            ROUND(SUM(profit_currency_brl), 2)        AS gross_revenue,
            ROUND(SUM(profit_currency_brl) / NULLIF(SUM(total_amount_currency_brl), 0) * 100, 2) AS take_rate,
            COUNT(DISTINCT uuid)                       AS trechos,
            ROUND(SUM(total_amount_currency_brl) / NULLIF(COUNT(DISTINCT uuid), 0), 2) AS ticket_medio
        FROM `{TABLE}` e
        WHERE type = 'flight' AND status = 2
          AND created_at >= '{inicio}'
          AND created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
          AND {_in_cias_sfo(variantes)}
    """
    row = list(bq_client().query(q).result())[0]
    return {
        "gmv":           float(row.gmv or 0),
        "gross_revenue": float(row.gross_revenue or 0),
        "take_rate":     float(row.take_rate or 0),
        "trechos":       int(row.trechos or 0),
        "ticket_medio":  float(row.ticket_medio or 0),
    }

# Mapeamento IATA canônico → standard_airline em silver_flight_orders (LATAM tem espaço, usar TRIM)
_IATA_TO_SFO_NAME = {
    "AD": "AZUL",  "G3": "GOL",
    "LA": "LATAM", "JJ": "LATAM", "LP": "LATAM", "XL": "LATAM", "4M": "LATAM",
    "T4": "LATAM", "P3": "LATAM", "M3": "LATAM",
}

@st.cache_data(ttl=300, show_spinner=False)
def q_onfly_liq_cia(inicio: str, fim: str, variantes: tuple) -> dict:
    """Passagem sem markup e sem taxas: onfly_amount − amount_taxes_v3_currency_brl via silver_flight_orders.
    Filtra diretamente por standard_airline no SFO — cobre cias internacionais sem TABLE_SEG."""
    sfo_vals = ", ".join(f"'{v}'" for v in variantes)
    sql = f"""
    SELECT
      ROUND(SUM(f.onfly_amount_currency_brl), 2)                                                     AS onfly_bruto,
      ROUND(SUM(IFNULL(f.amount_taxes_v3_currency_brl, 0)), 2)                                       AS taxas,
      ROUND(SUM(f.onfly_amount_currency_brl - IFNULL(f.amount_taxes_v3_currency_brl, 0)), 2)        AS onfly_liq
    FROM `{TABLE_FLIGHT_ORDERS}` f
    WHERE f.status = 2
      AND f.partition_date >= '{inicio}'
      AND f.partition_date < DATE_ADD('{fim}', INTERVAL 1 DAY)
      AND UPPER(TRIM(f.standard_airline)) IN ({sfo_vals})
    """
    row = list(bq_client().query(sql).result())[0]
    return {
        "onfly_bruto": float(row.onfly_bruto or 0),
        "taxas":       float(row.taxas or 0),
        "onfly_liq":   float(row.onfly_liq or 0),
    }

@st.cache_data(ttl=300, show_spinner=False)
def q_diario_cia(inicio: str, fim: str, variantes: tuple) -> pd.DataFrame:
    q = f"""
        SELECT
            DATE(created_at)                                AS dia,
            COUNT(DISTINCT uuid)                            AS qtd_reservas,
            ROUND(SUM(total_amount_currency_brl), 2)        AS gmv_dia
        FROM `{TABLE}` e
        WHERE type = 'flight' AND status = 2
          AND created_at >= '{inicio}'
          AND created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
          AND {_in_cias_sfo(variantes)}
        GROUP BY dia ORDER BY dia
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Data": str(r.dia), "Reservas": int(r.qtd_reservas), "GMV": float(r.gmv_dia or 0)}
        for r in rows
    ])

@st.cache_data(ttl=300, show_spinner=False)
def q_rotas_cia(inicio: str, fim: str, variantes: tuple) -> pd.DataFrame:
    q = f"""
        SELECT
            s.departure_airport_code      AS Origem,
            s.arrival_airport_code AS Destino,
            COUNT(DISTINCT e.uuid)                     AS Reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2) AS GMV
        FROM `{TABLE_SEG}` s
        JOIN `{TABLE}` e ON e.uuid = RTRIM(s.uuid, '_')
        WHERE e.type = 'flight' AND e.status = 2
          AND s.segment = 0 AND s.step = 1
          AND s.departure_airport_code != s.arrival_airport_code
          AND {_in_cias_sfo(variantes, seg_mode=True)}
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        GROUP BY 1, 2
        ORDER BY GMV DESC
        LIMIT 10
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Origem": r.Origem, "Destino": r.Destino,
         "Reservas": int(r.Reservas), "GMV": float(r.GMV or 0)}
        for r in rows
    ])

@st.cache_data(ttl=300, show_spinner=False)
def q_cabine_cia(inicio: str, fim: str, variantes: tuple) -> pd.DataFrame:
    q = f"""
        SELECT
            COALESCE(NULLIF(TRIM(f.fare_family), ''), 'Não informado') AS cabine,
            COUNT(DISTINCT e.uuid)                      AS qtd_reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)  AS gmv
        FROM `{TABLE}` e
        JOIN `{TABLE_FARE}` f
          ON CONCAT(SPLIT(f.uuid, '_flight_')[OFFSET(0)], '_flight') = e.uuid
        JOIN `{TABLE_SEG}` s ON RTRIM(s.uuid, '_') = e.uuid
        WHERE e.type = 'flight' AND e.status = 2
          AND s.segment = 0 AND s.step = 1
          AND {_in_cias_sfo(variantes, seg_mode=True)}
          AND f.type = 'flight'
          AND f.is_origin = 1
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        GROUP BY cabine ORDER BY gmv DESC LIMIT 15
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Cabine": r.cabine, "Reservas": int(r.qtd_reservas), "GMV": float(r.gmv or 0)}
        for r in rows
    ])

@st.cache_data(ttl=3600, show_spinner=False)
def q_gmv_anual_cia(variantes: tuple) -> pd.DataFrame:
    q = f"""
        SELECT
            CAST(EXTRACT(YEAR FROM created_at) AS INT64) AS ano,
            ROUND(SUM(total_amount_currency_brl), 2)     AS gmv,
            COUNT(DISTINCT uuid)                         AS trechos
        FROM `{TABLE}` e
        WHERE type = 'flight' AND status = 2
          AND {_in_cias_sfo(variantes)}
        GROUP BY ano ORDER BY ano
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {"Ano": int(r.ano), "GMV": float(r.gmv or 0), "Trechos": int(r.trechos or 0)}
        for r in rows
    ])


# ─── Mapeamento IATA → País ───────────────────────────────────────────────────
IATA_PAISES: dict[str, str] = {
    # Estados Unidos
    "JFK": "Estados Unidos", "MIA": "Estados Unidos", "ORD": "Estados Unidos",
    "ATL": "Estados Unidos", "IAH": "Estados Unidos", "DFW": "Estados Unidos",
    "MCO": "Estados Unidos", "LAX": "Estados Unidos", "EWR": "Estados Unidos",
    "SFO": "Estados Unidos", "LAS": "Estados Unidos", "BOS": "Estados Unidos",
    "FLL": "Estados Unidos", "IAD": "Estados Unidos", "DEN": "Estados Unidos",
    "AUS": "Estados Unidos", "SAN": "Estados Unidos", "PHX": "Estados Unidos",
    "SEA": "Estados Unidos", "MSP": "Estados Unidos", "PHL": "Estados Unidos",
    "CLT": "Estados Unidos", "DCA": "Estados Unidos", "BWI": "Estados Unidos",
    "SJU": "Estados Unidos", "HNL": "Estados Unidos",
    # México
    "MEX": "México", "GDL": "México", "MTY": "México", "CUN": "México",
    "QRO": "México", "MID": "México", "TQO": "México", "SJD": "México",
    "PVR": "México", "MZT": "México", "OAX": "México", "TLC": "México",
    # Portugal
    "LIS": "Portugal", "OPO": "Portugal", "FAO": "Portugal",
    # França
    "CDG": "França", "ORY": "França", "NCE": "França", "PAR": "França",
    "LYS": "França", "MRS": "França",
    # Alemanha
    "FRA": "Alemanha", "MUC": "Alemanha", "HAJ": "Alemanha", "STR": "Alemanha",
    "DUS": "Alemanha", "TXL": "Alemanha", "BER": "Alemanha", "HAM": "Alemanha",
    # Reino Unido
    "LHR": "Reino Unido", "LGW": "Reino Unido", "MAN": "Reino Unido",
    "EDI": "Reino Unido", "BHX": "Reino Unido",
    # Espanha
    "MAD": "Espanha", "BCN": "Espanha", "AGP": "Espanha", "PMI": "Espanha",
    "VLC": "Espanha", "SVQ": "Espanha", "BIO": "Espanha",
    # Itália
    "FCO": "Itália", "MXP": "Itália", "VCE": "Itália", "NAP": "Itália",
    "FLR": "Itália", "BLQ": "Itália",
    # Holanda
    "AMS": "Holanda",
    # Suíça
    "ZRH": "Suíça", "GVA": "Suíça",
    # Bélgica
    "BRU": "Bélgica",
    # Áustria
    "VIE": "Áustria",
    # Turquia
    "IST": "Turquia", "SAW": "Turquia", "AYT": "Turquia",
    # Emirados Árabes
    "DXB": "Emirados Árabes", "AUH": "Emirados Árabes", "SHJ": "Emirados Árabes",
    # Qatar
    "DOH": "Qatar",
    # Dinamarca
    "CPH": "Dinamarca",
    # Suécia
    "ARN": "Suécia", "GOT": "Suécia",
    # Finlândia
    "HEL": "Finlândia",
    # Irlanda
    "DUB": "Irlanda",
    # China
    "PEK": "China", "PVG": "China", "SZX": "China", "CAN": "China",
    "HGH": "China", "CTU": "China", "PKX": "China",
    # Hong Kong
    "HKG": "Hong Kong",
    # Japão
    "NRT": "Japão", "HND": "Japão", "KIX": "Japão", "NGO": "Japão",
    # Coreia do Sul
    "ICN": "Coreia do Sul", "GMP": "Coreia do Sul",
    # Singapura
    "SIN": "Singapura",
    # Tailândia
    "BKK": "Tailândia", "HKT": "Tailândia", "CNX": "Tailândia",
    # Indonésia
    "CGK": "Indonésia", "DPS": "Indonésia",
    # Índia
    "DEL": "Índia", "BOM": "Índia", "BLR": "Índia", "MAA": "Índia",
    # Austrália
    "SYD": "Austrália", "MEL": "Austrália", "BNE": "Austrália",
    # Nova Zelândia
    "AKL": "Nova Zelândia",
    # Canadá
    "YYZ": "Canadá", "YVR": "Canadá", "YUL": "Canadá", "YYC": "Canadá",
    # Argentina
    "AEP": "Argentina", "EZE": "Argentina", "MDZ": "Argentina", "COR": "Argentina",
    "BRC": "Argentina", "IGR": "Argentina", "MDP": "Argentina", "NQN": "Argentina",
    "SLA": "Argentina", "ROS": "Argentina", "TUC": "Argentina",
    # Chile
    "SCL": "Chile", "PMC": "Chile", "ANF": "Chile", "IQQ": "Chile",
    # Colômbia
    "BOG": "Colômbia", "MDE": "Colômbia", "CTG": "Colômbia", "CLO": "Colômbia",
    "BAQ": "Colômbia", "SMR": "Colômbia",
    # Peru
    "LIM": "Peru", "CUZ": "Peru",
    # Uruguai
    "MVD": "Uruguai", "PDP": "Uruguai",
    # Paraguai
    "ASU": "Paraguai", "CIO": "Paraguai",
    # Bolívia
    "VVI": "Bolívia", "LPB": "Bolívia", "CBB": "Bolívia",
    # Equador
    "GYE": "Equador", "UIO": "Equador",
    # Venezuela
    "CCS": "Venezuela",
    # Panamá
    "PTY": "Panamá",
    # Costa Rica
    "SJO": "Costa Rica",
    # Guatemala
    "GUA": "Guatemala",
    # República Dominicana
    "SDQ": "República Dominicana", "PUJ": "República Dominicana",
    # Cuba
    "HAV": "Cuba",
    # Jamaica
    "KIN": "Jamaica", "MBJ": "Jamaica",
    # Bahamas
    "NAS": "Bahamas",
    # Aruba
    "AUA": "Aruba",
    # Curaçao
    "CUR": "Curaçao",
    # África do Sul
    "JNB": "África do Sul", "CPT": "África do Sul",
    # Etiópia
    "ADD": "Etiópia",
    # Angola
    "LAD": "Angola",
    # Israel
    "TLV": "Israel",
}

# Índice inverso: país → lista de códigos IATA
PAISES_AEROPORTOS: dict[str, list[str]] = {}
for _ap, _pais in IATA_PAISES.items():
    PAISES_AEROPORTOS.setdefault(_pais, []).append(_ap)

# Mapeamento país → continente
PAIS_CONTINENTE: dict[str, str] = {
    # América do Norte
    "Estados Unidos": "🌎 América do Norte", "México": "🌎 América do Norte",
    "Canadá":         "🌎 América do Norte",
    # América Central e Caribe
    "Panamá":               "🌎 América Central e Caribe",
    "Costa Rica":           "🌎 América Central e Caribe",
    "Guatemala":            "🌎 América Central e Caribe",
    "República Dominicana": "🌎 América Central e Caribe",
    "Cuba":                 "🌎 América Central e Caribe",
    "Jamaica":              "🌎 América Central e Caribe",
    "Bahamas":              "🌎 América Central e Caribe",
    "Aruba":                "🌎 América Central e Caribe",
    "Curaçao":              "🌎 América Central e Caribe",
    # América do Sul
    "Argentina":  "🌎 América do Sul", "Chile":    "🌎 América do Sul",
    "Colômbia":   "🌎 América do Sul", "Peru":     "🌎 América do Sul",
    "Uruguai":    "🌎 América do Sul", "Paraguai": "🌎 América do Sul",
    "Bolívia":    "🌎 América do Sul", "Equador":  "🌎 América do Sul",
    "Venezuela":  "🌎 América do Sul",
    # Europa
    "Portugal":     "🌍 Europa", "França":      "🌍 Europa", "Alemanha":  "🌍 Europa",
    "Reino Unido":  "🌍 Europa", "Espanha":     "🌍 Europa", "Itália":    "🌍 Europa",
    "Holanda":      "🌍 Europa", "Suíça":       "🌍 Europa", "Bélgica":   "🌍 Europa",
    "Áustria":      "🌍 Europa", "Turquia":     "🌍 Europa", "Dinamarca": "🌍 Europa",
    "Suécia":       "🌍 Europa", "Finlândia":   "🌍 Europa", "Irlanda":   "🌍 Europa",
    # Oriente Médio
    "Emirados Árabes": "🌏 Oriente Médio", "Qatar": "🌏 Oriente Médio",
    "Israel":          "🌏 Oriente Médio",
    # Ásia
    "China":       "🌏 Ásia", "Hong Kong":   "🌏 Ásia", "Japão":        "🌏 Ásia",
    "Coreia do Sul": "🌏 Ásia", "Singapura": "🌏 Ásia", "Tailândia":    "🌏 Ásia",
    "Indonésia":   "🌏 Ásia", "Índia":       "🌏 Ásia",
    # Oceania
    "Austrália": "🌏 Oceania", "Nova Zelândia": "🌏 Oceania",
    # África
    "África do Sul": "🌍 África", "Etiópia": "🌍 África", "Angola": "🌍 África",
}

# Índice inverso: continente → lista de países
CONTINENTE_PAISES: dict[str, list[str]] = {}
for _p, _c in PAIS_CONTINENTE.items():
    if _p in PAISES_AEROPORTOS:
        CONTINENTE_PAISES.setdefault(_c, []).append(_p)
for _c in CONTINENTE_PAISES:
    CONTINENTE_PAISES[_c] = sorted(CONTINENTE_PAISES[_c])


@st.cache_data(ttl=300, show_spinner=False)
def q_potencial_pais(pais: str, inicio: str, fim: str) -> pd.DataFrame:
    return q_potencial_paises(tuple([pais]), inicio, fim)

@st.cache_data(ttl=300, show_spinner=False)
def q_potencial_paises(paises: tuple, inicio: str, fim: str) -> pd.DataFrame:
    aps = [a for p in paises for a in PAISES_AEROPORTOS.get(p, [])]
    if not aps:
        return pd.DataFrame()
    aps_sql = ", ".join(f"'{a}'" for a in aps)
    # ROW_NUMBER garante 1 segmento por uuid: o último trecho que chega no país destino
    # (voos com conexão interna no país — ex: GRU→MIA→JFK — teriam 2 segmentos no filtro)
    q = f"""
        WITH ranked AS (
            SELECT
                s.uuid,
                s.departure_airport_code                                       AS Origem,
                s.arrival_airport_code                                  AS Destino,
                UPPER(TRIM(s.company_operator))                         AS Cia,
                ROW_NUMBER() OVER (
                    PARTITION BY s.uuid
                    ORDER BY s.departure_date_hour DESC NULLS LAST
                )                                              AS rn
            FROM `{TABLE_SEG}` s
            JOIN `{TABLE}` e ON e.uuid = RTRIM(s.uuid, '_')
            WHERE e.type = 'flight'
              AND e.status = 2
              AND e.is_international = 1
              AND s.segment = 0 AND s.step = 1
              AND UPPER(TRIM(s.arrival_airport_code)) IN ({aps_sql})
              AND e.created_at >= '{inicio}'
              AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        )
        SELECT
            r.Origem,
            r.Destino,
            r.Cia,
            COUNT(DISTINCT r.uuid)                             AS Reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)         AS GMV,
            ROUND(AVG(e.total_amount_currency_brl), 2)         AS Ticket_Medio
        FROM ranked r
        JOIN `{TABLE}` e ON RTRIM(r.uuid, '_') = e.uuid
        WHERE r.rn = 1
        GROUP BY 1, 2, 3
        ORDER BY GMV DESC
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {
            "Origem":       r.Origem,
            "Destino":      r.Destino,
            "Cia":          r.Cia,
            "Reservas":     int(r.Reservas),
            "GMV":          float(r.GMV or 0),
            "Ticket Médio": float(r.Ticket_Medio or 0),
        }
        for r in rows
    ])

@st.cache_data(ttl=300, show_spinner=False)
def q_potencial_pais_saindo(pais: str, inicio: str, fim: str) -> pd.DataFrame:
    return q_potencial_paises_saindo(tuple([pais]), inicio, fim)

@st.cache_data(ttl=300, show_spinner=False)
def q_potencial_paises_saindo(paises: tuple, inicio: str, fim: str) -> pd.DataFrame:
    aps = [a for p in paises for a in PAISES_AEROPORTOS.get(p, [])]
    if not aps:
        return pd.DataFrame()
    aps_sql = ", ".join(f"'{a}'" for a in aps)
    # ROW_NUMBER garante 1 segmento por uuid: o primeiro trecho que sai do país origem
    # (voos com conexão interna no país — ex: GRU→VCP→MIA — teriam 2 segmentos no filtro)
    q = f"""
        WITH ranked AS (
            SELECT
                s.uuid,
                s.departure_airport_code                                       AS Origem,
                s.arrival_airport_code                                  AS Destino,
                UPPER(TRIM(s.company_operator))                         AS Cia,
                ROW_NUMBER() OVER (
                    PARTITION BY s.uuid
                    ORDER BY s.departure_date_hour ASC NULLS LAST
                )                                              AS rn
            FROM `{TABLE_SEG}` s
            JOIN `{TABLE}` e ON e.uuid = RTRIM(s.uuid, '_')
            WHERE e.type = 'flight'
              AND e.status = 2
              AND e.is_international = 1
              AND s.segment = 0 AND s.step = 1
              AND UPPER(TRIM(s.departure_airport_code)) IN ({aps_sql})
              AND e.created_at >= '{inicio}'
              AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
        )
        SELECT
            r.Origem,
            r.Destino,
            r.Cia,
            COUNT(DISTINCT r.uuid)                             AS Reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)         AS GMV,
            ROUND(AVG(e.total_amount_currency_brl), 2)         AS Ticket_Medio
        FROM ranked r
        JOIN `{TABLE}` e ON RTRIM(r.uuid, '_') = e.uuid
        WHERE r.rn = 1
        GROUP BY 1, 2, 3
        ORDER BY GMV DESC
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {
            "Origem":       r.Origem,
            "Destino":      r.Destino,
            "Cia":          r.Cia,
            "Reservas":     int(r.Reservas),
            "GMV":          float(r.GMV or 0),
            "Ticket Médio": float(r.Ticket_Medio or 0),
        }
        for r in rows
    ])


@st.cache_data(ttl=3600, show_spinner=False)
def q_lista_aeroportos() -> list[str]:
    q = f"""
        SELECT ap, SUM(gmv) AS total FROM (
            SELECT UPPER(TRIM(s.departure_airport_code))      AS ap,
                   SUM(e.total_amount_currency_brl) AS gmv
            FROM `{TABLE_SEG}` s
            JOIN `{TABLE}` e ON e.uuid = RTRIM(s.uuid, '_')
            WHERE e.type = 'flight' AND e.status = 2
            GROUP BY 1
            UNION ALL
            SELECT UPPER(TRIM(s.arrival_airport_code)) AS ap,
                   SUM(e.total_amount_currency_brl) AS gmv
            FROM `{TABLE_SEG}` s
            JOIN `{TABLE}` e ON e.uuid = RTRIM(s.uuid, '_')
            WHERE e.type = 'flight' AND e.status = 2
            GROUP BY 1
        )
        WHERE ap IS NOT NULL AND LENGTH(ap) = 3
        GROUP BY ap
        HAVING total > 0
        ORDER BY total DESC
    """
    rows = list(bq_client().query(q).result())
    return [r.ap for r in rows]

@st.cache_data(ttl=300, show_spinner=False)
def q_destinos_por_aeroporto(aeroporto: str, inicio: str, fim: str) -> pd.DataFrame:
    ap = aeroporto.upper().strip()
    q = f"""
        SELECT
            s.departure_airport_code                                           AS Origem,
            s.arrival_airport_code                                      AS Destino,
            COUNT(DISTINCT e.uuid)                             AS Reservas,
            ROUND(SUM(e.total_amount_currency_brl), 2)         AS GMV,
            STRING_AGG(DISTINCT UPPER(TRIM(s.company_operator)) ORDER BY UPPER(TRIM(s.company_operator))) AS Cias
        FROM `{TABLE_SEG}` s
        JOIN `{TABLE}` e ON e.uuid = RTRIM(s.uuid, '_')
        WHERE e.type = 'flight'
          AND e.status = 2
          AND s.departure_airport_code != s.arrival_airport_code
          AND e.created_at >= '{inicio}'
          AND e.created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
          AND (UPPER(TRIM(s.departure_airport_code)) = '{ap}' OR UPPER(TRIM(s.arrival_airport_code)) = '{ap}')
        GROUP BY 1, 2
        ORDER BY GMV DESC
    """
    rows = list(bq_client().query(q).result())
    return pd.DataFrame([
        {
            "Origem":   r.Origem,
            "Destino":  r.Destino,
            "Reservas": int(r.Reservas),
            "GMV":      float(r.GMV or 0),
            "Cias":     r.Cias or "—",
        }
        for r in rows
    ])


# ─── GMV de Incentivo (onfly_amount doméstico − taxas de embarque via CTE) ────
# Fórmula: SUM(sae.sae_onfly_amount) - SUM(IFNULL(consistency_case, 0))
# Dois CTEs: sae (silver_all_emissions, excl. México, todos os legs por protocolo)
# e sfo (silver_flight_orders, todos os legs por protocolo) para descontar taxas de embarque.
# Filtro de CIA via JOIN com TABLE_SEG (company_operator) — captura Amadeus + canal direto.

# Mapeamento nome da CIA → códigos IATA armazenados em company_operator no TABLE_SEG
_CIA_IATA_BILHETE = {
    "LATAM": ("LA", "LATAM", "JJ", "LP", "XL", "4M", "T4", "P3", "M3"),
    "AZUL":  ("AD", "AZUL"),
    "GOL":   ("G3", "GOL"),
}

def _cia_seg_join(cia: object) -> str:
    """Gera cláusula INNER JOIN com TABLE_SEG para filtrar por CIA operadora.
    Captura voos emitidos por qualquer canal (Amadeus, direto, etc.).
    Retorna string vazia quando cia=None (sem filtro de CIA)."""
    if not cia:
        return ""
    if isinstance(cia, (list, tuple)):
        codes: list = []
        for c in cia:
            codes.extend(_CIA_IATA_BILHETE.get(str(c).upper(), (str(c).upper(),)))
        iata_in = ", ".join(f"'{x}'" for x in dict.fromkeys(codes))
    else:
        codes = _CIA_IATA_BILHETE.get(str(cia).upper(), (str(cia).upper(),))
        iata_in = ", ".join(f"'{x}'" for x in codes)
    return f"""JOIN (
      SELECT DISTINCT REPLACE(RTRIM(uuid, '_'), '_flight', '') AS protocol
      FROM `{TABLE_SEG}` s
      WHERE s.segment = 0 AND s.step = 1
        AND UPPER(TRIM(s.company_operator)) IN ({iata_in})
    ) AS _seg ON _seg.protocol = sae.protocol"""


@st.cache_data(ttl=300, show_spinner=False)
def q_b_resumo(inicio, fim, cia=None):
    join_seg = _cia_seg_join(cia)
    sql = f"""
    WITH sae AS (
      SELECT
        REPLACE(uuid, '_flight', '')    AS protocol,
        SUM(total_amount_currency_brl)  AS sae_gmv,
        SUM(onfly_amount_currency_brl)  AS sae_onfly_amount
      FROM `{TABLE}`
      WHERE type = 'flight'
        AND status = 2
        AND (is_mexico_company = 0 OR is_mexico_company IS NULL)
        AND created_at >= '{inicio}'
        AND created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
      GROUP BY 1
    ),
    sfo AS (
      SELECT
        protocol,
        SUM(total_amount)    AS sfo_total,
        SUM(amount_taxes_v3) AS sfo_taxes
      FROM `{TABLE_FLIGHT_ORDERS}`
      WHERE status = 2
        AND partition_date >= '{inicio}'
        AND partition_date < DATE_ADD('{fim}', INTERVAL 1 DAY)
      GROUP BY 1
    )
    SELECT
      COUNT(*) AS qtd_reservas,
      ROUND(SUM(sae.sae_onfly_amount) - SUM(IFNULL(
        CASE WHEN ABS(sfo.sfo_total - sae.sae_gmv) < 1.0 THEN sfo.sfo_taxes ELSE NULL END, 0
      )), 2) AS gmv_total,
      ROUND((SUM(sae.sae_onfly_amount) - SUM(IFNULL(
        CASE WHEN ABS(sfo.sfo_total - sae.sae_gmv) < 1.0 THEN sfo.sfo_taxes ELSE NULL END, 0
      ))) / NULLIF(COUNT(*), 0), 2) AS ticket_medio
    FROM sae
    {join_seg}
    LEFT JOIN sfo ON sfo.protocol = sae.protocol
    """
    r = bq_client().query(sql).result()
    row = next(iter(r))
    return {"qtd_reservas": int(row.qtd_reservas or 0),
            "gmv_total":    float(row.gmv_total or 0),
            "ticket_medio": float(row.ticket_medio or 0)}

@st.cache_data(ttl=300, show_spinner=False)
def q_b_diario(inicio, fim, cia=None):
    join_seg = _cia_seg_join(cia)
    sql = f"""
    WITH sae AS (
      SELECT
        REPLACE(uuid, '_flight', '')    AS protocol,
        DATE(created_at)                AS dt,
        SUM(total_amount_currency_brl)  AS sae_gmv,
        SUM(onfly_amount_currency_brl)  AS sae_onfly_amount
      FROM `{TABLE}`
      WHERE type = 'flight'
        AND status = 2
        AND (is_mexico_company = 0 OR is_mexico_company IS NULL)
        AND created_at >= '{inicio}'
        AND created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
      GROUP BY 1, 2
    ),
    sfo AS (
      SELECT
        protocol,
        SUM(total_amount)    AS sfo_total,
        SUM(amount_taxes_v3) AS sfo_taxes
      FROM `{TABLE_FLIGHT_ORDERS}`
      WHERE status = 2
        AND partition_date >= '{inicio}'
        AND partition_date < DATE_ADD('{fim}', INTERVAL 1 DAY)
      GROUP BY 1
    )
    SELECT
      sae.dt                                                                          AS Data,
      COUNT(*)                                                                        AS Reservas,
      ROUND(SUM(sae.sae_onfly_amount) - SUM(IFNULL(
        CASE WHEN ABS(sfo.sfo_total - sae.sae_gmv) < 1.0 THEN sfo.sfo_taxes ELSE NULL END, 0
      )), 2)                                                                          AS GMV
    FROM sae
    {join_seg}
    LEFT JOIN sfo ON sfo.protocol = sae.protocol
    GROUP BY 1 ORDER BY 1
    """
    return bq_client().query(sql).to_dataframe()

@st.cache_data(ttl=300, show_spinner=False)
def q_b_por_cia(inicio, fim):
    sql = f"""
    WITH sae AS (
      SELECT
        REPLACE(uuid, '_flight', '')    AS protocol,
        consolidator_unified,
        SUM(total_amount_currency_brl)  AS sae_gmv,
        SUM(onfly_amount_currency_brl)  AS sae_onfly_amount
      FROM `{TABLE}`
      WHERE type = 'flight'
        AND status = 2
        AND (is_mexico_company = 0 OR is_mexico_company IS NULL)
        AND created_at >= '{inicio}'
        AND created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
      GROUP BY 1, 2
    ),
    sfo AS (
      SELECT
        protocol,
        SUM(total_amount)    AS sfo_total,
        SUM(amount_taxes_v3) AS sfo_taxes
      FROM `{TABLE_FLIGHT_ORDERS}`
      WHERE status = 2
        AND partition_date >= '{inicio}'
        AND partition_date < DATE_ADD('{fim}', INTERVAL 1 DAY)
      GROUP BY 1
    )
    SELECT
      sae.consolidator_unified                                                        AS Cia,
      COUNT(*)                                                                        AS Reservas,
      ROUND(SUM(sae.sae_onfly_amount) - SUM(IFNULL(
        CASE WHEN ABS(sfo.sfo_total - sae.sae_gmv) < 1.0 THEN sfo.sfo_taxes ELSE NULL END, 0
      )), 2)                                                                          AS GMV,
      ROUND((SUM(sae.sae_onfly_amount) - SUM(IFNULL(
        CASE WHEN ABS(sfo.sfo_total - sae.sae_gmv) < 1.0 THEN sfo.sfo_taxes ELSE NULL END, 0
      ))) / NULLIF(COUNT(*), 0), 2)                                                  AS ticket_medio
    FROM sae
    LEFT JOIN sfo ON sfo.protocol = sae.protocol
    GROUP BY 1 ORDER BY GMV DESC LIMIT 15
    """
    df = bq_client().query(sql).to_dataframe()
    df = df.rename(columns={"ticket_medio": "Ticket Médio"})
    return df

@st.cache_data(ttl=3600, show_spinner=False)
def q_b_anual(cia=None):
    join_seg = _cia_seg_join(cia)
    sql = f"""
    WITH sae AS (
      SELECT
        REPLACE(uuid, '_flight', '')         AS protocol,
        CAST(EXTRACT(YEAR FROM created_at) AS INT64) AS ano,
        SUM(total_amount_currency_brl)        AS sae_gmv,
        SUM(onfly_amount_currency_brl)        AS sae_onfly_amount
      FROM `{TABLE}`
      WHERE type = 'flight'
        AND status = 2
        AND (is_mexico_company = 0 OR is_mexico_company IS NULL)
      GROUP BY 1, 2
    ),
    sfo AS (
      SELECT
        protocol,
        SUM(total_amount)    AS sfo_total,
        SUM(amount_taxes_v3) AS sfo_taxes
      FROM `{TABLE_FLIGHT_ORDERS}`
      WHERE status = 2
      GROUP BY 1
    )
    SELECT
      sae.ano                                                                         AS Ano,
      ROUND(SUM(sae.sae_onfly_amount) - SUM(IFNULL(
        CASE WHEN ABS(sfo.sfo_total - sae.sae_gmv) < 1.0 THEN sfo.sfo_taxes ELSE NULL END, 0
      )), 2)                                                                          AS GMV,
      COUNT(*)                                                                        AS Reservas
    FROM sae
    {join_seg}
    LEFT JOIN sfo ON sfo.protocol = sae.protocol
    GROUP BY 1 ORDER BY 1
    """
    return bq_client().query(sql).to_dataframe()

@st.cache_data(ttl=3600, show_spinner=False)
def q_b_mensal(cia=None, ano: int = None):
    hoje = date.today()
    ano  = ano or hoje.year
    mes_limite = hoje.month if ano == hoje.year else 12
    join_seg = _cia_seg_join(cia)
    sql = f"""
    WITH sae AS (
      SELECT
        REPLACE(uuid, '_flight', '')              AS protocol,
        CAST(EXTRACT(MONTH FROM created_at) AS INT64) AS mes_num,
        SUM(total_amount_currency_brl)             AS sae_gmv,
        SUM(onfly_amount_currency_brl)             AS sae_onfly_amount
      FROM `{TABLE}`
      WHERE type = 'flight'
        AND status = 2
        AND (is_mexico_company = 0 OR is_mexico_company IS NULL)
        AND EXTRACT(YEAR FROM created_at) = {ano}
      GROUP BY 1, 2
    ),
    sfo AS (
      SELECT
        protocol,
        SUM(total_amount)    AS sfo_total,
        SUM(amount_taxes_v3) AS sfo_taxes
      FROM `{TABLE_FLIGHT_ORDERS}`
      WHERE status = 2
        AND EXTRACT(YEAR FROM partition_date) = {ano}
      GROUP BY 1
    )
    SELECT
      sae.mes_num,
      ROUND(SUM(sae.sae_onfly_amount) - SUM(IFNULL(
        CASE WHEN ABS(sfo.sfo_total - sae.sae_gmv) < 1.0 THEN sfo.sfo_taxes ELSE NULL END, 0
      )), 2)                                                                          AS GMV,
      COUNT(*)                                                                        AS Reservas
    FROM sae
    {join_seg}
    LEFT JOIN sfo ON sfo.protocol = sae.protocol
    GROUP BY 1 HAVING mes_num <= {mes_limite} ORDER BY 1
    """
    df = bq_client().query(sql).to_dataframe()
    df = df.rename(columns={"mes_num": "Mês"})
    return df


@st.cache_data(ttl=300, show_spinner=False)
def q_b_breakdown(inicio, fim, cia=None):
    """GMV de Incentivo separado por Nacional / Internacional, filtrado por CIA operadora."""
    join_seg = _cia_seg_join(cia)
    sql = f"""
    WITH sae AS (
      SELECT
        REPLACE(uuid, '_flight', '')    AS protocol,
        is_international,
        SUM(total_amount_currency_brl)  AS sae_gmv,
        SUM(onfly_amount_currency_brl)  AS sae_onfly_amount
      FROM `{TABLE}`
      WHERE type = 'flight'
        AND status = 2
        AND (is_mexico_company = 0 OR is_mexico_company IS NULL)
        AND created_at >= '{inicio}'
        AND created_at < DATE_ADD('{fim}', INTERVAL 1 DAY)
      GROUP BY 1, 2
    ),
    sfo AS (
      SELECT
        protocol,
        SUM(total_amount)    AS sfo_total,
        SUM(amount_taxes_v3) AS sfo_taxes
      FROM `{TABLE_FLIGHT_ORDERS}`
      WHERE status = 2
        AND partition_date >= '{inicio}'
        AND partition_date < DATE_ADD('{fim}', INTERVAL 1 DAY)
      GROUP BY 1
    )
    SELECT
      CASE WHEN sae.is_international = 1 THEN 'Internacional' ELSE 'Nacional' END AS tipo,
      COUNT(*)                                                                      AS reservas,
      ROUND(SUM(sae.sae_onfly_amount) - SUM(IFNULL(
        CASE WHEN ABS(sfo.sfo_total - sae.sae_gmv) < 1.0 THEN sfo.sfo_taxes ELSE NULL END, 0
      )), 2)                                                                        AS gmv_incentivo
    FROM sae
    {join_seg}
    LEFT JOIN sfo ON sfo.protocol = sae.protocol
    GROUP BY 1
    ORDER BY 1 DESC
    """
    return bq_client().query(sql).to_dataframe()


def render_bilhete_tab(cia, cor, cor_kpi1):
    """Idêntico a render_airline_tab mas usando queries de GMV de Incentivo.
    Usa company_operator (TABLE_SEG) para filtrar — captura Amadeus + canal direto."""
    # Variantes IATA para filtros de cabine/rotas/segmentos
    _var = tuple(_CIA_IATA_BILHETE.get(cia.upper(), (cia.upper(),)))
    with st.spinner(f"Carregando GMV de Incentivo {cia}..."):
        try:
            resumo    = q_b_resumo(i_str, f_str, cia)
            df_break  = q_b_breakdown(i_str, f_str, cia)
            df_day    = q_b_diario(i_str, f_str, cia)
            df_cab    = q_cabine_cia(i_str, f_str, _var)
            df_rot    = q_rotas_cia(i_str, f_str, _var)
            segs      = q_segmentos(i_str, f_str, cia)
            df_ano    = q_b_anual(cia)
            df_mensal = q_b_mensal(cia, ano=date.fromisoformat(i_str).year)
            comp      = q_b_resumo(ci_str, cf_str, cia) if ci_str else None
            segs_c    = q_segmentos(ci_str, cf_str, cia) if ci_str else None
            df_comp   = q_b_diario(ci_str, cf_str, cia) if ci_str else None
        except Exception as e:
            st.error(f"Erro ao conectar ao BigQuery: {e}")
            return
    render_kpis(resumo, comp, cor_kpi1, segmentos=segs, segmentos_comp=segs_c)

    # ── Nacional vs Internacional ─────────────────────────────────────────────
    if df_break is not None and not df_break.empty:
        st.markdown('<div class="sec-header-wrap"><p class="sec-header">Nacional vs Internacional</p></div>',
                    unsafe_allow_html=True)
        _bc1, _bc2 = st.columns(2)
        _cols = {"Nacional": _bc1, "Internacional": _bc2}
        _icons = {"Nacional": "🏠", "Internacional": "🌍"}
        for _, row in df_break.iterrows():
            _tipo = row["tipo"]
            _col  = _cols.get(_tipo, _bc2)
            with _col:
                st.markdown(
                    f'<div class="kpi-card">'
                    f'<p class="kpi-label">{_icons.get(_tipo, "")} {_tipo}</p>'
                    f'<p class="kpi-value">{brl(float(row["gmv_incentivo"]))}</p>'
                    f'<p style="font-size:0.75rem;color:#64748B;margin:2px 0 0 0;">'
                    f'{int(row["reservas"]):,} reservas</p>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    render_diario(df_day, cor, df_comp, comp_label)
    render_gmv_mensal(df_mensal, cor)
    render_cabine(df_cab, cor)
    render_rotas(df_rot)
    render_gmv_anual(df_ano, cor)


# ─── Tendência ───────────────────────────────────────────────────────────────
@st.cache_data(ttl=300, show_spinner=False)
def q_tendencia_diario(mes: int, ano: int) -> pd.DataFrame:
    from calendar import monthrange as _mr
    i = date(ano, mes, 1).strftime("%Y-%m-%d")
    f = date(ano, mes, _mr(ano, mes)[1]).strftime("%Y-%m-%d")
    sql = f"""
    SELECT
        DATE(created_at)                                                    AS dia,
        ROUND(SUM(total_amount_currency_brl), 2)                            AS gmv,
        ROUND(SUM(profit_currency_brl), 2)                                  AS gross_revenue,
        ROUND(SUM(profit_currency_brl) /
              NULLIF(SUM(total_amount_currency_brl), 0) * 100, 2)           AS take_rate
    FROM `{TABLE}`
    WHERE type = 'flight' AND status = 2

      AND created_at >= '{i}'
      AND created_at < DATE_ADD('{f}', INTERVAL 1 DAY)
    GROUP BY dia ORDER BY dia
    """
    df = bq_client().query(sql).to_dataframe()
    if not df.empty:
        df["dia"] = pd.to_datetime(df["dia"])
        df["dia_num"] = df["dia"].dt.day
    return df

@st.cache_data(ttl=300, show_spinner=False)
def q_todos_modais_diario(mes: int, ano: int) -> pd.DataFrame:
    """GMV diário por modal (flight, hotel, auto, bus) de silver_all_emissions."""
    from calendar import monthrange as _mr
    i = date(ano, mes, 1).strftime("%Y-%m-%d")
    f = date(ano, mes, _mr(ano, mes)[1]).strftime("%Y-%m-%d")
    sql = f"""
    SELECT
        DATE(created_at)                         AS dia,
        type                                     AS modal,
        ROUND(SUM(total_amount_currency_brl), 2) AS gmv
    FROM `dw-onfly-prd.travel_core.silver_all_emissions`
    WHERE status = 2
      AND created_at >= '{i}'
      AND created_at < DATE_ADD('{f}', INTERVAL 1 DAY)
    GROUP BY 1, 2
    ORDER BY 1, 2
    """
    df = bq_client().query(sql).to_dataframe()
    if not df.empty:
        df["dia"] = pd.to_datetime(df["dia"])
        df["dia_num"] = df["dia"].dt.day
    return df

@st.cache_data(ttl=3600, show_spinner=False)
def q_gmv_anual_todos_modais() -> pd.DataFrame:
    """GMV anual por modal de silver_all_emissions."""
    sql = """
    SELECT
        CAST(EXTRACT(YEAR FROM created_at) AS INT64) AS ano,
        type                                          AS modal,
        ROUND(SUM(total_amount_currency_brl), 2)      AS gmv,
        COUNT(*)                                      AS reservas
    FROM `dw-onfly-prd.travel_core.silver_all_emissions`
    WHERE status = 2
    GROUP BY 1, 2
    ORDER BY 1, 2
    """
    return bq_client().query(sql).to_dataframe()

@st.cache_data(ttl=300, show_spinner=False)
def q_tendencia_diario_cia(mes: int, ano: int, cia: str) -> pd.DataFrame:
    """Igual a q_tendencia_diario mas filtrado por consolidator_unified."""
    from calendar import monthrange as _mr
    i = date(ano, mes, 1).strftime("%Y-%m-%d")
    f = date(ano, mes, _mr(ano, mes)[1]).strftime("%Y-%m-%d")
    sql = f"""
    SELECT
        DATE(created_at)                                AS dia,
        ROUND(SUM(total_amount_currency_brl), 2)        AS gmv
    FROM `{TABLE}`
    WHERE type = 'flight' AND status = 2

      AND consolidator_unified = '{cia}'
      AND created_at >= '{i}'
      AND created_at < DATE_ADD('{f}', INTERVAL 1 DAY)
    GROUP BY dia ORDER BY dia
    """
    df = bq_client().query(sql).to_dataframe()
    if not df.empty:
        df["dia"] = pd.to_datetime(df["dia"])
        df["dia_num"] = df["dia"].dt.day
    return df


# ─── Análises com IA ─────────────────────────────────────────────────────────
_ANALISES_HISTORY_FILE = os.path.join(os.path.dirname(__file__), ".streamlit", "analises_history.json")

def _history_save(history: list):
    """Persiste histórico em JSON (sem o DataFrame completo — salva só as primeiras linhas)."""
    import json as _json
    records = []
    for e in history:
        df = e.get("df", pd.DataFrame())
        records.append({
            "pergunta": e.get("pergunta", ""),
            "sql":      e.get("sql", ""),
            "resumo":   e.get("resumo", ""),
            "analise":  e.get("analise", ""),
            "df_records":  df.head(200).to_dict(orient="records") if not df.empty else [],
            "df_columns":  list(df.columns) if not df.empty else [],
        })
    try:
        with open(_ANALISES_HISTORY_FILE, "w", encoding="utf-8") as _f:
            _json.dump(records, _f, ensure_ascii=False, default=str)
    except Exception:
        pass

def _history_load() -> list:
    """Carrega histórico do JSON."""
    import json as _json
    if not os.path.exists(_ANALISES_HISTORY_FILE):
        return []
    try:
        with open(_ANALISES_HISTORY_FILE, "r", encoding="utf-8") as _f:
            raw = _json.load(_f)
        history = []
        for e in raw:
            cols = e.get("df_columns", [])
            recs = e.get("df_records", [])
            df = pd.DataFrame(recs, columns=cols) if recs else pd.DataFrame()
            history.append({
                "pergunta": e.get("pergunta", ""),
                "sql":      e.get("sql", ""),
                "resumo":   e.get("resumo", ""),
                "analise":  e.get("analise", ""),
                "df":       df,
            })
        return history
    except Exception:
        return []


# ══════════════════════════════════════════════════════════════════════════════
# CRM Aéreo — queries
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(ttl=600, show_spinner=False)
def q_crm_airlines() -> list:
    q = f"""
        SELECT UPPER(TRIM(standard_airline)) AS a, COUNT(DISTINCT protocol) AS c
        FROM `{TABLE_FLIGHT_ORDERS}`
        WHERE partition_date >= DATE_SUB(CURRENT_DATE(), INTERVAL 12 MONTH)
          AND status = 2
          AND standard_airline IS NOT NULL AND TRIM(standard_airline) != ''
        GROUP BY 1
        HAVING SUM(onfly_amount_currency_brl) > 0
        ORDER BY c DESC
        LIMIT 50
    """
    try:
        rows = list(bq_client().query(q).result())
        return [{"a": r.a, "c": int(r.c)} for r in rows]
    except Exception:
        return []


@st.cache_data(ttl=300, show_spinner=False)
def q_crm_transacional(airline: str, from_date: str, to_date: str) -> dict:
    client = bq_client()
    a = airline.replace("'", "\\'")
    T = TABLE_FLIGHT_ORDERS
    w = (
        f"UPPER(TRIM(standard_airline)) = '{a}' AND status = 2 "
        f"AND partition_date BETWEEN '{from_date}' AND '{to_date}' "
        f"AND DATE(purchase_date) BETWEEN '{from_date}' AND '{to_date}'"
    )
    wDate = (
        f"status = 2 "
        f"AND partition_date BETWEEN '{from_date}' AND '{to_date}' "
        f"AND DATE(purchase_date) BETWEEN '{from_date}' AND '{to_date}'"
    )

    def run(sql):
        try:
            rows = list(client.query(sql).result())
            out = []
            for r in rows:
                d = {}
                for k in r.keys():
                    v = r[k]
                    if isinstance(v, bool):
                        d[k] = int(v)
                    elif isinstance(v, (int, float)):
                        d[k] = v
                    else:
                        d[k] = str(v) if v is not None else None
                out.append(d)
            return out
        except Exception:
            return []

    stats = run(f"""
        SELECT
          COUNT(DISTINCT protocol) AS tickets,
          ROUND(SUM(onfly_amount_currency_brl - IFNULL(amount_taxes_v3_currency_brl,0)),2) AS net,
          ROUND(AVG(onfly_amount_currency_brl - IFNULL(amount_taxes_v3_currency_brl,0)),2) AS avg_t,
          COUNTIF(is_international = TRUE)  AS intl,
          COUNTIF(is_international = FALSE OR is_international IS NULL) AS dom,
          COUNTIF(has_agreement = TRUE) AS agree,
          ROUND(SUM(total_amount_currency_brl),2) AS gmv,
          ROUND(SUM(IFNULL(amount_fares_v3_currency_brl, 0)),2) AS fare_base,
          ROUND(SUM(CASE WHEN is_international = FALSE OR is_international IS NULL
                    THEN total_amount_currency_brl ELSE 0 END),2) AS dom_gmv
        FROM `{T}` WHERE {w}
    """)

    routes = run(f"""
        SELECT
          CONCAT(TRIM(origin),'/',TRIM(destination)) AS rota,
          COUNT(DISTINCT protocol) AS t,
          ROUND(SUM(onfly_amount_currency_brl - IFNULL(amount_taxes_v3_currency_brl,0)),2) AS net,
          COUNTIF(is_international = TRUE) AS intl
        FROM `{T}` WHERE {w}
        GROUP BY 1 ORDER BY 2 DESC LIMIT 12
    """)

    trend = run(f"""
        SELECT
          FORMAT_DATE('%Y-%m', DATE(purchase_date)) AS mo,
          COUNT(DISTINCT protocol) AS t,
          ROUND(SUM(onfly_amount_currency_brl - IFNULL(amount_taxes_v3_currency_brl,0)),2) AS net
        FROM `{T}` WHERE {w}
        GROUP BY 1 ORDER BY 1
    """)

    channels = run(f"""
        SELECT
          COALESCE(TRIM(consolidator),'Desconhecido') AS consolidator,
          COUNT(DISTINCT protocol) AS t,
          ROUND(SUM(onfly_amount_currency_brl - IFNULL(amount_taxes_v3_currency_brl,0)),2) AS net
        FROM `{T}` WHERE {w}
        GROUP BY 1 ORDER BY 2 DESC
    """)

    totals = run(f"""
        SELECT
          ROUND(SUM(total_amount_currency_brl),2) AS total_mundo,
          ROUND(SUM(CASE WHEN is_international = FALSE OR is_international IS NULL
                    THEN total_amount_currency_brl ELSE 0 END),2) AS total_br
        FROM `{T}` WHERE {wDate}
    """)

    nationals = run(f"""
        SELECT
          UPPER(TRIM(standard_airline)) AS cia,
          ROUND(SUM(CASE WHEN is_international = FALSE OR is_international IS NULL
                    THEN total_amount_currency_brl ELSE 0 END),2) AS gmv_br
        FROM `{T}`
        WHERE UPPER(TRIM(standard_airline)) IN ('GOL','AZUL','LATAM') AND {wDate}
        GROUP BY 1
    """)

    return {
        "stats": stats, "routes": routes, "trend": trend,
        "channels": channels, "totals": totals, "nationals": nationals,
    }


@st.cache_data(ttl=300, show_spinner=False)
def q_crm_gmail(airline: str) -> list | None:
    """Busca e-mails sobre a cia via Gmail API OAuth. Retorna None se não configurado."""
    _needed = ("GMAIL_CLIENT_ID", "GMAIL_CLIENT_SECRET", "GMAIL_REFRESH_TOKEN")
    if not all(k in st.secrets for k in _needed):
        return None
    try:
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build

        creds = Credentials(
            token=None,
            refresh_token=st.secrets["GMAIL_REFRESH_TOKEN"],
            token_uri="https://oauth2.googleapis.com/token",
            client_id=st.secrets["GMAIL_CLIENT_ID"],
            client_secret=st.secrets["GMAIL_CLIENT_SECRET"],
            scopes=["https://www.googleapis.com/auth/gmail.readonly"],
        )
        service = build("gmail", "v1", credentials=creds, cache_discovery=False)

        query_str = airline.strip().split()[0].lower()
        result = service.users().threads().list(
            userId="me", q=query_str, maxResults=8
        ).execute()

        out = []
        for t in result.get("threads", []):
            detail = service.users().threads().get(
                userId="me", threadId=t["id"], format="metadata",
                metadataHeaders=["Subject", "From", "Date"],
            ).execute()
            msgs = detail.get("messages", [])
            if not msgs:
                continue
            last = msgs[-1]
            hdrs = {h["name"]: h["value"]
                    for h in last.get("payload", {}).get("headers", [])}
            out.append({
                "subject": hdrs.get("Subject", "(sem assunto)"),
                "from":    hdrs.get("From", ""),
                "date":    hdrs.get("Date", ""),
                "snippet": last.get("snippet", ""),
            })
        return out
    except Exception:
        return []


_ANALISES_SYSTEM_TPL = """\
Você é um analista de dados especializado em viagens corporativas da Onfly.
Responda sempre em português brasileiro.

Você deve gerar queries BigQuery Standard SQL usando EXATAMENTE os nomes de tabela abaixo (com crase e path completo):

1. `dw-onfly-prd.travel_core.silver_all_emissions`
   Campos principais:
   - uuid: identificador único da reserva
   - created_at (DATETIME, já em horário de Brasília): data/hora da emissão
   - type: tipo (sempre filtrar: type = 'flight')
   - status: status (sempre filtrar: status = 2)
   - total_amount_currency_brl (FLOAT): GMV total da reserva em BRL (tarifa + taxas)
   - onfly_amount_currency_brl (FLOAT): valor do bilhete em BRL — usado para GMV de Incentivo (descontando amount_taxes_v3 de silver_flight_orders via CTE)
   - profit_currency_brl (FLOAT): receita bruta (gross revenue) da Onfly em BRL
   - consolidator_unified (STRING): cia aérea / consolidador (ex: LATAM, AZUL, GOL, AMADEUS)
   - company_id (INTEGER): ID da empresa cliente

   Fórmulas importantes:
   - GMV = SUM(total_amount_currency_brl)
   - Gross Revenue = SUM(profit_currency_brl)
   - Take Rate (%) = SUM(profit_currency_brl) / NULLIF(SUM(total_amount_currency_brl), 0) * 100
   - GMV de Incentivo = SUM(sae_onfly_amount) - SUM(IFNULL(consistency_case, 0))  # taxas via CTE sfo (silver_flight_orders, todos os legs)
   - Ticket Médio = AVG(total_amount_currency_brl)

2. `dw-onfly-prd.travel_core.gold_item_summaries_flight_by_protocol_traveler_segment_leg`
   Campos: uuid (FK para tabela 1 via USING(uuid)), departure_airport_code (IATA origem),
   arrival_airport_code (IATA destino), segment (INTEGER, 0=ida/1=volta),
   step (INTEGER, 1=primeiro leg), company_operator (código IATA da cia),
   departure_date_hour (DATETIME), arrival_date_hour (DATETIME)
   — Para pegar apenas o primeiro trecho de saída: WHERE segment = 0 AND step = 1

3. `dw-onfly-prd.onfly_dim_shared.silver_companies`
   Campos: company_id (INTEGER), name (STRING), social_name (STRING), cnpj (STRING)
   Join com tabela 1: ON c.company_id = e.company_id

Contexto de datas:
- Hoje: {hoje}
- Ontem: {ontem}
- Anteontem: {anteontem}
- Período selecionado no dashboard: {i_str} a {f_str}
- O campo created_at é do tipo DATETIME (já em horário de Brasília). Para filtrar por dia use: DATE(created_at) = 'YYYY-MM-DD'

Regras obrigatórias:
- Sempre inclua: WHERE type = 'flight' AND status = 2
- Quando a pergunta mencionar datas relativas (hoje, ontem, anteontem, esta semana etc.), use as datas acima
- Se a pergunta não mencionar período específico, use o período selecionado no dashboard ({i_str} a {f_str})
- Para filtrar por dia: use DATE(created_at) = 'YYYY-MM-DD'
- Use APENAS os nomes de tabela exatos acima — nunca abrevie ou altere o path
- Use Standard SQL do BigQuery (sem legacy SQL)

ATENÇÃO — evitar duplicação de GMV:
- A tabela de emissões (tabela 1) tem UMA linha por reserva. O GMV correto é SUM(total_amount_currency_brl) NESSA tabela sem JOIN.
- A tabela de segmentos (tabela 2) tem VÁRIAS linhas por reserva (uma por trecho). Se você fizer JOIN entre as duas e depois SUM(total_amount_currency_brl), o GMV será multiplicado pelo número de trechos e ficará ERRADO.
- Regra: só faça JOIN com a tabela de segmentos quando precisar de campos exclusivos dela (departure_airport_code, arrival_airport_code, company_operator). Nesse caso, use COUNT(DISTINCT e.uuid) para contagens e agrupe apenas pelo que vier da tabela de segmentos.
- Para GMV, ticket médio e contagem de reservas: use APENAS a tabela de emissões, sem JOIN.

Retorne APENAS um JSON válido — sem markdown, sem texto extra:
{{"sql": "SELECT ...", "resumo": "O que a query busca em 1 linha"}}
"""


def _analises_query(pergunta: str, i_str: str, f_str: str) -> dict:
    """Interpreta pergunta via Claude, executa SQL no BQ e retorna análise."""
    import json as _json
    client = _get_anthropic_client()
    if client is None:
        raise RuntimeError("API key da Anthropic não configurada.")
    _hoje      = date.today()
    _ontem     = _hoje - timedelta(days=1)
    _anteontem = _hoje - timedelta(days=2)
    system = _ANALISES_SYSTEM_TPL.format(
        i_str=i_str, f_str=f_str,
        hoje=_hoje.strftime("%Y-%m-%d"),
        ontem=_ontem.strftime("%Y-%m-%d"),
        anteontem=_anteontem.strftime("%Y-%m-%d"),
    )
    # Passo 1: gerar SQL
    r1 = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=1024,
        system=system,
        messages=[{"role": "user", "content": pergunta}],
    )
    raw = r1.content[0].text.strip()
    # Remove possível markdown ```json ... ```
    if raw.startswith("```"):
        raw = re.sub(r"^```[a-z]*\n?", "", raw)
        raw = re.sub(r"\n?```$", "", raw)
    data = _json.loads(raw)
    sql, resumo = data["sql"], data.get("resumo", "")
    # Passo 2: executar no BigQuery
    df = bq_client().query(sql).to_dataframe()
    # Passo 3: interpretar resultados
    try:
        tabela_md = df.head(30).to_markdown(index=False) if not df.empty else "Sem resultados."
    except Exception:
        tabela_md = df.head(30).to_string(index=False) if not df.empty else "Sem resultados."
    r2 = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=1024,
        system="Você é analista de viagens corporativas da Onfly. Responda em português, de forma objetiva e clara.",
        messages=[{
            "role": "user",
            "content": (
                f"Pergunta do usuário: {pergunta}\n\n"
                f"Resultados da query:\n{tabela_md}\n\n"
                "Faça uma análise concisa dos resultados acima."
            ),
        }],
    )
    return {"pergunta": pergunta, "sql": sql, "resumo": resumo, "df": df, "analise": r2.content[0].text}



# ══════════════════════════════════════════════════════════════════════════════
# TELA INICIAL / SEÇÕES
# ══════════════════════════════════════════════════════════════════════════════
if secao is None:
    st.markdown("""
        <div style="display:flex;flex-direction:column;align-items:center;justify-content:center;
                    height:60vh;gap:16px;text-align:center;">
            <span style="font-size:3rem;">✈️</span>
            <p style="font-size:1.4rem;font-weight:700;color:#1D2939;margin:0;">
                Bem-vindo ao Onfly Dashboard - Sourcing Aéreas
            </p>
            <p style="font-size:0.95rem;color:#6B7280;margin:0;">
                Selecione um grupo no menu à esquerda para começar.
            </p>
        </div>
    """, unsafe_allow_html=True)

elif secao == "📊  GMV":
    tab_geral, tab_latam, tab_azul, tab_gol, tab_amadeus = st.tabs([
        "✈️  GMV Aéreo", "🔴  LATAM", "🔵  AZUL", "🟠  GOL", "🔷  Amadeus",
    ])

    with tab_geral:
        resumo = None
        with st.spinner("Consultando BigQuery..."):
            try:
                resumo         = q_resumo(i_str, f_str)
                df_cia         = q_por_cia(i_str, f_str)
                df_daily       = q_diario(i_str, f_str)
                df_onhappy     = q_onhappy_diario(i_str, f_str)
                df_rot         = q_rotas(i_str, f_str)
                segs           = q_segmentos(i_str, f_str)
                df_ano         = q_gmv_anual()
                df_mensal      = q_gmv_mensal(ano=date.fromisoformat(i_str).year)
                comp           = q_resumo(ci_str, cf_str) if ci_str else None
                segs_c         = q_segmentos(ci_str, cf_str) if ci_str else None
                df_comp        = q_diario(ci_str, cf_str) if ci_str else None
            except Exception as e:
                st.error(f"Erro ao conectar ao BigQuery: {e}")

        if resumo is not None:
            # ── Projeção do mês (dias úteis D-1 — exclui sáb/dom e dia atual) ────
            from calendar import monthrange as _mr
            import numpy as _np_g
            _hoje_g   = date.today()
            _df_tend  = q_tendencia_diario(_hoje_g.month, _hoje_g.year)
            if not _df_tend.empty:
                # Exclui o dia atual (D-1)
                _df_tend_d1 = _df_tend[_df_tend["dia_num"] < _hoje_g.day]
                _gmv_g    = float(_df_tend_d1["gmv"].sum())
                # Dias úteis decorridos até D-1
                _uteis_passados = sum(
                    1 for d in _df_tend_d1["dia_num"].unique()
                    if date(_hoje_g.year, _hoje_g.month, int(d)).weekday() < 5
                )
                # Total de dias úteis no mês
                _uteis_mes = sum(
                    1 for d in range(1, _mr(_hoje_g.year, _hoje_g.month)[1] + 1)
                    if date(_hoje_g.year, _hoje_g.month, d).weekday() < 5
                )
                _proj_gmv = (_gmv_g / _uteis_passados * _uteis_mes) if _uteis_passados > 0 else None
            else:
                _proj_gmv = None

            # ── Projeção anual com sazonalidade do ano anterior ───────────────────
            # Ratio = GMV atual acumulado ÷ GMV mesmo período ano anterior
            # Projeção = GMV total ano anterior × ratio
            _ano_ant_g   = _hoje_g.year - 1
            _mes_g       = _hoje_g.month
            _dias_mes_g  = _hoje_g.day
            _tot_dias_g  = _mr(_hoje_g.year, _mes_g)[1]
            _gmv_prev    = q_gmv_mensal_ano(_ano_ant_g)
            _gmv_curr_m  = q_gmv_mensal_ano(_hoje_g.year)

            # Mesmo período no ano anterior (meses completos + mês atual pro-rated)
            _gmv_prev_periodo = (
                sum(_gmv_prev.get(m, 0) for m in range(1, _mes_g)) +
                _gmv_prev.get(_mes_g, 0) * (_dias_mes_g / _tot_dias_g)
            )
            _gmv_curr_acum = sum(_gmv_curr_m.get(m, 0) for m in range(1, _mes_g + 1))
            _ratio_g       = (_gmv_curr_acum / _gmv_prev_periodo) if _gmv_prev_periodo > 0 else 1.0
            _gmv_prev_ano  = sum(_gmv_prev.values())
            _proj_gmv_anual = _gmv_prev_ano * _ratio_g if _gmv_prev_ano > 0 else None

            render_kpis(resumo, comp, segmentos=segs, segmentos_comp=segs_c)

            # ── Card ONHAPPY ──────────────────────────────────────────────────
            if df_onhappy is not None and not df_onhappy.empty:
                _gmv_onhappy  = float(df_onhappy["GMV"].sum())
                _gmv_total_p  = float(resumo.get("gmv_total", 0))
                _pct_onhappy  = (_gmv_onhappy / _gmv_total_p * 100) if _gmv_total_p else 0
                _col_oh, _    = st.columns([1, 3])
                with _col_oh:
                    st.markdown(
                        f'<div class="kpi-card" style="border-top-color:#F59E0B;">'
                        f'<p class="kpi-label">🏢 ONHAPPY</p>'
                        f'<p class="kpi-value">{brl(_gmv_onhappy)}</p>'
                        f'<p style="font-size:0.75rem;color:#64748B;margin:2px 0 0 0;">'
                        f'{_pct_onhappy:.1f}% do GMV total</p>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

            render_diario(df_daily, ONFLY_BLUE, df_comp, comp_label,
                          df_extra=df_onhappy, extra_label="ONHAPPY", extra_cor="#F59E0B")
            render_gmv_mensal(df_mensal, ONFLY_BLUE, proj_gmv=_proj_gmv)

            st.markdown('<div class="sec-header-wrap"><p class="sec-header">GMV por Cia Aérea</p></div>', unsafe_allow_html=True)
            if not df_cia.empty:
                col_chart, col_table = st.columns([1, 1], gap="large")
                gmv_total = resumo["gmv_total"]
                with col_chart:
                    fig_bar = px.bar(
                        df_cia.head(10).sort_values("GMV"),
                        x="GMV", y="Cia", orientation="h",
                        color="GMV",
                        color_continuous_scale=[[0, "#93C5FD"], [1, ONFLY_BLUE]],
                        labels={"GMV": "GMV (R$)", "Cia": ""},
                    )
                    fig_bar.update_coloraxes(showscale=False)
                    fig_bar.update_xaxes(tickprefix="R$ ")
                    st.plotly_chart(plotly_layout(fig_bar, 380), use_container_width=True)
                with col_table:
                    df_show = df_cia.copy()
                    df_show["% GMV"] = (df_show["GMV"] / gmv_total * 100).round(1)
                    st.dataframe(
                        _brl_df(df_show[["Cia", "Reservas", "GMV", "Ticket Médio", "% GMV"]]),
                        use_container_width=True, hide_index=True, height=380,
                        column_config={
                            "GMV":          st.column_config.TextColumn("GMV"),
                            "Ticket Médio": st.column_config.TextColumn("Ticket Médio"),
                            "Reservas":     st.column_config.NumberColumn("Reservas",     format="%d"),
                            "% GMV":        st.column_config.NumberColumn("% GMV",        format="%.1f%%"),
                        },
                    )
            render_rotas(df_rot)
            render_gmv_anual(df_ano, ONFLY_BLUE, proj_gmv_anual=_proj_gmv_anual)

    # Carrega variantes (airline codes) para cada cia — usa cache, sem query extra
    _lista_cias_gmv = q_lista_cias()
    # Códigos LATAM confirmados no banco: LA (principal), LATAM, JJ (ex-TAM Brasil),
    # LP (LATAM Peru), XL (LATAM Equador), 4M (LATAM Argentina), T4, P3, M3
    _var_latam = ("LA", "LATAM", "JJ", "LP", "XL", "4M", "T4", "P3", "M3")
    _var_azul  = _lista_cias_gmv.get("AD", ("AZUL",))
    _var_gol   = _lista_cias_gmv.get("G3", ("GOL",))

    with tab_latam:
        render_airline_tab("LATAM", COR_LATAM, "red", variantes=_var_latam)

    with tab_azul:
        render_airline_tab("AZUL", COR_AZUL, "azul", variantes=_var_azul)

    with tab_gol:
        render_airline_tab("GOL", COR_GOL, "gol", variantes=_var_gol)

    with tab_amadeus:
        resumo = None
        with st.spinner("Carregando dados Amadeus..."):
            try:
                resumo  = q_resumo(i_str, f_str, AMADEUS_CONSOLIDATORS)
                df_day  = q_diario(i_str, f_str, AMADEUS_CONSOLIDATORS)
                df_cab  = q_cabine(i_str, f_str, "AMADEUS")
                df_rot  = q_rotas(i_str, f_str, "AMADEUS")
                df_cias = q_cias_amadeus(i_str, f_str)
                segs      = q_segmentos(i_str, f_str, "AMADEUS")
                df_ano    = q_gmv_anual(AMADEUS_CONSOLIDATORS)
                df_mensal = q_gmv_mensal(AMADEUS_CONSOLIDATORS, ano=date.fromisoformat(i_str).year)
                comp      = q_resumo(ci_str, cf_str, AMADEUS_CONSOLIDATORS) if ci_str else None
                segs_c    = q_segmentos(ci_str, cf_str, "AMADEUS") if ci_str else None
                df_comp   = q_diario(ci_str, cf_str, AMADEUS_CONSOLIDATORS) if ci_str else None
            except Exception as e:
                st.error(f"Erro ao conectar ao BigQuery: {e}")

        if resumo is not None:
            render_kpis(resumo, comp, "amadeus", "orange", "green", segmentos=segs, segmentos_comp=segs_c)
            render_diario(df_day, COR_AMADEUS, df_comp, comp_label)
            render_gmv_mensal(df_mensal, COR_AMADEUS, proj_gmv=_proj_mensal_de_df(df_mensal))

            st.markdown('<div class="sec-header-wrap"><p class="sec-header">Cias Aéreas com Emissões via Amadeus</p></div>', unsafe_allow_html=True)
            if not df_cias.empty:
                gmv_total = df_cias["GMV"].sum()
                col_chart, col_table = st.columns([1, 1], gap="large")
                with col_chart:
                    fig_bar = px.bar(
                        df_cias.head(12).sort_values("GMV"),
                        x="GMV", y="Sigla", orientation="h",
                        color="GMV",
                        color_continuous_scale=[[0, "#BAE6FD"], [1, COR_AMADEUS]],
                        labels={"GMV": "GMV (R$)", "Sigla": ""},
                    )
                    fig_bar.update_coloraxes(showscale=False)
                    fig_bar.update_xaxes(tickprefix="R$ ")
                    st.plotly_chart(plotly_layout(fig_bar, 380), use_container_width=True)
                with col_table:
                    df_show = df_cias.copy()
                    df_show["% GMV"] = (df_show["GMV"] / gmv_total * 100).round(1)
                    st.dataframe(
                        _brl_df(df_show[["Sigla", "Companhia", "Reservas", "GMV", "Ticket Médio", "% GMV"]]),
                        use_container_width=True, hide_index=True, height=380,
                        column_config={
                            "GMV":          st.column_config.TextColumn("GMV"),
                            "Ticket Médio": st.column_config.TextColumn("Ticket Médio"),
                            "Reservas":     st.column_config.NumberColumn("Reservas",     format="%d"),
                            "% GMV":        st.column_config.NumberColumn("% GMV",        format="%.1f%%"),
                        },
                    )
            render_cabine(df_cab, COR_AMADEUS)
            render_rotas(df_rot)
            render_gmv_anual(df_ano, COR_AMADEUS)

    # ── GMV Total Ano a Ano — Todos os Modais ────────────────────────────────
    st.markdown("---")
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">Evolução do GMV Ano a Ano — Todos os Modais</p></div>', unsafe_allow_html=True)
    with st.spinner("Carregando dados multimodal..."):
        try:
            from calendar import monthrange as _mr_mm
            _hoje_mm   = date.today()
            _df_mm_ano = q_gmv_anual_todos_modais()
            _df_mm_mes = q_todos_modais_diario(_hoje_mm.month, _hoje_mm.year)
        except Exception as _e_mm:
            _df_mm_ano = pd.DataFrame()
            _df_mm_mes = pd.DataFrame()
            st.error(f"Erro ao carregar dados multimodal: {_e_mm}")

    if not _df_mm_ano.empty:
        _MODAL_LABELS = {"flight": "Aéreo", "hotel": "Hotel", "auto": "Carro", "bus": "Ônibus", "generic": "Outros"}
        _MODAL_CORES  = {"flight": ONFLY_BLUE, "hotel": "#10B981", "auto": "#F59E0B", "bus": "#8B5CF6", "generic": "#94A3B8"}
        _MODAIS_ORDER = ["flight", "hotel", "auto", "bus", "generic"]

        # Projeção do ano atual (D-1, dias úteis)
        _proj_ano_mm = 0.0
        if not _df_mm_mes.empty:
            _df_mm_d1 = _df_mm_mes[_df_mm_mes["dia_num"] < _hoje_mm.day]
            _gmv_mm_real = float(_df_mm_d1["gmv"].sum())
            _uteis_passados_mm = sum(
                1 for d in _df_mm_d1["dia_num"].unique()
                if date(_hoje_mm.year, _hoje_mm.month, int(d)).weekday() < 5
            )
            _uteis_mes_mm = sum(
                1 for d in range(1, _mr_mm(_hoje_mm.year, _hoje_mm.month)[1] + 1)
                if date(_hoje_mm.year, _hoje_mm.month, d).weekday() < 5
            )
            _proj_mes_mm = (_gmv_mm_real / _uteis_passados_mm * _uteis_mes_mm) if _uteis_passados_mm > 0 else 0
            # Extrapola para o ano com base na proporção do mês atual vs ano anterior
            _gmv_ano_ant_mm = float(_df_mm_ano[_df_mm_ano["ano"] == _hoje_mm.year - 1]["gmv"].sum()) if not _df_mm_ano.empty else 0
            _gmv_per_mes_ant = _gmv_ano_ant_mm / 12 if _gmv_ano_ant_mm > 0 else 0
            _gmv_ano_real_mm = float(_df_mm_ano[_df_mm_ano["ano"] == _hoje_mm.year]["gmv"].sum())
            _ratio_mm = (_gmv_ano_real_mm / (_gmv_per_mes_ant * _hoje_mm.month)) if _gmv_per_mes_ant > 0 else 1
            _proj_ano_mm = _gmv_ano_ant_mm * _ratio_mm if _gmv_ano_ant_mm > 0 else 0

        # Pivot: anos × modais
        _df_pivot_ano = _df_mm_ano.pivot_table(index="ano", columns="modal", values="gmv", aggfunc="sum", fill_value=0).reset_index()
        _df_pivot_ano.columns = [str(c) for c in _df_pivot_ano.columns]
        _modais_ok = [m for m in _MODAIS_ORDER if m in _df_pivot_ano.columns]
        _anos_str  = _df_pivot_ano["ano"].astype(str).tolist()
        _ano_atual_str = str(_hoje_mm.year)

        # Totais por ano (para o label no topo)
        _totais_ano = _df_pivot_ano[_modais_ok].sum(axis=1).tolist()

        col_chart_mm, col_table_mm = st.columns([3, 2], gap="large")
        with col_chart_mm:
            _fig_mm = go.Figure()
            for _m in _modais_ok:
                _fig_mm.add_trace(go.Bar(
                    name=_MODAL_LABELS.get(_m, _m),
                    x=_anos_str,
                    y=_df_pivot_ano[_m],
                    marker_color=_MODAL_CORES.get(_m, "#94A3B8"),
                    hovertemplate=f"<b>%{{x}}</b><br>{_MODAL_LABELS.get(_m, _m)}: R$ %{{y:,.0f}}<extra></extra>",
                ))
            # Projeção empilhada no ano atual
            _mask_ano_at = _df_pivot_ano["ano"].astype(int) == _hoje_mm.year
            _gmv_atual_total = float(_df_pivot_ano[_mask_ano_at][_modais_ok].sum(axis=1).iloc[0]) if _mask_ano_at.any() else 0
            _proj_resto_mm = max(0.0, _proj_ano_mm - _gmv_atual_total)
            if _proj_resto_mm > 0:
                _proj_vals_mm = [_proj_resto_mm if a == _ano_atual_str else 0 for a in _anos_str]
                _fig_mm.add_trace(go.Bar(
                    name="Projeção",
                    x=_anos_str, y=_proj_vals_mm,
                    marker_color=hex_to_rgba("#F59E0B", 0.55),
                    hovertemplate="<b>%{x}</b><br>Projeção restante: R$ %{y:,.0f}<extra></extra>",
                ))
            # Totais no topo (sem R$, sem decimais, em milhões)
            def _fmt_m(v): return f"{v/1_000_000:.0f}M"
            _labels_topo = []
            for _a, _t in zip(_anos_str, _totais_ano):
                if _a == _ano_atual_str and _proj_ano_mm > 0:
                    _labels_topo.append(f"<b>{_fmt_m(_proj_ano_mm)}*</b>")
                else:
                    _labels_topo.append(f"<b>{_fmt_m(_t)}</b>")
            _fig_mm.add_trace(go.Scatter(
                x=_anos_str,
                y=[_t + _proj_resto_mm if _a == _ano_atual_str else _t for _a, _t in zip(_anos_str, _totais_ano)],
                mode="text",
                text=_labels_topo,
                textposition="top center",
                textfont=dict(size=11, color="#334155"),
                showlegend=False,
                hoverinfo="skip",
            ))
            _fig_mm.update_layout(
                barmode="stack",
                showlegend=True,
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0, font=dict(size=12)),
            )
            _fig_mm.update_yaxes(tickprefix="R$ ")
            _fig_mm.update_xaxes(type="category")
            st.plotly_chart(plotly_layout(_fig_mm, 320), use_container_width=True)

        with col_table_mm:
            _resumo_anos = _df_mm_ano.groupby("ano").agg(gmv=("gmv", "sum"), reservas=("reservas", "sum")).reset_index()
            _resumo_anos = _resumo_anos.sort_values("ano")
            _resumo_anos["Var. YoY"] = _resumo_anos["gmv"].pct_change() * 100
            _resumo_anos["GMV"] = _resumo_anos.apply(
                lambda r: f"{brl(r['gmv'])} *" if r["ano"] == _hoje_mm.year and _proj_ano_mm > 0 else brl(r["gmv"]), axis=1
            )
            st.dataframe(
                _resumo_anos[["ano", "reservas", "GMV", "Var. YoY"]].rename(columns={"ano": "Ano", "reservas": "Reservas"}),
                use_container_width=True, hide_index=True, height=300,
                column_config={
                    "Ano":      st.column_config.NumberColumn("Ano",      format="%d"),
                    "GMV":      st.column_config.TextColumn("GMV"),
                    "Reservas": st.column_config.NumberColumn("Reservas", format="%d"),
                    "Var. YoY": st.column_config.NumberColumn("Var. YoY", format="%.1f%%"),
                },
            )
            if _proj_ano_mm > 0:
                st.caption("* Projeção baseada na sazonalidade do ano anterior")

# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO: GMV DE INCENTIVO
# ══════════════════════════════════════════════════════════════════════════════
elif secao == "🎫  GMV de Incentivo":
    st.markdown(
        "<p style='color:#6B7280;font-size:0.82rem;margin-bottom:8px;'>"
        "GMV de Incentivo — onfly_amount doméstico (excl. México) descontadas as taxas de embarque, "
        "base para cálculo de incentivo das cias aéreas.</p>",
        unsafe_allow_html=True,
    )
    tab_latam, tab_azul, tab_gol = st.tabs([
        "🔴  LATAM", "🔵  AZUL", "🟠  GOL",
    ])

    with tab_latam:
        render_bilhete_tab("LATAM", COR_LATAM, "red")

    with tab_azul:
        render_bilhete_tab("AZUL", COR_AZUL, "azul")

    with tab_gol:
        render_bilhete_tab("GOL", COR_GOL, "gol")

# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO: TAKE RATE
# ══════════════════════════════════════════════════════════════════════════════
elif secao == "💰  Take Rate":
    META_TR = 10.7  # % meta anual Onfly 2026 — aéreo

    with st.spinner("Consultando BigQuery..."):
        try:
            tr          = q_tr_resumo(i_str, f_str)
            df_daily    = q_tr_diario(i_str, f_str)
            df_cia      = q_tr_por_cia(i_str, f_str)
            df_negativos= q_tr_negativos(i_str, f_str)
            tr_comp     = q_tr_resumo(ci_str, cf_str) if ci_str else None
        except Exception as e:
            st.error(f"Erro ao conectar ao BigQuery: {e}")
            st.stop()

    # ── KPI Cards ──
    meta_diff = tr["take_rate"] - META_TR
    meta_cor  = "#16A34A" if meta_diff >= 0 else "#DC2626"
    meta_icon = "▲" if meta_diff >= 0 else "▼"

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f"""
            <div class="kpi-card" style="border-top-color:#8B5CF6;">
                <p class="kpi-label">Take Rate Aéreo</p>
                <p class="kpi-value">{tr["take_rate"]:.2f}%</p>
                <p style="font-size:0.75rem;color:{meta_cor};font-weight:600;margin:2px 0 4px 0;">
                    {meta_icon} {abs(meta_diff):.2f}pp vs meta ({META_TR}%)
                </p>
                {delta_html(tr["take_rate"], tr_comp["take_rate"] if tr_comp else None)}
            </div>
        """, unsafe_allow_html=True)
    with c2:
        st.markdown(f"""
            <div class="kpi-card orange">
                <p class="kpi-label">Gross Revenue Aéreo</p>
                <p class="kpi-value">{brl(tr["gross_revenue"])}</p>
                {delta_html(tr["gross_revenue"], tr_comp["gross_revenue"] if tr_comp else None)}
            </div>
        """, unsafe_allow_html=True)
    with c3:
        st.markdown(f"""
            <div class="kpi-card green">
                <p class="kpi-label">GMV Aéreo</p>
                <p class="kpi-value">{brl(tr["gmv"])}</p>
                {delta_html(tr["gmv"], tr_comp["gmv"] if tr_comp else None)}
            </div>
        """, unsafe_allow_html=True)

    # ── Evolução diária ──
    st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">Evolução Diária do Take Rate</p></div>', unsafe_allow_html=True)
    if not df_daily.empty:
        fig_tr = go.Figure()
        fig_tr.add_hline(y=META_TR, line_dash="dot", line_color="#8B5CF6",
                         annotation_text=f"Meta {META_TR}%", annotation_position="top right")
        fig_tr.add_trace(go.Scatter(
            x=df_daily["Data"], y=df_daily["Take Rate (%)"],
            mode="lines+markers", name="Take Rate",
            line=dict(color="#8B5CF6", width=2.5),
            marker=dict(size=5, color="#8B5CF6"),
            fill="tozeroy", fillcolor="rgba(139,92,246,0.07)",
            hovertemplate="<b>%{x}</b><br>Take Rate: %{y:.2f}%<extra></extra>",
        ))
        fig_tr.update_yaxes(ticksuffix="%")
        st.plotly_chart(plotly_layout(fig_tr, 300), use_container_width=True)

    # ── Por Cia Aérea ──
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">Take Rate por Cia Aérea</p></div>', unsafe_allow_html=True)
    col_chart, col_table = st.columns([1, 1], gap="large")
    with col_chart:
        fig_cia = px.bar(
            df_cia.head(12).sort_values("Take Rate (%)"),
            x="Take Rate (%)", y="Cia", orientation="h",
            color="Take Rate (%)",
            color_continuous_scale=[[0, "#DDD6FE"], [1, "#7C3AED"]],
            text="Take Rate (%)",
            labels={"Take Rate (%)": "Take Rate (%)", "Cia": ""},
        )
        fig_cia.update_traces(texttemplate="%{text:.2f}%", textposition="outside")
        fig_cia.update_coloraxes(showscale=False)
        fig_cia.update_xaxes(ticksuffix="%")
        st.plotly_chart(plotly_layout(fig_cia, 420), use_container_width=True)
    with col_table:
        st.dataframe(
            _brl_df(df_cia),
            use_container_width=True, hide_index=True, height=420,
            column_config={
                "GMV":           st.column_config.TextColumn("GMV"),
                "Gross Revenue": st.column_config.TextColumn("Gross Revenue"),
                "Take Rate (%)": st.column_config.NumberColumn("Take Rate (%)", format="%.2f%%"),
            },
        )

    # ── Cias com Take Rate negativo ──
    if not df_negativos.empty:
        st.markdown('<div class="sec-header-wrap"><p class="sec-header">⚠️ Cias com Take Rate Negativo</p></div>', unsafe_allow_html=True)
        st.caption("Exibe cias com take rate negativo **no acumulado do período selecionado**. Em períodos mais longos (ex: ano inteiro), uma cia que foi negativa em um mês isolado pode não aparecer aqui caso tenha recuperado o resultado nos demais meses.")
        st.dataframe(
            _brl_df(df_negativos),
            use_container_width=True, hide_index=True,
            column_config={
                "GMV":           st.column_config.TextColumn("GMV"),
                "Gross Revenue": st.column_config.TextColumn("Gross Revenue"),
                "Take Rate (%)": st.column_config.NumberColumn("Take Rate (%)", format="%.2f%%"),
            },
        )

    # ── Drill-down por Dia ──
    if dia_filtro:
        st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
        st.markdown(f'<div class="sec-header-wrap"><p class="sec-header">🔎 Drill-down — {dia_filtro.strftime("%d/%m/%Y")}</p></div>', unsafe_allow_html=True)

        with st.spinner("Carregando transações do dia..."):
            try:
                df_trans   = q_tr_drill_transacoes(i_str, f_str)
                df_cli     = q_tr_drill_clientes(i_str, f_str)
            except Exception as e:
                st.error(f"Erro: {e}")
                df_trans = df_cli = pd.DataFrame()

        tab_cli, tab_trans = st.tabs(["👥  Por Cliente", "📄  Transações"])

        with tab_cli:
            if not df_cli.empty:
                st.dataframe(
                    _brl_df(df_cli),
                    use_container_width=True, hide_index=True,
                    height=min(50 + len(df_cli) * 35, 600),
                    column_config={
                        "Reservas":      st.column_config.NumberColumn("Reservas",      format="%d"),
                        "GMV":           st.column_config.TextColumn("GMV"),
                        "Gross Revenue": st.column_config.TextColumn("Gross Revenue"),
                        "Take Rate (%)": st.column_config.NumberColumn("Take Rate (%)", format="%.2f%%"),
                    },
                )

        with tab_trans:
            if not df_trans.empty:
                # Filtro rápido por empresa
                empresas = ["Todas"] + sorted(df_trans["Empresa"].unique().tolist())
                col_f, _ = st.columns([1, 2])
                with col_f:
                    emp_sel = st.selectbox("Empresa", empresas, index=0, label_visibility="collapsed", key="drill_emp")
                df_t = df_trans if emp_sel == "Todas" else df_trans[df_trans["Empresa"] == emp_sel]
                st.dataframe(
                    _brl_df(df_t),
                    use_container_width=True, hide_index=True,
                    height=min(50 + len(df_t) * 35, 600),
                    column_config={
                        "Horário":       st.column_config.TextColumn("Horário",        width="small"),
                        "Reserva":       st.column_config.TextColumn("Reserva",        width="small"),
                        "Empresa":       st.column_config.TextColumn("Empresa"),
                        "Cia":           st.column_config.TextColumn("Cia",            width="small"),
                        "GMV":           st.column_config.TextColumn("GMV"),
                        "Gross Revenue": st.column_config.TextColumn("Gross Revenue"),
                        "Take Rate (%)": st.column_config.NumberColumn("Take Rate (%)", format="%.2f%%"),
                    },
                )

# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO: TAKE RATE C/ INCENTIVO
# ══════════════════════════════════════════════════════════════════════════════
elif secao == "💎  Take Rate c/ Incentivo":
    tab_analise, tab_cadastro = st.tabs(["📊  Análise", "📝  Cadastrar Incentivos"])

    # ── ABA: CADASTRO ──────────────────────────────────────────────────────────
    with tab_cadastro:
        st.markdown('<div class="sec-header-wrap"><p class="sec-header">Incentivos por Cia Aérea</p></div>', unsafe_allow_html=True)

        col_conf, _ = st.columns([1, 2])
        with col_conf:
            mes_ref = st.date_input(
                "Mês de referência",
                value=date.today().replace(day=1),
                format="MM/DD/YYYY",
            ).strftime("%Y-%m")

        dados = load_incentivos()
        mes_atual = dados.get(mes_ref, {})

        # Busca cias do período atual para sugerir as linhas
        with st.spinner("Carregando cias..."):
            try:
                df_cias_ref = q_tr_por_cia(
                    f"{mes_ref}-01",
                    (date.fromisoformat(f"{mes_ref}-01") + relativedelta(months=1) - timedelta(days=1)).strftime("%Y-%m-%d")
                )
                cias_sugeridas = df_cias_ref["Cia"].tolist()
            except Exception:
                cias_sugeridas = []

        # Monta dataframe para edição
        todas_cias = sorted({c for c in (set(cias_sugeridas) | set(mes_atual.keys())) if c is not None})
        df_editor = pd.DataFrame({
            "Cia Aérea":      todas_cias,
            "Incentivo (R$)": [float(mes_atual.get(c, 0.0)) for c in todas_cias],
        })

        st.markdown(f"**Mês: {mes_ref}** — edite os valores abaixo e clique em Salvar.")
        edited = st.data_editor(
            df_editor,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            column_config={
                "Cia Aérea":      st.column_config.TextColumn("Cia Aérea"),
                "Incentivo (R$)": st.column_config.NumberColumn("Incentivo (R$)", format="R$ %.2f", min_value=0.0),
            },
        )

        if st.button("💾  Salvar", type="primary"):
            dados[mes_ref] = {
                row["Cia Aérea"]: float(row["Incentivo (R$)"])
                for _, row in edited.iterrows()
                if row["Cia Aérea"] and float(row["Incentivo (R$)"]) != 0
            }
            save_incentivos(dados)
            st.success(f"Incentivos de {mes_ref} salvos com sucesso!")
            st.cache_data.clear()

        # Mostra resumo de todos os meses cadastrados
        if dados:
            st.markdown('<div class="sec-header-wrap"><p class="sec-header">Histórico Cadastrado</p></div>', unsafe_allow_html=True)
            rows_hist = []
            for mes, cias in sorted(dados.items()):
                for cia, val in cias.items():
                    rows_hist.append({"Mês": mes, "Cia Aérea": cia, "Incentivo (R$)": val})
            st.dataframe(
                pd.DataFrame(rows_hist),
                use_container_width=True, hide_index=True,
                column_config={
                    "Incentivo (R$)": st.column_config.TextColumn("Incentivo (R$)"),
                },
            )

    # ── ABA: ANÁLISE ───────────────────────────────────────────────────────────
    with tab_analise:
        tr = None
        with st.spinner("Consultando BigQuery..."):
            try:
                tr      = q_tr_resumo(i_str, f_str)
                df_cia  = q_tr_por_cia(i_str, f_str)
                tr_comp = q_tr_resumo(ci_str, cf_str) if ci_str else None
            except Exception as e:
                st.error(f"Erro ao conectar ao BigQuery: {e}")

        if tr is not None:
            incentivos = get_incentivo_periodo(i_str, f_str)
            total_incentivo = sum(incentivos.values())

            gmv_total    = tr["gmv"]
            gr_sem       = tr["gross_revenue"]
            gr_com       = gr_sem + total_incentivo
            tr_sem_pct   = tr["take_rate"]
            tr_com_pct   = (gr_com / gmv_total * 100) if gmv_total else 0

            # Rebate: denominador = onfly_amount das cias com incentivo (exclui fee Onfly)
            cias_com_incentivo    = {cia for cia, val in incentivos.items() if val > 0}
            onfly_amt_c_incentivo = df_cia[df_cia["Cia"].isin(cias_com_incentivo)]["Onfly Amount"].sum()
            rebate_pct            = (total_incentivo / onfly_amt_c_incentivo * 100) if onfly_amt_c_incentivo else 0

            # KPI Cards
            c1, c2, c3, c4, c5 = st.columns(5)
            with c1:
                st.markdown(f"""
                    <div class="kpi-card" style="border-top-color:#64748B;">
                        <p class="kpi-label">Take Rate s/ Incentivo</p>
                        <p class="kpi-value">{tr_sem_pct:.2f}%</p>
                        {delta_html(tr_sem_pct, tr_comp["take_rate"] if tr_comp else None)}
                    </div>
                """, unsafe_allow_html=True)
            with c2:
                st.markdown(f"""
                    <div class="kpi-card" style="border-top-color:#8B5CF6;">
                        <p class="kpi-label">Take Rate c/ Incentivo</p>
                        <p class="kpi-value">{tr_com_pct:.2f}%</p>
                        <p style="font-size:0.75rem;color:#16A34A;font-weight:600;margin:2px 0 0 0;">
                            +{tr_com_pct - tr_sem_pct:.2f}pp de incentivo
                        </p>
                    </div>
                """, unsafe_allow_html=True)
            with c3:
                st.markdown(f"""
                    <div class="kpi-card orange">
                        <p class="kpi-label">Total de Incentivos</p>
                        <p class="kpi-value">{brl(total_incentivo)}</p>
                    </div>
                """, unsafe_allow_html=True)
            with c4:
                st.markdown(f"""
                    <div class="kpi-card green">
                        <p class="kpi-label">Gross Revenue c/ Incentivo</p>
                        <p class="kpi-value">{brl(gr_com)}</p>
                    </div>
                """, unsafe_allow_html=True)
            with c5:
                st.markdown(f"""
                    <div class="kpi-card" style="border-top-color:#F59E0B;">
                        <p class="kpi-label">Rebate Médio (% GMV)</p>
                        <p class="kpi-value">{rebate_pct:.2f}%</p>
                        <p style="font-size:0.72rem;color:#94A3B8;font-weight:500;margin:2px 0 0 0;">
                            incentivo ÷ GMV
                        </p>
                    </div>
                """, unsafe_allow_html=True)

            # Take Rate por cia com incentivo
            st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
            st.markdown('<div class="sec-header-wrap"><p class="sec-header">Take Rate por Cia Aérea c/ Incentivo</p></div>', unsafe_allow_html=True)
            df_inc = df_cia.copy()
            df_inc["Incentivo"] = df_inc["Cia"].map(lambda c: incentivos.get(c, 0.0))
            df_inc["GR c/ Incentivo"] = df_inc["Gross Revenue"] + df_inc["Incentivo"]
            df_inc["TR c/ Incentivo (%)"] = (df_inc["GR c/ Incentivo"] / df_inc["GMV"].replace(0, float("nan")) * 100).round(2)
            df_inc["Rebate (%)"] = (df_inc["Incentivo"] / df_inc["Onfly Amount"].replace(0, float("nan")) * 100).round(2)

            col_chart, col_table = st.columns([1, 1], gap="large")
            with col_chart:
                df_plot = df_inc.head(12).sort_values("TR c/ Incentivo (%)").copy()
                df_plot["Incremento (%)"] = (df_plot["TR c/ Incentivo (%)"] - df_plot["Take Rate (%)"]).clip(lower=0)
                fig = go.Figure()
                fig.add_trace(go.Bar(
                    x=df_plot["Take Rate (%)"], y=df_plot["Cia"],
                    orientation="h", name="s/ Incentivo",
                    marker_color=ONFLY_BLUE,
                    hovertemplate="<b>%{y}</b><br>TR s/ Incentivo: %{x:.2f}%<extra></extra>",
                ))
                fig.add_trace(go.Bar(
                    x=df_plot["Incremento (%)"], y=df_plot["Cia"],
                    orientation="h", name="Incentivo",
                    marker_color="#8B5CF6",
                    hovertemplate="<b>%{y}</b><br>Incremento: +%{x:.2f}pp<extra></extra>",
                    text=[f"{v:.2f}%" for v in df_plot["TR c/ Incentivo (%)"]],
                    textposition="outside",
                    textfont=dict(size=12, color="#0F172A", weight="bold"),
                ))
                fig.update_layout(
                    barmode="stack",
                    xaxis_ticksuffix="%",
                    xaxis=dict(range=[0, df_plot["TR c/ Incentivo (%)"].max() * 1.25]),
                )
                st.plotly_chart(plotly_layout(fig, 420), use_container_width=True)

            with col_table:
                st.dataframe(
                    _brl_df(df_inc[["Cia", "GMV", "Incentivo", "Rebate (%)", "Take Rate (%)", "TR c/ Incentivo (%)"]]),
                    use_container_width=True, hide_index=True, height=420,
                    column_config={
                        "GMV":                 st.column_config.TextColumn("GMV"),
                        "Incentivo":           st.column_config.TextColumn("Incentivo"),
                        "Rebate (%)":          st.column_config.NumberColumn("Rebate % (inc÷GMV)",  format="%.2f%%"),
                        "Take Rate (%)":       st.column_config.NumberColumn("TR s/ Incentivo (%)", format="%.2f%%"),
                        "TR c/ Incentivo (%)": st.column_config.NumberColumn("TR c/ Incentivo (%)", format="%.2f%%"),
                    },
                )

# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO: CIA AÉREA
# ══════════════════════════════════════════════════════════════════════════════
elif secao == "✈️  Cia Aérea":
    # ── Origem das emissões ──
    with st.spinner("Carregando origens de emissão..."):
        try:
            df_origem = q_origem_emissoes(i_str, f_str)
        except Exception as e:
            st.error(f"Erro ao conectar ao BigQuery: {e}")
            st.stop()

    # ── Carrega lista de cias disponíveis ──
    with st.spinner("Carregando companhias..."):
        try:
            lista_cias = q_lista_cias()
        except Exception as e:
            st.error(f"Erro ao conectar ao BigQuery: {e}")
            st.stop()

    if not lista_cias:
        st.info("Nenhuma companhia encontrada.")
        st.stop()

    opcoes = {cia: f"{cia} — {IATA_NOMES.get(cia, 'Não identificada')}" for cia in lista_cias}

    col_sel, _ = st.columns([2, 5])
    with col_sel:
        cia_sel = st.selectbox(
            "Companhia Aérea",
            options=[""] + list(opcoes.keys()),
            format_func=lambda x: "Selecione uma companhia..." if x == "" else opcoes[x],
        )

    if not cia_sel:
        st.info("Selecione uma companhia aérea para visualizar os dados.")
        st.stop()

    variantes = lista_cias[cia_sel]  # tuple com todos os valores raw da cia

    st.markdown(f"""
        <h3 style="margin:16px 0 4px 0; color:#0F172A; font-size:1.4rem; font-weight:800;">
            {opcoes[cia_sel]}
        </h3>
    """, unsafe_allow_html=True)

    # ── Carrega dados da cia selecionada ──
    with st.spinner(f"Carregando dados de {cia_sel}..."):
        try:
            kpis      = q_tr_cia(i_str, f_str, variantes)
            onfly_liq = q_onfly_liq_cia(i_str, f_str, variantes)
            df_day    = q_diario_cia(i_str, f_str, variantes)
            df_cab    = q_cabine_cia(i_str, f_str, variantes)
            df_rot    = q_rotas_cia(i_str, f_str, variantes)
            df_ano    = q_gmv_anual_cia(variantes)
            kpis_c    = q_tr_cia(ci_str, cf_str, variantes) if ci_str else None
            liq_c     = q_onfly_liq_cia(ci_str, cf_str, variantes) if ci_str else None
            df_comp   = q_diario_cia(ci_str, cf_str, variantes) if ci_str else None
        except Exception as e:
            st.error(f"Erro ao consultar BigQuery: {e}")
            st.stop()

    # ── KPI Cards ──
    c1, c2, c3, c4 = st.columns(4)
    kpi_cards = [
        (c1, "kpi-card",        "GMV",          brl(kpis["gmv"]),            kpis["gmv"],          kpis_c["gmv"]          if kpis_c else None),
        (c2, "kpi-card orange", "Trechos",       f'{kpis["trechos"]:,}',      kpis["trechos"],      kpis_c["trechos"]      if kpis_c else None),
        (c3, "kpi-card green",  "Take Rate",     f'{kpis["take_rate"]:.2f}%', kpis["take_rate"],    kpis_c["take_rate"]    if kpis_c else None),
        (c4, "kpi-card",        "Ticket Médio",  brl(kpis["ticket_medio"]),   kpis["ticket_medio"], kpis_c["ticket_medio"] if kpis_c else None),
    ]
    for col, css, label, valor, atual, anterior in kpi_cards:
        with col:
            st.markdown(f"""
                <div class="{css}">
                    <p class="kpi-label">{label}</p>
                    <p class="kpi-value">{valor}</p>
                    {delta_html(atual, anterior) if anterior is not None else ""}
                </div>
            """, unsafe_allow_html=True)

    # ── Card: Passagem sem markup e sem taxas ──
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">Passagem Líquida de Taxas</p></div>', unsafe_allow_html=True)
    _lc1, _lc2, _lc3, _ = st.columns(4)
    _liq_cards = [
        (_lc1, "kpi-card",        "🎫 Passagem (onfly − taxas)", brl(onfly_liq["onfly_liq"]),   onfly_liq["onfly_liq"],   liq_c["onfly_liq"]   if liq_c else None),
        (_lc2, "kpi-card orange", "onfly_amount (bruto)",        brl(onfly_liq["onfly_bruto"]), onfly_liq["onfly_bruto"], liq_c["onfly_bruto"] if liq_c else None),
        (_lc3, "kpi-card",        "Taxas de Embarque",           brl(onfly_liq["taxas"]),       onfly_liq["taxas"],       liq_c["taxas"]       if liq_c else None),
    ]
    for col, css, label, valor, atual, anterior in _liq_cards:
        with col:
            st.markdown(f"""
                <div class="{css}">
                    <p class="kpi-label">{label}</p>
                    <p class="kpi-value">{valor}</p>
                    {delta_html(atual, anterior) if anterior is not None else ""}
                </div>
            """, unsafe_allow_html=True)
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    # ── Evolução Diária ──
    render_diario(df_day, ONFLY_BLUE, df_comp, comp_label)

    # ── Cabine ──
    render_cabine(df_cab, ONFLY_BLUE)

    # ── Rotas ──
    render_rotas(df_rot)

    # ── GMV Histórico por Ano ──
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">GMV Histórico por Ano</p></div>', unsafe_allow_html=True)
    if not df_ano.empty:
        col_chart, col_table = st.columns([3, 2], gap="large")
        with col_chart:
            fig_ano = go.Figure()
            fig_ano.add_trace(go.Bar(
                x=df_ano["Ano"].astype(str), y=df_ano["GMV"],
                marker_color=ONFLY_BLUE,
                text=[brl(v) for v in df_ano["GMV"]],
                textposition="outside",
                textfont=dict(size=11, color="#334155"),
                hovertemplate="<b>%{x}</b><br>GMV: R$ %{y:,.2f}<extra></extra>",
            ))
            fig_ano.update_yaxes(tickprefix="R$ ")
            st.plotly_chart(plotly_layout(fig_ano, 280), use_container_width=True)
        with col_table:
            gmv_total_hist = df_ano["GMV"].sum()
            df_ano_show = df_ano.copy()
            df_ano_show["% Total"] = (df_ano_show["GMV"] / gmv_total_hist * 100).round(1)
            st.dataframe(
                _brl_df(df_ano_show[["Ano", "Trechos", "GMV", "% Total"]]),
                use_container_width=True, hide_index=True, height=280,
                column_config={
                    "Ano":      st.column_config.NumberColumn("Ano",     format="%d"),
                    "GMV":      st.column_config.TextColumn("GMV"),
                    "Trechos":  st.column_config.NumberColumn("Trechos", format="%d"),
                    "% Total":  st.column_config.NumberColumn("% Total", format="%.1f%%"),
                },
            )

    # ── Origem das Emissões (filtrado pela cia selecionada) ──
    with st.spinner("Carregando origens de emissão..."):
        try:
            df_origem = q_origem_emissoes(i_str, f_str, variantes)
        except Exception as e:
            st.error(f"Erro: {e}")
            df_origem = pd.DataFrame()

    if not df_origem.empty:
        st.markdown('<div class="sec-header-wrap"><p class="sec-header">Origem das Emissões</p></div>', unsafe_allow_html=True)
        col_pie, col_tbl = st.columns([1, 1], gap="large")
        gmv_total_orig = df_origem["GMV"].sum()

        with col_pie:
            fig_pie = px.pie(
                df_origem,
                names="Origem", values="GMV",
                hole=0.45,
                color_discrete_sequence=px.colors.qualitative.Set2,
            )
            fig_pie.update_traces(
                textposition="inside",
                textinfo="percent+label",
                hovertemplate="<b>%{label}</b><br>GMV: R$ %{value:,.2f}<br>%{percent}<extra></extra>",
            )
            fig_pie.update_layout(showlegend=False)
            st.plotly_chart(plotly_layout(fig_pie, 340), use_container_width=True)

        with col_tbl:
            df_orig_show = df_origem.copy()
            df_orig_show["% GMV"] = (df_orig_show["GMV"] / gmv_total_orig * 100).round(1)
            st.dataframe(
                _brl_df(df_orig_show[["Origem", "Reservas", "GMV", "Gross Revenue", "Take Rate (%)", "% GMV"]]),
                use_container_width=True, hide_index=True, height=340,
                column_config={
                    "Reservas":      st.column_config.NumberColumn("Reservas",      format="%d"),
                    "GMV":           st.column_config.TextColumn("GMV"),
                    "Gross Revenue": st.column_config.TextColumn("Gross Revenue"),
                    "Take Rate (%)": st.column_config.NumberColumn("Take Rate (%)", format="%.2f%%"),
                    "% GMV":         st.column_config.NumberColumn("% GMV",         format="%.1f%%"),
                },
            )

    # ── Clientes que emitiram nesta cia ──────────────────────────────────────
    st.markdown("<div style='height:48px'></div>", unsafe_allow_html=True)
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">Clientes que Emitiram</p></div>', unsafe_allow_html=True)

    with st.spinner("Carregando clientes..."):
        df_cli_cia = q_clientes_por_cia(variantes, i_str, f_str)

    if df_cli_cia.empty:
        st.info("Nenhum cliente encontrado para esta cia no período.")
    else:
        def _fmt_cnpj(c):
            c = str(c).strip()
            return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}" if len(c) == 14 else c

        df_cli_cia_show = df_cli_cia.copy()
        df_cli_cia_show["CNPJ"] = df_cli_cia_show["CNPJ"].apply(_fmt_cnpj)
        _gmv_total_cia = df_cli_cia_show["GMV"].sum()
        df_cli_cia_show["% do Total"] = (df_cli_cia_show["GMV"] / _gmv_total_cia * 100).round(1)
        df_cli_cia_show["GMV"]          = df_cli_cia_show["GMV"].apply(brl)
        df_cli_cia_show["Ticket Médio"] = df_cli_cia_show["Ticket Médio"].apply(brl)

        st.caption(f"{len(df_cli_cia_show)} clientes · GMV total {brl(_gmv_total_cia)}")
        st.dataframe(
            df_cli_cia_show[["Cliente", "CNPJ", "Razão Social", "Reservas", "GMV", "Ticket Médio", "% do Total"]],
            use_container_width=True,
            hide_index=True,
            height=min(55 + len(df_cli_cia_show) * 35, 600),
            column_config={
                "Cliente":      st.column_config.TextColumn("Cliente"),
                "CNPJ":         st.column_config.TextColumn("CNPJ"),
                "Razão Social": st.column_config.TextColumn("Razão Social"),
                "Reservas":     st.column_config.NumberColumn("Reservas", format="%d"),
                "GMV":          st.column_config.TextColumn("GMV"),
                "Ticket Médio": st.column_config.TextColumn("Ticket Médio"),
                "% do Total":   st.column_config.NumberColumn("% do Total", format="%.1f%%"),
            },
        )

# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO: DISTRIBUIÇÃO
# ══════════════════════════════════════════════════════════════════════════════
elif secao == "🗺️  Distribuição":

    GDS_LISTA = ["Amadeus", "Sabre", "Travelport", "Galileo", "Worldspan", "Apollo"]
    TIPOS_CONEXAO = ["Direta", "Direta NDC", "NDC offline", "Via Consolidador", "Via Wooba", "Via Wooba NDC", "Via GDS", "Via GDS NDC offline"]
    TIPOS_INCENTIVO_DIST = ["Upfront", "Backend", "Comissão via BSP", "Desconto na Tarifa", "OBT"]

    # ── Carrega lista de cias ──
    with st.spinner("Carregando companhias..."):
        try:
            lista_cias_dist = q_lista_cias()
        except Exception as e:
            st.error(f"Erro ao conectar ao BigQuery: {e}")
            st.stop()

    opcoes_cias = {cia: f"{cia} — {IATA_NOMES.get(cia, 'Não identificada')}" for cia in lista_cias_dist}
    dist_dados  = load_distribuicao()

    # ── Session state ─────────────────────────────────────────────────────────
    if "dist_editing" not in st.session_state:
        st.session_state["dist_editing"] = False
    if "dist_cia_editar_pre" in st.session_state:
        st.session_state["dist_cia_form"] = st.session_state.pop("dist_cia_editar_pre")
        st.session_state["dist_editing"]  = True

    import base64 as _b64

    # ── Header + botão Nova Cia ────────────────────────────────────────────────
    _hcol, _ncol = st.columns([5, 1])
    with _hcol:
        st.markdown('<div class="sec-header-wrap"><p class="sec-header">🗺️ Distribuição</p></div>',
                    unsafe_allow_html=True)
    with _ncol:
        st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)
        if st.button("➕  Nova Cia", use_container_width=True, key="dist_nova_cia"):
            st.session_state["dist_editing"] = True
            st.session_state["dist_cia_form"] = ""
            st.rerun()

    # ── Tabela ────────────────────────────────────────────────────────────────
    if not dist_dados:
        st.info("Nenhuma cia cadastrada ainda. Clique em **➕ Nova Cia** para adicionar.")
    else:
        rows_vis = []
        for _cia, _cfg in sorted(dist_dados.items()):
            _anx_list = _get_anexos(_cfg)
            _qtd_anx  = len(_anx_list)
            _img_url  = None
            for _anx in _anx_list:
                if _anx.get("mime", "").startswith("image/"):
                    _img_url = f"data:{_anx['mime']};base64,{_anx['dados']}"
                    break
            if "dist_ver_anx" not in st.session_state:
                st.session_state["dist_ver_anx"] = None

            rows_vis.append({
                "🗑️":                 False,
                "Sigla":              _cia,
                "Companhia":          IATA_NOMES.get(_cia) or IATA_NOMES.get(IATA_NORMALIZACAO.get(_cia, _cia), "—"),
                "Tipo de Conexão":    ", ".join(_cfg.get("conexao_tipo") or []) or "—",
                "Consolidador":       ", ".join(_cfg.get("consolidador") or []) or "—",
                "GDS":                _cfg.get("gds") or "—",
                "Tipos de Incentivo": ", ".join(_cfg.get("tipos_incentivo", [])) or "—",
                "Descrição":          _cfg.get("descricao") or "—",
                "📎":                 _img_url,
                "Nº":                 _qtd_anx if _qtd_anx else None,
            })

        edited_vis = st.data_editor(
            pd.DataFrame(rows_vis),
            use_container_width=True,
            hide_index=True,
            column_config={
                "🗑️":                 st.column_config.CheckboxColumn("🗑️", help="Excluir", width="small"),
                "Sigla":              st.column_config.TextColumn("Sigla", width="small"),
                "Companhia":          st.column_config.TextColumn("Companhia"),
                "Tipo de Conexão":    st.column_config.TextColumn("Tipo de Conexão"),
                "Consolidador":       st.column_config.TextColumn("Consolidador"),
                "GDS":                st.column_config.TextColumn("GDS", width="small"),
                "Tipos de Incentivo": st.column_config.TextColumn("Tipos de Incentivo"),
                "Descrição":          st.column_config.TextColumn("Descrição"),
                "📎":                 st.column_config.ImageColumn("📎", help="Passe o mouse para ampliar", width="small"),
                "Nº":                 st.column_config.NumberColumn("Nº", help="Total de anexos", width="small", format="%d"),
            },
            disabled=["Sigla", "Companhia", "Tipo de Conexão", "Consolidador",
                      "GDS", "Tipos de Incentivo", "Descrição", "📎", "Nº"],
        )

        # ── Botões de edição — um por cia ─────────────────────────────────────
        st.markdown(
            "<p style='font-size:0.72rem;font-weight:700;color:#8C9BAB;"
            "text-transform:uppercase;letter-spacing:0.07em;margin:10px 0 6px 0;'>"
            "✏️ Editar</p>",
            unsafe_allow_html=True,
        )
        _cias_ord = sorted(dist_dados.keys())
        _btn_cols = st.columns(min(len(_cias_ord), 8))
        for _bi, _bc in enumerate(_cias_ord):
            with _btn_cols[_bi % 8]:
                _ativo = st.session_state.get("dist_cia_form") == _bc and st.session_state.get("dist_editing")
                if st.button(
                    _bc,
                    key=f"edit_btn_{_bc}",
                    type="primary" if _ativo else "secondary",
                    use_container_width=True,
                    help=IATA_NOMES.get(_bc) or IATA_NOMES.get(IATA_NORMALIZACAO.get(_bc, _bc), _bc),
                ):
                    if _ativo:
                        st.session_state["dist_editing"]  = False
                        st.session_state["dist_cia_form"] = ""
                    else:
                        st.session_state["dist_cia_form"] = _bc
                        st.session_state["dist_editing"]  = True
                    st.rerun()

        para_excluir = edited_vis[edited_vis["🗑️"] == True]["Sigla"].tolist()
        if para_excluir:
            _ce1, _ce2 = st.columns([3, 1])
            with _ce1:
                st.warning(f"**{len(para_excluir)} registro(s) marcado(s):** {', '.join(para_excluir)}")
            with _ce2:
                if st.button("🗑️  Confirmar Exclusão", type="primary", key="dist_del_confirm"):
                    for _cia in para_excluir:
                        dist_dados.pop(_cia, None)
                    save_distribuicao(dist_dados)
                    st.rerun()

    # ── Galeria de anexos ─────────────────────────────────────────────────────
    _cias_com_anx = [(c, _get_anexos(cfg)) for c, cfg in sorted(dist_dados.items()) if _get_anexos(cfg)]
    if _cias_com_anx:
        st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)
        st.markdown('<div class="sec-header-wrap"><p class="sec-header">📎 Galeria de Anexos</p></div>',
                    unsafe_allow_html=True)
        _ops_galeria = {c: f"{c} — {IATA_NOMES.get(c) or IATA_NOMES.get(IATA_NORMALIZACAO.get(c,c), c)}  ({len(a)} arquivo{'s' if len(a)>1 else ''})"
                        for c, a in _cias_com_anx}
        _col_gsel, _ = st.columns([2, 4])
        with _col_gsel:
            _sel_gal = st.selectbox(
                "Ver anexos de:",
                options=[""] + list(_ops_galeria.keys()),
                format_func=lambda x: "Selecione uma companhia..." if x == "" else _ops_galeria[x],
                key="dist_galeria_sel",
            )
        if _sel_gal:
            _anx_ver  = _get_anexos(dist_dados[_sel_gal])
            _ncols_g  = min(len(_anx_ver), 4)
            _gcols_v  = st.columns(_ncols_g)
            for _gi_v, _anx_v in enumerate(_anx_ver):
                with _gcols_v[_gi_v % _ncols_g]:
                    st.markdown(
                        f"<p style='font-size:0.72rem;font-weight:600;color:#475569;"
                        f"margin-bottom:4px;text-align:center;'>{_anx_v['nome']}</p>",
                        unsafe_allow_html=True,
                    )
                    try:
                        _bytes_gv = _b64.b64decode(_anx_v["dados"])
                        if _anx_v["mime"] == "application/pdf":
                            st.markdown(
                                "<div style='background:#F1F5F9;border-radius:8px;padding:24px;"
                                "text-align:center;font-size:2rem;'>📄</div>",
                                unsafe_allow_html=True,
                            )
                            st.download_button("⬇️ Baixar PDF", _bytes_gv, file_name=_anx_v["nome"],
                                               mime="application/pdf",
                                               key=f"gal_dl_{_sel_gal}_{_gi_v}",
                                               use_container_width=True)
                        else:
                            st.image(_bytes_gv, use_container_width=True)
                    except Exception as _eg:
                        st.warning(f"Erro: {_eg}")

    # ── Formulário inline ─────────────────────────────────────────────────────
    if st.session_state["dist_editing"]:
        st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

        with st.container(border=True):
            # Seletor de cia no topo do form
            _pre_form = st.session_state.get("dist_cia_form", "")
            _opts_form = [""] + list(opcoes_cias.keys())
            _idx_form  = _opts_form.index(_pre_form) if _pre_form in _opts_form else 0
            _fh, _fc = st.columns([5, 1])
            with _fh:
                cia_dist = st.selectbox(
                    "Companhia Aérea",
                    options=_opts_form,
                    index=_idx_form,
                    format_func=lambda x: "Selecione uma companhia..." if x == "" else opcoes_cias[x],
                    key="dist_cia_sel",
                )
            with _fc:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                if st.button("✖️  Fechar", key="dist_fechar", use_container_width=True):
                    st.session_state["dist_editing"]  = False
                    st.session_state["dist_cia_form"] = ""
                    st.rerun()

            if not cia_dist:
                st.info("Selecione uma companhia aérea acima.")
            else:
                existente = dist_dados.get(cia_dist, {})
                st.markdown(
                    f"<h3 style='margin:12px 0;color:#0F172A;font-size:1.1rem;font-weight:800;'>"
                    f"{opcoes_cias[cia_dist]}</h3>",
                    unsafe_allow_html=True,
                )

                col1, col2 = st.columns(2)
                with col1:
                    default_conexao = [t for t in existente.get("conexao_tipo", []) if t in TIPOS_CONEXAO] if isinstance(existente.get("conexao_tipo"), list) else []
                    conexao_tipo = st.multiselect("Tipo de Conexão", options=TIPOS_CONEXAO,
                                                  default=default_conexao, key="dist_conexao_tipo")
                with col2:
                    default_cons = [c for c in existente.get("consolidador", []) if c in CONSOLIDADORAS] if isinstance(existente.get("consolidador"), list) else []
                    consolidador_sel = st.multiselect("Consolidador", options=sorted(CONSOLIDADORAS),
                                                      default=default_cons, key="dist_consolidador")

                col3, col4 = st.columns(2)
                with col3:
                    gds_sel = st.selectbox(
                        "GDS", options=[""] + GDS_LISTA,
                        index=(GDS_LISTA.index(existente["gds"]) + 1) if existente.get("gds") in GDS_LISTA else 0,
                        format_func=lambda x: "Selecione..." if x == "" else x,
                        key="dist_gds",
                    )
                with col4:
                    _compat_tipos = {"Comissão": "Comissão via BSP"}
                    default_tipos = [_compat_tipos.get(t, t) for t in existente.get("tipos_incentivo", [])
                                     if _compat_tipos.get(t, t) in TIPOS_INCENTIVO_DIST]
                    tipos_incentivo_sel = st.multiselect("Tipos de Incentivo", options=TIPOS_INCENTIVO_DIST,
                                                         default=default_tipos, key="dist_tipos_incentivo")

                descricao = st.text_area(
                    "Descrição", value=existente.get("descricao", ""), height=100,
                    placeholder="Condições, observações ou detalhes...",
                    key="dist_descricao",
                )

                # ── Anexos ────────────────────────────────────────────────────
                st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
                st.markdown("**📎 Anexos**")
                _rm_key         = f"dist_rm_{cia_dist}"
                if _rm_key not in st.session_state:
                    st.session_state[_rm_key] = set()
                _anx_existentes = _get_anexos(existente)
                _anx_visiveis   = [(_i, _a) for _i, _a in enumerate(_anx_existentes)
                                   if _i not in st.session_state[_rm_key]]
                if _anx_visiveis:
                    _ncols = min(len(_anx_visiveis), 4)
                    _gcols = st.columns(_ncols)
                    for _gi, (_ri, _anx) in enumerate(_anx_visiveis):
                        with _gcols[_gi % _ncols]:
                            try:
                                _bytes_a = _b64.b64decode(_anx["dados"])
                                if _anx["mime"] == "application/pdf":
                                    st.markdown(
                                        f"<div style='background:#F1F5F9;border-radius:6px;padding:10px;"
                                        f"text-align:center;font-size:0.75rem;color:#475569;'>"
                                        f"📄<br><b>{_anx['nome']}</b></div>",
                                        unsafe_allow_html=True,
                                    )
                                    st.download_button("⬇️ PDF", _bytes_a, file_name=_anx["nome"],
                                                       mime="application/pdf",
                                                       key=f"dl_anx_{cia_dist}_{_ri}",
                                                       use_container_width=True)
                                else:
                                    st.image(_bytes_a, caption=_anx["nome"], use_container_width=True)
                            except Exception:
                                st.markdown(f"⚠️ {_anx['nome']}")
                            if st.button("🗑️ Remover", key=f"rm_anx_{cia_dist}_{_ri}",
                                         use_container_width=True):
                                st.session_state[_rm_key].add(_ri)
                                st.rerun()

                _novos_uploads = st.file_uploader(
                    "Adicionar imagens ou PDFs",
                    type=["png", "jpg", "jpeg", "pdf"],
                    accept_multiple_files=True,
                    key=f"dist_upload_{cia_dist}",
                )

                st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
                _sb, _db, _ = st.columns([1, 1, 4])
                with _sb:
                    if st.button("💾  Salvar", type="primary", key="dist_save"):
                        _rm_set   = st.session_state.get(_rm_key, set())
                        _mantidos = [a for i, a in enumerate(_anx_existentes) if i not in _rm_set]
                        _adicionados = []
                        for _nf in (_novos_uploads or []):
                            _ext_nf  = _nf.name.rsplit(".", 1)[-1].lower() if "." in _nf.name else "png"
                            _mime_nf = "application/pdf" if _ext_nf == "pdf" else ("image/jpeg" if _ext_nf in ("jpg","jpeg") else "image/png")
                            _adicionados.append({"nome": _nf.name,
                                                 "dados": _b64.b64encode(_nf.read()).decode(),
                                                 "mime":  _mime_nf})
                        dist_dados[cia_dist] = {
                            "conexao_tipo":            conexao_tipo,
                            "consolidador":            consolidador_sel,
                            "gds":                     gds_sel or "",
                            "tipos_incentivo":         tipos_incentivo_sel,
                            "descricao":               descricao,
                            "anexos":                  _mantidos + _adicionados,
                            "comissao_bsp_imagem":     None,
                            "comissao_bsp_imagem_nome": None,
                        }
                        save_distribuicao(dist_dados)
                        st.session_state.pop(_rm_key, None)
                        st.session_state["dist_editing"]  = False
                        st.session_state["dist_cia_form"] = ""
                        st.rerun()
                with _db:
                    if cia_dist in dist_dados:
                        if st.button("🗑️  Excluir", key="dist_del"):
                            dist_dados.pop(cia_dist)
                            save_distribuicao(dist_dados)
                            st.session_state["dist_editing"]  = False
                            st.session_state["dist_cia_form"] = ""
                            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO: DESTINO
# ══════════════════════════════════════════════════════════════════════════════
elif secao == "🌍  Destino":
    with st.spinner("Carregando aeroportos..."):
        try:
            lista_aeroportos = q_lista_aeroportos()
        except Exception as e:
            st.error(f"Erro ao conectar ao BigQuery: {e}")
            st.stop()

    col_sel, _ = st.columns([2, 4])
    with col_sel:
        aeroporto = st.selectbox(
            "Aeroporto",
            options=[""] + lista_aeroportos,
            format_func=lambda x: "Selecione um aeroporto..." if x == "" else x,
            key="destino_aeroporto_sel",
        )

    if not aeroporto:
        st.info("Selecione um aeroporto para ver as rotas.")
    else:
        with st.spinner(f"Consultando rotas para {aeroporto}..."):
            try:
                df_dest = q_destinos_por_aeroporto(aeroporto, i_str, f_str)
            except Exception as e:
                st.error(f"Erro ao consultar BigQuery: {e}")
                st.stop()

        if df_dest.empty:
            st.warning(f"Nenhuma rota encontrada para **{aeroporto}**.")
        else:
            # KPIs resumo
            total_reservas = df_dest["Reservas"].sum()
            total_gmv      = df_dest["GMV"].sum()
            total_rotas    = len(df_dest)

            st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
            st.markdown('<div class="sec-header-wrap"><p class="sec-header">✈️ TOTAL</p></div>', unsafe_allow_html=True)
            k1, k2, k3 = st.columns(3)
            for col, label, valor in [
                (k1, "Rotas",    f"{total_rotas:,}"),
                (k2, "Reservas", f"{total_reservas:,}"),
                (k3, "GMV",      brl(total_gmv)),
            ]:
                with col:
                    st.markdown(f"""
                        <div class="kpi-card">
                            <p class="kpi-label">{label}</p>
                            <p class="kpi-value">{valor}</p>
                        </div>
                    """, unsafe_allow_html=True)

            st.markdown("<div style='height:32px'></div>", unsafe_allow_html=True)
            def _tabela_rotas(df_bloco: pd.DataFrame, gmv_ref: float):
                df_s = df_bloco.copy()
                df_s["% GMV"] = (df_s["GMV"] / gmv_ref * 100).round(1)
                st.dataframe(
                    _brl_df(df_s[["Origem", "Destino", "Reservas", "GMV", "% GMV", "Cias"]]),
                    use_container_width=True,
                    hide_index=True,
                    height=min(50 + len(df_s) * 35, 600),
                    column_config={
                        "GMV":      st.column_config.TextColumn("GMV"),
                        "Reservas": st.column_config.NumberColumn("Reservas", format="%d"),
                        "% GMV":    st.column_config.NumberColumn("% GMV",    format="%.1f%%"),
                    },
                )

            # ── Saindo deste aeroporto ────────────────────────────────────────
            df_saindo = df_dest[df_dest["Origem"].str.upper().str.strip() == aeroporto].copy()
            gmv_saindo = df_saindo["GMV"].sum()
            res_saindo = df_saindo["Reservas"].sum()
            st.markdown(f'<div class="sec-header-wrap"><p class="sec-header">✈️ Saindo de {aeroporto}</p></div>', unsafe_allow_html=True)
            c1, c2, c3 = st.columns(3)
            with c1:
                st.markdown(f"""<div class="kpi-card"><p class="kpi-label">Rotas</p>
                    <p class="kpi-value">{len(df_saindo):,}</p></div>""", unsafe_allow_html=True)
            with c2:
                st.markdown(f"""<div class="kpi-card"><p class="kpi-label">Reservas</p>
                    <p class="kpi-value">{res_saindo:,}</p></div>""", unsafe_allow_html=True)
            with c3:
                st.markdown(f"""<div class="kpi-card"><p class="kpi-label">GMV</p>
                    <p class="kpi-value">{brl(gmv_saindo)}</p></div>""", unsafe_allow_html=True)
            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
            if df_saindo.empty:
                st.info("Nenhuma rota saindo deste aeroporto no período.")
            else:
                _tabela_rotas(df_saindo, total_gmv)

            st.markdown("<div style='height:32px'></div>", unsafe_allow_html=True)

            # ── Chegando neste aeroporto ──────────────────────────────────────
            df_chegando = df_dest[df_dest["Destino"].str.upper().str.strip() == aeroporto].copy()
            gmv_chegando = df_chegando["GMV"].sum()
            res_chegando = df_chegando["Reservas"].sum()
            st.markdown(f'<div class="sec-header-wrap"><p class="sec-header">🛬 Chegando em {aeroporto}</p></div>', unsafe_allow_html=True)
            c4, c5, c6 = st.columns(3)
            with c4:
                st.markdown(f"""<div class="kpi-card"><p class="kpi-label">Rotas</p>
                    <p class="kpi-value">{len(df_chegando):,}</p></div>""", unsafe_allow_html=True)
            with c5:
                st.markdown(f"""<div class="kpi-card"><p class="kpi-label">Reservas</p>
                    <p class="kpi-value">{res_chegando:,}</p></div>""", unsafe_allow_html=True)
            with c6:
                st.markdown(f"""<div class="kpi-card"><p class="kpi-label">GMV</p>
                    <p class="kpi-value">{brl(gmv_chegando)}</p></div>""", unsafe_allow_html=True)
            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
            if df_chegando.empty:
                st.info("Nenhuma rota chegando neste aeroporto no período.")
            else:
                _tabela_rotas(df_chegando, total_gmv)

# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO: POTENCIAL DE VOO
# ══════════════════════════════════════════════════════════════════════════════
elif secao == "🌐  Potencial de Voo":
    continentes_ordenados = sorted(CONTINENTE_PAISES.keys())

    col_cont, col_pais, _ = st.columns([2, 2, 3])
    with col_cont:
        continente_sel = st.selectbox(
            "Continente",
            options=[""] + continentes_ordenados,
            format_func=lambda x: "Selecione um continente..." if x == "" else x,
            key="potencial_cont_sel",
        )

    paises_sel = []
    if continente_sel:
        paises_continente = CONTINENTE_PAISES.get(continente_sel, [])
        with col_pais:
            paises_sel = st.multiselect(
                "País",
                options=paises_continente,
                placeholder="Selecione um ou mais países...",
                key="potencial_pais_sel",
            )

    if not continente_sel:
        st.info("Selecione um continente para começar.")
    elif not paises_sel:
        st.markdown(f"<p style='color:#6B7280;font-size:0.9rem;margin-top:8px;'>📍 {len(CONTINENTE_PAISES.get(continente_sel,[]))} países disponíveis em {continente_sel.split(' ',1)[-1]}: {', '.join(CONTINENTE_PAISES.get(continente_sel,[]))}</p>", unsafe_allow_html=True)
        st.info("Selecione um ou mais países para ver os voos emitidos.")
    else:
        _label_paises = ", ".join(paises_sel) if len(paises_sel) <= 3 else f"{len(paises_sel)} países selecionados"
        st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
        st.markdown(f'<div class="sec-header-wrap"><p class="sec-header">✈️ Chegando em {_label_paises} de qualquer destino</p></div>',
                    unsafe_allow_html=True)

        with st.spinner(f"Consultando voos para {_label_paises}..."):
            try:
                df_pot = q_potencial_paises(tuple(paises_sel), i_str, f_str)
            except Exception as e:
                st.error(f"Erro ao consultar BigQuery: {e}")
                st.stop()

        if df_pot.empty:
            st.warning(f"Nenhuma emissão encontrada para **{_label_paises}** no período.")
        else:
            total_reservas = df_pot["Reservas"].sum()
            total_gmv      = df_pot["GMV"].sum()
            total_rotas    = df_pot[["Origem", "Destino"]].drop_duplicates().shape[0]
            cias_unicas    = df_pot["Cia"].nunique()

            k1, k2, k3, k4 = st.columns(4)
            for col, label, valor in [
                (k1, "Rotas",       f"{total_rotas:,}"),
                (k2, "Reservas",    f"{total_reservas:,}"),
                (k3, "GMV",         brl(total_gmv)),
                (k4, "Cias Aéreas", f"{cias_unicas:,}"),
            ]:
                with col:
                    st.markdown(f"""
                        <div class="kpi-card">
                            <p class="kpi-label">{label}</p>
                            <p class="kpi-value">{valor}</p>
                        </div>
                    """, unsafe_allow_html=True)

            st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

            # ── Por aeroporto de destino ───────────────────────────────────
            st.markdown('<div class="sec-header-wrap"><p class="sec-header">Por Aeroporto de Destino</p></div>', unsafe_allow_html=True)
            df_por_dest = (
                df_pot.groupby("Destino", as_index=False)
                .agg(Reservas=("Reservas", "sum"), GMV=("GMV", "sum"))
                .sort_values("GMV", ascending=False)
            )
            df_por_dest["% GMV"] = (df_por_dest["GMV"] / total_gmv * 100).round(1)
            col_chart, col_tbl = st.columns([3, 2], gap="large")
            with col_chart:
                fig_dest = go.Figure(go.Bar(
                    x=df_por_dest["Destino"],
                    y=df_por_dest["GMV"],
                    marker_color=ONFLY_BLUE,
                    text=[brl(v) for v in df_por_dest["GMV"]],
                    textposition="outside",
                    textfont=dict(size=10, color="#334155"),
                    hovertemplate="<b>%{x}</b><br>GMV: R$ %{y:,.2f}<extra></extra>",
                ))
                fig_dest.update_yaxes(tickprefix="R$ ")
                st.plotly_chart(plotly_layout(fig_dest, 300), use_container_width=True)
            with col_tbl:
                st.dataframe(
                    _brl_df(df_por_dest[["Destino", "Reservas", "GMV", "% GMV"]]),
                    use_container_width=True, hide_index=True, height=300,
                    column_config={
                        "GMV":      st.column_config.TextColumn("GMV"),
                        "Reservas": st.column_config.NumberColumn("Reservas", format="%d"),
                        "% GMV":    st.column_config.NumberColumn("% GMV",    format="%.1f%%"),
                    },
                )

            # ── Por cia aérea ──────────────────────────────────────────────
            st.markdown('<div class="sec-header-wrap"><p class="sec-header">Por Cia Aérea</p></div>', unsafe_allow_html=True)
            df_por_cia = (
                df_pot.groupby("Cia", as_index=False)
                .agg(Reservas=("Reservas", "sum"), GMV=("GMV", "sum"))
                .sort_values("GMV", ascending=False)
            )
            df_por_cia["% GMV"] = (df_por_cia["GMV"] / total_gmv * 100).round(1)
            col_c2, col_t2 = st.columns([3, 2], gap="large")
            with col_c2:
                fig_cia = go.Figure(go.Bar(
                    x=df_por_cia["Cia"],
                    y=df_por_cia["GMV"],
                    marker_color=ONFLY_BLUE,
                    text=[brl(v) for v in df_por_cia["GMV"]],
                    textposition="outside",
                    textfont=dict(size=10, color="#334155"),
                    hovertemplate="<b>%{x}</b><br>GMV: R$ %{y:,.2f}<extra></extra>",
                ))
                fig_cia.update_yaxes(tickprefix="R$ ")
                st.plotly_chart(plotly_layout(fig_cia, 300), use_container_width=True)
            with col_t2:
                st.dataframe(
                    _brl_df(df_por_cia[["Cia", "Reservas", "GMV", "% GMV"]]),
                    use_container_width=True, hide_index=True, height=300,
                    column_config={
                        "GMV":      st.column_config.TextColumn("GMV"),
                        "Reservas": st.column_config.NumberColumn("Reservas", format="%d"),
                        "% GMV":    st.column_config.NumberColumn("% GMV",    format="%.1f%%"),
                    },
                )

            # ── Detalhe completo ───────────────────────────────────────────
            st.markdown('<div class="sec-header-wrap"><p class="sec-header">Detalhe por Rota e Cia Aérea</p></div>', unsafe_allow_html=True)
            df_show = df_pot.copy()
            df_show["% GMV"] = (df_show["GMV"] / total_gmv * 100).round(1)
            st.dataframe(
                _brl_df(df_show[["Origem", "Destino", "Cia", "Reservas", "GMV", "Ticket Médio", "% GMV"]]),
                use_container_width=True, hide_index=True,
                height=min(50 + len(df_show) * 35, 600),
                column_config={
                    "GMV":          st.column_config.TextColumn("GMV"),
                    "Ticket Médio": st.column_config.TextColumn("Ticket Médio"),
                    "Reservas":     st.column_config.NumberColumn("Reservas",     format="%d"),
                    "% GMV":        st.column_config.NumberColumn("% GMV",        format="%.1f%%"),
                },
            )

            # ══ SEÇÃO: SAINDO DO PAÍS ════════════════════════════════════════
            st.markdown("<hr style='margin:32px 0 24px 0; border:none; border-top:1px solid #E2E8F0;'>",
                        unsafe_allow_html=True)
            st.markdown(f'<div class="sec-header-wrap"><p class="sec-header">✈️ Saindo de {_label_paises} para qualquer destino</p></div>',
                        unsafe_allow_html=True)

            with st.spinner(f"Consultando voos saindo de {_label_paises}..."):
                try:
                    df_saindo = q_potencial_paises_saindo(tuple(paises_sel), i_str, f_str)
                except Exception as e:
                    st.error(f"Erro ao consultar BigQuery: {e}")
                    df_saindo = pd.DataFrame()

            if df_saindo.empty:
                st.info(f"Nenhuma emissão encontrada saindo de **{_label_paises}** no período.")
            else:
                tot_r_s = df_saindo["Reservas"].sum()
                tot_g_s = df_saindo["GMV"].sum()
                tot_rt_s = df_saindo[["Origem", "Destino"]].drop_duplicates().shape[0]
                cias_s   = df_saindo["Cia"].nunique()

                ks1, ks2, ks3, ks4 = st.columns(4)
                for col, label, valor in [
                    (ks1, "Rotas",       f"{tot_rt_s:,}"),
                    (ks2, "Reservas",    f"{tot_r_s:,}"),
                    (ks3, "GMV",         brl(tot_g_s)),
                    (ks4, "Cias Aéreas", f"{cias_s:,}"),
                ]:
                    with col:
                        st.markdown(f"""
                            <div class="kpi-card">
                                <p class="kpi-label">{label}</p>
                                <p class="kpi-value">{valor}</p>
                            </div>
                        """, unsafe_allow_html=True)

                st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

                st.markdown('<div class="sec-header-wrap"><p class="sec-header">Por Destino</p></div>', unsafe_allow_html=True)
                df_dest_s = (
                    df_saindo.groupby("Destino", as_index=False)
                    .agg(Reservas=("Reservas", "sum"), GMV=("GMV", "sum"))
                    .sort_values("GMV", ascending=False)
                )
                df_dest_s["% GMV"] = (df_dest_s["GMV"] / tot_g_s * 100).round(1)
                col_cs1, col_ts1 = st.columns([3, 2], gap="large")
                with col_cs1:
                    fig_ds = go.Figure(go.Bar(
                        x=df_dest_s["Destino"], y=df_dest_s["GMV"],
                        marker_color=ONFLY_BLUE,
                        text=[brl(v) for v in df_dest_s["GMV"]],
                        textposition="outside",
                        textfont=dict(size=10, color="#334155"),
                        hovertemplate="<b>%{x}</b><br>GMV: R$ %{y:,.2f}<extra></extra>",
                    ))
                    fig_ds.update_yaxes(tickprefix="R$ ")
                    st.plotly_chart(plotly_layout(fig_ds, 300), use_container_width=True)
                with col_ts1:
                    st.dataframe(
                        _brl_df(df_dest_s[["Destino", "Reservas", "GMV", "% GMV"]]),
                        use_container_width=True, hide_index=True, height=300,
                        column_config={
                            "GMV":      st.column_config.TextColumn("GMV"),
                            "Reservas": st.column_config.NumberColumn("Reservas", format="%d"),
                            "% GMV":    st.column_config.NumberColumn("% GMV",    format="%.1f%%"),
                        },
                    )

                st.markdown('<div class="sec-header-wrap"><p class="sec-header">Por Cia Aérea</p></div>', unsafe_allow_html=True)
                df_cia_s = (
                    df_saindo.groupby("Cia", as_index=False)
                    .agg(Reservas=("Reservas", "sum"), GMV=("GMV", "sum"))
                    .sort_values("GMV", ascending=False)
                )
                df_cia_s["% GMV"] = (df_cia_s["GMV"] / tot_g_s * 100).round(1)
                col_cs2, col_ts2 = st.columns([3, 2], gap="large")
                with col_cs2:
                    fig_cs = go.Figure(go.Bar(
                        x=df_cia_s["Cia"], y=df_cia_s["GMV"],
                        marker_color=ONFLY_BLUE,
                        text=[brl(v) for v in df_cia_s["GMV"]],
                        textposition="outside",
                        textfont=dict(size=10, color="#334155"),
                        hovertemplate="<b>%{x}</b><br>GMV: R$ %{y:,.2f}<extra></extra>",
                    ))
                    fig_cs.update_yaxes(tickprefix="R$ ")
                    st.plotly_chart(plotly_layout(fig_cs, 300), use_container_width=True)
                with col_ts2:
                    st.dataframe(
                        _brl_df(df_cia_s[["Cia", "Reservas", "GMV", "% GMV"]]),
                        use_container_width=True, hide_index=True, height=300,
                        column_config={
                            "GMV":      st.column_config.TextColumn("GMV"),
                            "Reservas": st.column_config.NumberColumn("Reservas", format="%d"),
                            "% GMV":    st.column_config.NumberColumn("% GMV",    format="%.1f%%"),
                        },
                    )

                st.markdown('<div class="sec-header-wrap"><p class="sec-header">Detalhe por Rota e Cia Aérea</p></div>', unsafe_allow_html=True)
                df_show_s = df_saindo.copy()
                df_show_s["% GMV"] = (df_show_s["GMV"] / tot_g_s * 100).round(1)
                st.dataframe(
                    _brl_df(df_show_s[["Origem", "Destino", "Cia", "Reservas", "GMV", "Ticket Médio", "% GMV"]]),
                    use_container_width=True, hide_index=True,
                    height=min(50 + len(df_show_s) * 35, 600),
                    column_config={
                        "GMV":          st.column_config.TextColumn("GMV"),
                        "Ticket Médio": st.column_config.TextColumn("Ticket Médio"),
                        "Reservas":     st.column_config.NumberColumn("Reservas",     format="%d"),
                        "% GMV":        st.column_config.NumberColumn("% GMV",        format="%.1f%%"),
                    },
                )


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO: AEROPORTOS
# ══════════════════════════════════════════════════════════════════════════════
elif secao == "🛬  Aeroportos":

    st.markdown('<div class="sec-header-wrap"><p class="sec-header">🛬 Aeroportos</p></div>',
                unsafe_allow_html=True)

    with st.spinner("Carregando ranking de aeroportos..."):
        try:
            df_ap_dest, df_ap_orig = q_aeroportos_ranking(i_str, f_str)
        except Exception as _e:
            st.error(f"Erro ao consultar BigQuery: {_e}")
            st.stop()

    if df_ap_dest.empty and df_ap_orig.empty:
        st.info("Nenhum dado encontrado para o período selecionado.")
        st.stop()

    _ap_col1, _ap_col2 = st.columns(2)

    # ── Ranking Destinos ──────────────────────────────────────────────────────
    with _ap_col1:
        st.markdown(
            "<p style='font-size:0.75rem;font-weight:700;text-transform:uppercase;"
            "letter-spacing:0.08em;color:#1890FF;margin-bottom:12px;'>✈️ Top Destinos</p>",
            unsafe_allow_html=True,
        )
        if df_ap_dest.empty:
            st.caption("Sem dados.")
        else:
            _df_dest_fmt = df_ap_dest.copy()
            _df_dest_fmt["GMV"] = _df_dest_fmt["GMV"] / 2          # ÷2: cada reserva conta em origem e destino
            _gmv_dest_tot = _df_dest_fmt["GMV"].sum()
            _bil_dest_tot = int(_df_dest_fmt["Bilhetes"].sum())
            _df_dest_fmt["GMV"] = _df_dest_fmt["GMV"].apply(brl)
            _tot_dest = pd.DataFrame([{
                "Nº": pd.NA, "IATA": "", "Aeroporto": "TOTAL",
                "Bilhetes": _bil_dest_tot, "GMV": brl(_gmv_dest_tot),
            }])
            _df_dest_fmt = pd.concat(
                [_df_dest_fmt[["Nº", "IATA", "Aeroporto", "Bilhetes", "GMV"]], _tot_dest],
                ignore_index=True,
            )
            st.dataframe(
                _df_dest_fmt,
                use_container_width=True,
                hide_index=True,
                height=min(50 + len(_df_dest_fmt) * 35, 600),
                column_config={
                    "Nº":        st.column_config.NumberColumn("Nº",        width="small"),
                    "IATA":      st.column_config.TextColumn("IATA",        width="small"),
                    "Aeroporto": st.column_config.TextColumn("Aeroporto"),
                    "Bilhetes":  st.column_config.NumberColumn("Bilhetes",  width="small"),
                    "GMV":       st.column_config.TextColumn("GMV"),
                },
            )

    # ── Ranking Origens ───────────────────────────────────────────────────────
    with _ap_col2:
        st.markdown(
            "<p style='font-size:0.75rem;font-weight:700;text-transform:uppercase;"
            "letter-spacing:0.08em;color:#FF6B00;margin-bottom:12px;'>🛫 Top Origens</p>",
            unsafe_allow_html=True,
        )
        if df_ap_orig.empty:
            st.caption("Sem dados.")
        else:
            _df_orig_fmt = df_ap_orig.copy()
            _df_orig_fmt["GMV"] = _df_orig_fmt["GMV"] / 2          # ÷2: cada reserva conta em origem e destino
            _gmv_orig_tot = _df_orig_fmt["GMV"].sum()
            _bil_orig_tot = int(_df_orig_fmt["Bilhetes"].sum())
            _df_orig_fmt["GMV"] = _df_orig_fmt["GMV"].apply(brl)
            _tot_orig = pd.DataFrame([{
                "Nº": pd.NA, "IATA": "", "Aeroporto": "TOTAL",
                "Bilhetes": _bil_orig_tot, "GMV": brl(_gmv_orig_tot),
            }])
            _df_orig_fmt = pd.concat(
                [_df_orig_fmt[["Nº", "IATA", "Aeroporto", "Bilhetes", "GMV"]], _tot_orig],
                ignore_index=True,
            )
            st.dataframe(
                _df_orig_fmt,
                use_container_width=True,
                hide_index=True,
                height=min(50 + len(_df_orig_fmt) * 35, 600),
                column_config={
                    "Nº":        st.column_config.NumberColumn("Nº",        width="small"),
                    "IATA":      st.column_config.TextColumn("IATA",        width="small"),
                    "Aeroporto": st.column_config.TextColumn("Aeroporto"),
                    "Bilhetes":  st.column_config.NumberColumn("Bilhetes",  width="small"),
                    "GMV":       st.column_config.TextColumn("GMV"),
                },
            )

# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO: BALANCEAMENTO DE EMISSÃO
# ══════════════════════════════════════════════════════════════════════════════
elif secao == "⚖️  Balanceamento":

    st.markdown('<div class="sec-header-wrap"><p class="sec-header">⚖️ Balanceamento de Emissão</p></div>',
                unsafe_allow_html=True)

    # ── Filtro de cias ────────────────────────────────────────────────────────
    with st.spinner("Carregando cias aéreas..."):
        _bal_cias_todas = q_cias_balanceamento(i_str, f_str)

    _bal_cias_sel = st.multiselect(
        "Cias Aéreas",
        options=_bal_cias_todas,
        default=_bal_cias_todas,
        placeholder="Selecione as cias...",
        label_visibility="collapsed",
        key="bal_cias",
    )

    # tuple vazia = sem filtro (todas) para otimizar a query
    _bal_filtro = tuple(sorted(_bal_cias_sel)) if len(_bal_cias_sel) < len(_bal_cias_todas) else ()

    st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)

    # ── Busca ─────────────────────────────────────────────────────────────────
    with st.spinner("Carregando balanceamento..."):
        try:
            df_bal = q_balanceamento(i_str, f_str, _bal_filtro)
        except Exception as _e:
            st.error(f"Erro ao consultar BigQuery: {_e}")
            st.stop()

    if df_bal.empty:
        st.info("Nenhum dado encontrado para o período selecionado.")
    else:
        _bal_total_gmv = df_bal["GMV"].sum()
        _bal_total_res = int(df_bal["Reservas"].sum())

        # ── helper: agrega por dimensão ──────────────────────────────────────
        def _bal_agg(col):
            return df_bal.groupby(col, as_index=False).agg(
                Reservas=("Reservas", "sum"), GMV=("GMV", "sum")
            ).assign(Pct=lambda d: (d["GMV"] / _bal_total_gmv * 100).round(1))

        # ═══ BLOCO 1 — Nacional vs Internacional ════════════════════════════
        st.markdown(
            "<p style='font-size:0.75rem;font-weight:700;text-transform:uppercase;"
            "letter-spacing:0.08em;color:#1890FF;margin-bottom:12px;'>🌍 Nacional vs Internacional</p>",
            unsafe_allow_html=True,
        )

        df_escopo = _bal_agg("Escopo").sort_values("GMV", ascending=False)
        _ncols = len(df_escopo)
        _cols_e = st.columns(_ncols)
        _cores_e = {"Nacional": ONFLY_BLUE, "Internacional": "#7C3AED"}
        for i, row in df_escopo.iterrows():
            _cor = _cores_e.get(row["Escopo"], ONFLY_BLUE)
            with _cols_e[list(df_escopo.index).index(i)]:
                st.markdown(
                    f'<div class="kpi-card" style="border-top-color:{_cor};">'
                    f'<p class="kpi-label">{row["Escopo"]}</p>'
                    f'<p class="kpi-value">{brl(row["GMV"])}</p>'
                    f'<p class="kpi-label" style="font-size:0.85rem;font-weight:700;color:{_cor};">'
                    f'{row["Pct"]:.1f}%</p>'
                    f'<p class="kpi-label">{row["Reservas"]:,} reservas</p>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        # Gráfico de barras horizontais — escopo
        _fig_e = px.bar(
            df_escopo, x="GMV", y="Escopo", orientation="h",
            color="Escopo",
            color_discrete_map=_cores_e,
            text=df_escopo["Pct"].apply(lambda v: f"{v:.1f}%"),
            labels={"GMV": "GMV (R$)", "Escopo": ""},
        )
        _fig_e.update_traces(textposition="outside")
        _fig_e.update_xaxes(tickprefix="R$ ")
        _fig_e.update_layout(showlegend=False)
        st.plotly_chart(plotly_layout(_fig_e, 180), use_container_width=True)

        st.markdown("<div style='height:32px'></div>", unsafe_allow_html=True)

        # ═══ BLOCO 2 — Manual vs Automático ═════════════════════════════════
        st.markdown(
            "<p style='font-size:0.75rem;font-weight:700;text-transform:uppercase;"
            "letter-spacing:0.08em;color:#FF6B00;margin-bottom:12px;'>🤖 Manual vs Automático</p>",
            unsafe_allow_html=True,
        )

        df_canal = _bal_agg("Canal").sort_values("GMV", ascending=False)
        _cores_c = {"Automático": ONFLY_GREEN, "Manual": ONFLY_ORANGE}
        _cols_c = st.columns(len(df_canal))
        for i, row in df_canal.iterrows():
            _cor = _cores_c.get(row["Canal"], ONFLY_BLUE)
            with _cols_c[list(df_canal.index).index(i)]:
                st.markdown(
                    f'<div class="kpi-card" style="border-top-color:{_cor};">'
                    f'<p class="kpi-label">{row["Canal"]}</p>'
                    f'<p class="kpi-value">{brl(row["GMV"])}</p>'
                    f'<p class="kpi-label" style="font-size:0.85rem;font-weight:700;color:{_cor};">'
                    f'{row["Pct"]:.1f}%</p>'
                    f'<p class="kpi-label">{row["Reservas"]:,} reservas</p>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        # Gráfico de barras horizontais — canal
        _fig_c = px.bar(
            df_canal, x="GMV", y="Canal", orientation="h",
            color="Canal",
            color_discrete_map=_cores_c,
            text=df_canal["Pct"].apply(lambda v: f"{v:.1f}%"),
            labels={"GMV": "GMV (R$)", "Canal": ""},
        )
        _fig_c.update_traces(textposition="outside")
        _fig_c.update_xaxes(tickprefix="R$ ")
        _fig_c.update_layout(showlegend=False)
        st.plotly_chart(plotly_layout(_fig_c, 180), use_container_width=True)

        st.markdown("<div style='height:32px'></div>", unsafe_allow_html=True)

        # ═══ BLOCO 3 — Matriz cruzada ════════════════════════════════════════
        st.markdown(
            "<p style='font-size:0.75rem;font-weight:700;text-transform:uppercase;"
            "letter-spacing:0.08em;color:#5A6475;margin-bottom:12px;'>📊 Detalhamento Cruzado</p>",
            unsafe_allow_html=True,
        )

        df_cross = df_bal.copy()
        # % dentro do próprio escopo (Nacional soma 100%, Internacional soma 100%)
        df_cross["% no Escopo"] = (
            df_cross.groupby("Escopo")["GMV"]
            .transform(lambda x: (x / x.sum() * 100).round(1))
        )
        # % sobre o total geral (todas as 4 células somam 100%)
        df_cross["% do Total"]  = (df_cross["GMV"] / _bal_total_gmv * 100).round(1)
        df_cross["GMV fmt"]     = df_cross["GMV"].apply(brl)
        df_cross = df_cross.sort_values(["Escopo", "Canal"], ascending=[False, True])
        st.dataframe(
            df_cross[["Escopo", "Canal", "Reservas", "GMV fmt", "% no Escopo", "% do Total"]].rename(
                columns={"GMV fmt": "GMV"}
            ),
            use_container_width=True,
            hide_index=True,
            column_config={
                "Escopo":      st.column_config.TextColumn("Escopo"),
                "Canal":       st.column_config.TextColumn("Canal"),
                "Reservas":    st.column_config.NumberColumn("Reservas",   format="%d"),
                "GMV":         st.column_config.TextColumn("GMV"),
                "% no Escopo": st.column_config.NumberColumn("% no Escopo", format="%.1f%%"),
                "% do Total":  st.column_config.NumberColumn("% do Total",  format="%.1f%%"),
            },
        )

        st.markdown("<div style='height:40px'></div>", unsafe_allow_html=True)

        # ═══ BLOCO 4 — Origem das emissões manuais ═══════════════════════════
        st.markdown(
            "<p style='font-size:0.75rem;font-weight:700;text-transform:uppercase;"
            "letter-spacing:0.08em;color:#5A6475;margin-bottom:12px;'>📡 Origem das Emissões Manuais</p>",
            unsafe_allow_html=True,
        )

        with st.spinner("Carregando canais manuais..."):
            try:
                df_canais = q_balanceamento_canais(i_str, f_str, _bal_filtro)
            except Exception as _ec:
                st.error(f"Erro ao consultar BigQuery: {_ec}")
                df_canais = pd.DataFrame()

        if df_canais.empty:
            st.info("Nenhuma emissão manual encontrada no período.")
        else:
            _col_nac, _col_int = st.columns(2, gap="large")

            for _col_graf, _escopo, _cor in [
                (_col_nac, "Nacional",      ONFLY_BLUE),
                (_col_int, "Internacional", "#7C3AED"),
            ]:
                _df_esc = df_canais[df_canais["Escopo"] == _escopo].copy()
                with _col_graf:
                    st.markdown(
                        f"<p style='font-size:0.78rem;font-weight:700;color:{_cor};"
                        f"margin-bottom:8px;'>{'🇧🇷' if _escopo == 'Nacional' else '🌍'} {_escopo}</p>",
                        unsafe_allow_html=True,
                    )
                    if _df_esc.empty:
                        st.caption("Sem emissões manuais.")
                    else:
                        _gmv_esc  = _df_esc["GMV"].sum()
                        _df_esc["Pct"] = (_df_esc["GMV"] / _gmv_esc * 100).round(1)
                        _df_esc = _df_esc.sort_values("GMV", ascending=True)

                        _fig_m = px.bar(
                            _df_esc, x="GMV", y="Canal", orientation="h",
                            text=_df_esc["Pct"].apply(lambda v: f"{v:.1f}%"),
                            color_discrete_sequence=[_cor],
                            labels={"GMV": "GMV (R$)", "Canal": ""},
                        )
                        _fig_m.update_traces(
                            textposition="outside",
                            marker_color=_cor,
                            hovertemplate="<b>%{y}</b><br>GMV: R$ %{x:,.2f}<extra></extra>",
                        )
                        _fig_m.update_xaxes(tickprefix="R$ ")
                        _fig_m.update_layout(showlegend=False)
                        _altura_m = max(180, 50 + len(_df_esc) * 36)
                        st.plotly_chart(plotly_layout(_fig_m, _altura_m), use_container_width=True)

        st.markdown("<div style='height:40px'></div>", unsafe_allow_html=True)

        # ═══ BLOCO 5 — Emissores manuais ════════════════════════════════════
        st.markdown(
            "<p style='font-size:0.75rem;font-weight:700;text-transform:uppercase;"
            "letter-spacing:0.08em;color:#5A6475;margin-bottom:12px;'>👤 Emissores Manuais</p>",
            unsafe_allow_html=True,
        )

        with st.spinner("Carregando emissores..."):
            try:
                df_emiss = q_balanceamento_emissores(i_str, f_str, _bal_filtro)
            except Exception as _ee:
                st.error(f"Erro ao consultar BigQuery: {_ee}")
                df_emiss = pd.DataFrame()

        if df_emiss.empty:
            st.info("Nenhum emissor manual encontrado no período.")
        else:
            _em_col_nac, _em_col_int = st.columns(2, gap="large")

            for _em_col, _escopo, _cor in [
                (_em_col_nac, "Nacional",      ONFLY_BLUE),
                (_em_col_int, "Internacional", "#7C3AED"),
            ]:
                _df_em = df_emiss[df_emiss["Escopo"] == _escopo].copy()
                with _em_col:
                    st.markdown(
                        f"<p style='font-size:0.78rem;font-weight:700;color:{_cor};"
                        f"margin-bottom:8px;'>{'🇧🇷' if _escopo == 'Nacional' else '🌍'} {_escopo}</p>",
                        unsafe_allow_html=True,
                    )
                    if _df_em.empty:
                        st.caption("Sem emissores manuais.")
                    else:
                        _gmv_em_tot = _df_em["GMV"].sum()
                        _df_em["% Escopo"] = (_df_em["GMV"] / _gmv_em_tot * 100).round(1)
                        _df_em["GMV"] = _df_em["GMV"].apply(brl)
                        st.dataframe(
                            _df_em[["Emissor", "Canal", "Reservas", "GMV", "% Escopo"]],
                            use_container_width=True,
                            hide_index=True,
                            height=min(50 + len(_df_em) * 35, 500),
                            column_config={
                                "Emissor":   st.column_config.TextColumn("Emissor"),
                                "Canal":     st.column_config.TextColumn("Canal",    width="small"),
                                "Reservas":  st.column_config.NumberColumn("Reservas", format="%d", width="small"),
                                "GMV":       st.column_config.TextColumn("GMV"),
                                "% Escopo":  st.column_config.NumberColumn("% Escopo", format="%.1f%%", width="small"),
                            },
                        )

# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO: CIA AÉREA LEGADO
# ══════════════════════════════════════════════════════════════════════════════
elif secao == "📋  Cia Aérea Legado":
    with st.spinner("Consultando BigQuery..."):
        try:
            df_leg = q_cia_legado(i_str, f_str)
        except Exception as e:
            st.error(f"Erro ao conectar ao BigQuery: {e}")
            st.stop()

    st.markdown('<div class="sec-header-wrap"><p class="sec-header">GMV por Cia Aérea — Legado</p></div>', unsafe_allow_html=True)

    if df_leg.empty:
        st.info("Nenhum dado encontrado para o período selecionado.")
    else:
        # Filtro por cia aérea
        opcoes_cia = ["Todas"] + [
            f"{row['Sigla']} — {row['Companhia']}" if row['Companhia'] != row['Sigla'] else row['Sigla']
            for _, row in df_leg.iterrows()
        ]
        col_busca, _ = st.columns([1, 2])
        with col_busca:
            cia_sel = st.selectbox("Cia Aérea", opcoes_cia, index=0, label_visibility="collapsed")
        if cia_sel != "Todas":
            sigla_sel = cia_sel.split(" — ")[0].strip()
            df_leg = df_leg[df_leg["Sigla"] == sigla_sel].reset_index(drop=True)

        gmv_total     = df_leg["GMV"].sum()
        bilhete_total = df_leg["GMV Incentivo"].sum()

        # KPIs de totais
        k1, k2, k3 = st.columns(3)
        k1.metric("Total Reservas",     f"{df_leg['Reservas'].sum():,.0f}".replace(",", "."))
        k2.metric("Total GMV",          f"R$ {gmv_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        k3.metric("Total GMV Incentivo",f"R$ {bilhete_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

        st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

        # Tabela principal
        df_show = df_leg.copy()
        df_show["% GMV"]       = (df_show["GMV"] / gmv_total * 100).round(1)
        df_show["GMV"]           = df_show["GMV"].apply(lambda x: f"R$ {x:,.2f}".replace(",","X").replace(".",",").replace("X","."))
        df_show["GMV Incentivo"] = df_show["GMV Incentivo"].apply(lambda x: f"R$ {x:,.2f}".replace(",","X").replace(".",",").replace("X","."))

        st.dataframe(
            df_show[["Sigla", "Companhia", "Reservas", "GMV", "GMV Incentivo", "% GMV"]],
            use_container_width=True,
            hide_index=True,
            column_config={
                "Sigla":          st.column_config.TextColumn("Sigla",         width="small"),
                "Companhia":      st.column_config.TextColumn("Companhia"),
                "Reservas":       st.column_config.NumberColumn("Reservas",    format="%d"),
                "GMV":            st.column_config.TextColumn("GMV"),
                "GMV Incentivo":  st.column_config.TextColumn("GMV Incentivo"),
                "% GMV":          st.column_config.NumberColumn("% GMV",       format="%.1f%%"),
            },
        )

        # Gráfico de barras duplo
        st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
        st.markdown('<div class="sec-header-wrap"><p class="sec-header">Top 15 — GMV vs GMV Incentivo</p></div>', unsafe_allow_html=True)
        df_top = df_leg.head(15).copy()
        df_melt = df_top.melt(id_vars=["Sigla"], value_vars=["GMV", "GMV Incentivo"], var_name="Tipo", value_name="Valor")
        fig = px.bar(
            df_melt,
            x="Sigla", y="Valor", color="Tipo", barmode="group",
            color_discrete_map={"GMV": ONFLY_BLUE, "GMV Incentivo": "#60A5FA"},
            labels={"Valor": "R$", "Sigla": ""},
        )
        fig.update_yaxes(tickprefix="R$ ")
        st.plotly_chart(plotly_layout(fig, 420), use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO: CLIENTES
# ══════════════════════════════════════════════════════════════════════════════
elif secao == "🏢  Clientes":

    # ── Ranking Geral ─────────────────────────────────────────────────────────
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">Ranking de Clientes</p></div>', unsafe_allow_html=True)

    with st.spinner("Carregando ranking..."):
        df_rank = q_clientes_ranking(i_str, f_str)

    if df_rank.empty:
        st.info("Nenhum dado encontrado para o período selecionado.")
    else:
        total_gmv      = df_rank["GMV"].sum()
        total_reservas = int(df_rank["Reservas"].sum())
        df_rank_show = df_rank.copy()
        df_rank_show.insert(0, "Nº", range(1, len(df_rank_show) + 1))
        df_rank_show["% GMV"] = (df_rank_show["GMV"] / total_gmv * 100).round(1).apply(lambda v: f"{v:.1f}%")
        df_rank_show["GMV"]          = df_rank_show["GMV"].apply(brl)
        df_rank_show["Ticket Médio"] = df_rank_show["Ticket Médio"].apply(brl)

        _tot_rank = pd.DataFrame([{
            "Nº": pd.NA, "Cliente": "TOTAL", "GMV": brl(total_gmv),
            "Reservas": total_reservas, "Ticket Médio": "", "% GMV": "100,0%",
        }])
        df_rank_show = pd.concat(
            [df_rank_show[["Nº", "Cliente", "GMV", "Reservas", "Ticket Médio", "% GMV"]], _tot_rank],
            ignore_index=True,
        )

        st.dataframe(
            df_rank_show,
            use_container_width=True,
            hide_index=True,
            height=min(50 + len(df_rank_show) * 35, 500),
        )

    st.markdown("<div style='height:48px'></div>", unsafe_allow_html=True)

    # ── Detalhe por Cliente ───────────────────────────────────────────────────
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">Detalhe por Cliente</p></div>', unsafe_allow_html=True)

    if not df_rank.empty:
        clientes_opcoes = [""] + df_rank.sort_values("Cliente")["Cliente"].tolist()
        cliente_sel = st.selectbox("Selecione o cliente", clientes_opcoes, index=0, label_visibility="collapsed")

        if not cliente_sel:
            st.info("Selecione um cliente acima para ver o detalhamento.")
        else:
            company_id_sel = int(df_rank.loc[df_rank["Cliente"] == cliente_sel, "company_id"].iloc[0])

            _info = q_cliente_info(company_id_sel)
            _cnpj = _info.get("cnpj", "")
            _cnpj_fmt = f"{_cnpj[:2]}.{_cnpj[2:5]}.{_cnpj[5:8]}/{_cnpj[8:12]}-{_cnpj[12:]}" if _cnpj and len(_cnpj) == 14 else _cnpj
            _social = _info.get("social_name", "")
            st.markdown(f"""
                <div style="display:flex;gap:12px;flex-wrap:wrap;margin:12px 0 4px 0;">
                    <div style="padding:8px 14px;background:#F0F7FF;border:1px solid #BAD7FF;
                                border-radius:8px;font-size:12px;color:#1890FF;">
                        <span style="font-weight:600;color:#5B6B7A;">ID Onfly</span>
                        &nbsp;<span style="font-weight:700;">{company_id_sel}</span>
                    </div>
                    <div style="padding:8px 14px;background:#F6F8FA;border:1px solid #E1E4E8;
                                border-radius:8px;font-size:12px;color:#24292E;">
                        <span style="font-weight:600;color:#5B6B7A;">CNPJ</span>
                        &nbsp;<span style="font-weight:700;">{_cnpj_fmt or '—'}</span>
                    </div>
                    {f'<div style="padding:8px 14px;background:#F6F8FA;border:1px solid #E1E4E8;border-radius:8px;font-size:12px;color:#24292E;"><span style="font-weight:600;color:#5B6B7A;">Razão Social</span>&nbsp;<span style="font-weight:700;">{_social}</span></div>' if _social else ''}
                </div>
            """, unsafe_allow_html=True)

            st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

            with st.spinner("Carregando dados do cliente..."):
                resumo_cli = df_rank.loc[df_rank["company_id"] == company_id_sel].iloc[0]
                df_cli_diario = q_cliente_diario(company_id_sel, i_str, f_str)
                df_cli_rotas  = q_cliente_rotas(company_id_sel, i_str, f_str)
                df_cli_cias   = q_cliente_cias(company_id_sel, i_str, f_str)

            # KPIs
            c1, c2, c3 = st.columns(3)
            with c1:
                st.markdown(f"""
                    <div class="kpi-card">
                        <p class="kpi-label">GMV Total</p>
                        <p class="kpi-value">{brl(resumo_cli["GMV"])}</p>
                    </div>""", unsafe_allow_html=True)
            with c2:
                st.markdown(f"""
                    <div class="kpi-card orange">
                        <p class="kpi-label">Reservas</p>
                        <p class="kpi-value">{int(resumo_cli["Reservas"]):,}</p>
                    </div>""", unsafe_allow_html=True)
            with c3:
                st.markdown(f"""
                    <div class="kpi-card green">
                        <p class="kpi-label">Ticket Médio</p>
                        <p class="kpi-value">{brl(resumo_cli["Ticket Médio"])}</p>
                    </div>""", unsafe_allow_html=True)

            st.markdown("<div style='height:48px'></div>", unsafe_allow_html=True)

            # Evolução Diária
            st.markdown('<div class="sec-header-wrap"><p class="sec-header">Evolução Diária</p></div>', unsafe_allow_html=True)
            if not df_cli_diario.empty:
                import plotly.graph_objects as go
                fig_d = go.Figure()
                fig_d.add_trace(go.Scatter(
                    x=df_cli_diario["Data"], y=df_cli_diario["GMV"],
                    mode="lines", fill="tozeroy",
                    line=dict(color=ONFLY_BLUE, width=2),
                    fillcolor=f"rgba(24,144,255,0.15)",
                    name="GMV",
                    hovertemplate="<b>%{x}</b><br>GMV: R$ %{y:,.2f}<extra></extra>",
                ))
                plotly_layout(fig_d)
                st.plotly_chart(fig_d, use_container_width=True)
            else:
                st.info("Sem dados de evolução diária para este cliente no período.")

            st.markdown("<div style='height:48px'></div>", unsafe_allow_html=True)

            # Cias Aéreas
            st.markdown('<div class="sec-header-wrap"><p class="sec-header">Cias Aéreas</p></div>', unsafe_allow_html=True)
            if not df_cli_cias.empty:
                import plotly.express as px
                total_gmv_cias = df_cli_cias["GMV"].sum()
                df_cli_cias["% GMV"] = (df_cli_cias["GMV"] / total_gmv_cias * 100).round(1)
                fig_cias = px.bar(
                    df_cli_cias.head(10), x="GMV", y="Cia", orientation="h",
                    color_discrete_sequence=[ONFLY_BLUE],
                    text=df_cli_cias.head(10)["% GMV"].apply(lambda v: f"{v:.1f}%"),
                )
                fig_cias.update_traces(textposition="outside")
                plotly_layout(fig_cias, height=300)
                st.plotly_chart(fig_cias, use_container_width=True)

                df_cli_cias_show = df_cli_cias.copy()
                df_cli_cias_show["GMV"]   = df_cli_cias_show["GMV"].apply(brl)
                df_cli_cias_show["% GMV"] = df_cli_cias_show["% GMV"].apply(lambda v: f"{v:.1f}%")
                st.dataframe(
                    _brl_df(df_cli_cias_show[["Cia", "Reservas", "GMV", "% GMV"]]),
                    use_container_width=True, hide_index=True,
                )

            st.markdown("<div style='height:48px'></div>", unsafe_allow_html=True)

            # Rotas
            st.markdown('<div class="sec-header-wrap"><p class="sec-header">Top Rotas</p></div>', unsafe_allow_html=True)
            if not df_cli_rotas.empty:
                total_gmv_rotas = df_cli_rotas["GMV"].sum()
                df_cli_rotas["% GMV"] = (df_cli_rotas["GMV"] / total_gmv_rotas * 100).round(1)
                df_cli_rotas["Rota"] = df_cli_rotas["Origem"] + " → " + df_cli_rotas["Destino"]
                df_cli_rotas_show = df_cli_rotas.copy()
                df_cli_rotas_show["GMV"]   = df_cli_rotas_show["GMV"].apply(brl)
                df_cli_rotas_show["% GMV"] = df_cli_rotas_show["% GMV"].apply(lambda v: f"{v:.1f}%")
                st.dataframe(
                    _brl_df(df_cli_rotas_show[["Rota", "Reservas", "GMV", "% GMV"]]),
                    use_container_width=True, hide_index=True,
                )
            else:
                st.info("Sem dados de rotas para este cliente no período.")

    # ── Todas as Emissões do Período ──────────────────────────────────────────
    st.markdown("<div style='height:48px'></div>", unsafe_allow_html=True)
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">📋 Todas as Emissões do Período</p></div>',
                unsafe_allow_html=True)

    with st.spinner("Carregando emissões..."):
        try:
            df_emissoes = q_clientes_emissoes(i_str, f_str)
        except Exception as _e:
            st.error(f"Erro ao carregar emissões: {_e}")
            df_emissoes = pd.DataFrame()

    if df_emissoes.empty:
        st.info("Nenhuma emissão encontrada no período.")
    elif len(df_emissoes) == 5000:
        st.caption("⚠️ Resultado limitado a 5.000 registros — refine o período ou use os filtros abaixo.")
    else:
        # Filtros
        _ef1, _ef2, _ef3 = st.columns([2, 2, 2])
        with _ef1:
            _e_cli_opts = ["Todos"] + sorted(df_emissoes["Cliente"].dropna().unique().tolist())
            _e_cli = st.selectbox("Filtrar por Cliente", _e_cli_opts, key="em_cli")
        with _ef2:
            _e_cons_opts = ["Todas"] + sorted(df_emissoes["Consolidadora"].dropna().unique().tolist())
            _e_cons = st.selectbox("Filtrar por Consolidadora", _e_cons_opts, key="em_cons")
        with _ef3:
            _e_emissor = st.text_input("Emissor", placeholder="Nome do emissor...", key="em_pax")

        _df_em_fil = df_emissoes.copy()
        if _e_cli != "Todos":
            _df_em_fil = _df_em_fil[_df_em_fil["Cliente"] == _e_cli]
        if _e_cons != "Todas":
            _df_em_fil = _df_em_fil[_df_em_fil["Consolidadora"] == _e_cons]
        if _e_emissor.strip():
            _df_em_fil = _df_em_fil[
                _df_em_fil["Emissor"].astype(str).str.contains(_e_emissor.strip(), case=False, na=False)
            ]

        # Paginação
        _EM_PAGE = 50
        _em_total = len(_df_em_fil)
        _em_pages = max(1, -(-_em_total // _EM_PAGE))

        _em_fkey = (_e_cli, _e_cons, _e_emissor.strip())
        if st.session_state.get("em_filtros_prev") != _em_fkey:
            st.session_state["em_page"] = 0
            st.session_state["em_filtros_prev"] = _em_fkey
        _em_page = st.session_state.get("em_page", 0)

        _em_start = _em_page * _EM_PAGE
        _em_end   = min(_em_start + _EM_PAGE, _em_total)
        _df_em_page = _df_em_fil.iloc[_em_start:_em_end].copy()

        _gmv_em = _df_em_fil["GMV"].sum()
        _df_em_page["GMV"] = _df_em_page["GMV"].apply(brl)

        st.dataframe(
            _df_em_page,
            use_container_width=True,
            hide_index=True,
            height=min(520, (_em_end - _em_start + 1) * 35 + 10),
        )

        # Rodapé
        _er1, _er2 = st.columns([3, 2])
        with _er1:
            st.markdown(
                f"<div style='font-size:0.88rem;font-weight:700;color:#0F172A;padding:6px 0 0 0;'>"
                f"Total GMV: <span style='color:#0EA5E9;font-size:1rem;'>{brl(_gmv_em)}</span>"
                f"&nbsp;&nbsp;·&nbsp;&nbsp;{_em_total} emissõe{'s' if _em_total != 1 else ''}"
                f"</div>",
                unsafe_allow_html=True,
            )
        with _er2:
            _ep1, _ep2, _ep3 = st.columns([1, 2, 1])
            with _ep1:
                if st.button("◀", key="em_prev", disabled=(_em_page == 0)):
                    st.session_state["em_page"] = _em_page - 1
                    st.rerun()
            with _ep2:
                st.markdown(
                    f"<div style='text-align:center;font-size:0.82rem;color:#64748B;padding-top:6px;'>"
                    f"Pág. {_em_page + 1} de {_em_pages}"
                    f"&nbsp;·&nbsp;{_em_start + 1}–{_em_end}"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            with _ep3:
                if st.button("▶", key="em_next", disabled=(_em_page >= _em_pages - 1)):
                    st.session_state["em_page"] = _em_page + 1
                    st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO: PRICING
# ══════════════════════════════════════════════════════════════════════════════
elif secao == "💲  Pricing":
    from calendar import monthrange
    import plotly.graph_objects as go

    MESES_NOME = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                  "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
    _anos = list(range(2020, hoje.year + 1))

    # ── Seleção de trecho ─────────────────────────────────────────────────────
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">Trecho</p></div>', unsafe_allow_html=True)

    with st.spinner("Carregando aeroportos..."):
        lista_ap = sorted(q_lista_aeroportos())

    col_orig, col_dest = st.columns(2)
    with col_orig:
        st.markdown("<p style='font-size:0.75rem;font-weight:600;color:#5A6475;margin-bottom:4px;'>Origem</p>", unsafe_allow_html=True)
        origem_sel = st.selectbox("Origem", [""] + lista_ap, index=0, label_visibility="collapsed", key="pricing_orig")
    with col_dest:
        st.markdown("<p style='font-size:0.75rem;font-weight:600;color:#5A6475;margin-bottom:4px;'>Destino</p>", unsafe_allow_html=True)
        destino_opts = [ap for ap in lista_ap if ap != origem_sel]
        destino_sel = st.selectbox("Destino", [""] + destino_opts, index=0, label_visibility="collapsed", key="pricing_dest")

    st.markdown("<div style='height:32px'></div>", unsafe_allow_html=True)

    # ── Seleção de período ────────────────────────────────────────────────────
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">Período de Comparação</p></div>', unsafe_allow_html=True)

    pc1, pc2 = st.columns(2)
    with pc1:
        st.markdown("<p style='font-size:0.75rem;font-weight:700;color:#1890FF;margin-bottom:4px;'>📅 Período 1</p>", unsafe_allow_html=True)
        pm1a, pm1b = st.columns(2)
        with pm1a:
            mes1_nome = st.selectbox("Mês 1", MESES_NOME, index=max(hoje.month-2, 0), label_visibility="collapsed", key="pr_m1")
        with pm1b:
            ano1 = st.selectbox("Ano 1", _anos, index=len(_anos)-1, label_visibility="collapsed", key="pr_a1")
    with pc2:
        st.markdown("<p style='font-size:0.75rem;font-weight:700;color:#FF6B00;margin-bottom:4px;'>📅 Período 2</p>", unsafe_allow_html=True)
        pm2a, pm2b = st.columns(2)
        with pm2a:
            mes2_nome = st.selectbox("Mês 2", MESES_NOME, index=hoje.month-1, label_visibility="collapsed", key="pr_m2")
        with pm2b:
            ano2 = st.selectbox("Ano 2", _anos, index=len(_anos)-1, label_visibility="collapsed", key="pr_a2")

    mes1 = MESES_NOME.index(mes1_nome) + 1
    mes2 = MESES_NOME.index(mes2_nome) + 1
    i1 = f"{ano1}-{mes1:02d}-01"
    f1 = f"{ano1}-{mes1:02d}-{monthrange(ano1, mes1)[1]:02d}"
    i2 = f"{ano2}-{mes2:02d}-01"
    f2 = f"{ano2}-{mes2:02d}-{monthrange(ano2, mes2)[1]:02d}"
    label1 = f"{mes1_nome[:3]}/{ano1}"
    label2 = f"{mes2_nome[:3]}/{ano2}"

    st.markdown("<div style='height:32px'></div>", unsafe_allow_html=True)

    if not origem_sel or not destino_sel:
        st.info("Selecione a origem e o destino para buscar os preços.")
    else:

        # ── Busca ─────────────────────────────────────────────────────────────
        with st.spinner(f"Buscando preços {origem_sel} → {destino_sel}..."):
            df_p1 = q_pricing(origem_sel, destino_sel, i1, f1)
            df_p2 = q_pricing(origem_sel, destino_sel, i2, f2)

        df_p1["Período"] = label1
        df_p2["Período"] = label2
        df_pricing = pd.concat([df_p1, df_p2], ignore_index=True)

        if df_pricing.empty:
            st.info(f"Nenhuma emissão encontrada para **{origem_sel} → {destino_sel}** nos períodos selecionados.")
        else:
            COR_P1, COR_P2 = ONFLY_BLUE, "#FF6B00"

            # ── KPIs lado a lado ──────────────────────────────────────────────
            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

            def _kpis(df):
                if df.empty: return None
                return {"medio": df["Preço (R$)"].mean(), "min": df["Preço (R$)"].min(),
                        "max": df["Preço (R$)"].max(), "n": len(df)}

            def _card_color(v2, v1, maior_melhor=False):
                if v1 is None or v1 == 0: return "#FFFFFF", ""
                diff = (v2 - v1) / v1 * 100
                subiu  = diff > 1
                desceu = diff < -1
                if subiu:
                    bg    = "#F0FFF4" if maior_melhor else "#FFF0F0"
                    badge = f"▲ {diff:.1f}%"
                elif desceu:
                    bg    = "#FFF0F0" if maior_melhor else "#F0FFF4"
                    badge = f"▼ {abs(diff):.1f}%"
                else:
                    bg, badge = "#FFFBEB", "— 0%"
                return bg, badge

            kpi1 = _kpis(df_p1)
            kpi2 = _kpis(df_p2)
            kc1, kc2 = st.columns(2)

            with kc1:
                if kpi1:
                    st.markdown(f"<p style='font-size:0.8rem;font-weight:700;color:{COR_P1};margin-bottom:8px;'>📅 {label1}</p>", unsafe_allow_html=True)
                    ka, kb, kc2_, kd = st.columns(4)
                    for _c, _l, _v in [(ka,"Médio",brl(kpi1["medio"])),(kb,"Mínimo",brl(kpi1["min"])),(kc2_,"Máximo",brl(kpi1["max"])),(kd,"Emissões",f"{kpi1['n']:,}")]:
                        with _c:
                            st.markdown(f'<div class="kpi-card"><p class="kpi-label" style="font-size:0.65rem;">{_l}</p><p class="kpi-value" style="font-size:0.85rem;">{_v}</p></div>', unsafe_allow_html=True)

            with kc2:
                if kpi2:
                    st.markdown(f"<p style='font-size:0.8rem;font-weight:700;color:{COR_P2};margin-bottom:8px;'>📅 {label2}</p>", unsafe_allow_html=True)
                    ka2, kb2, kc2_2, kd2 = st.columns(4)
                    metricas2 = [
                        (ka2,   "Médio",    brl(kpi2["medio"]), kpi2["medio"], kpi1["medio"] if kpi1 else None, False),
                        (kb2,   "Mínimo",   brl(kpi2["min"]),   kpi2["min"],   kpi1["min"]   if kpi1 else None, False),
                        (kc2_2, "Máximo",   brl(kpi2["max"]),   kpi2["max"],   kpi1["max"]   if kpi1 else None, False),
                        (kd2,   "Emissões", f"{kpi2['n']:,}",   kpi2["n"],     kpi1["n"]     if kpi1 else None, True),
                    ]
                    for _c, _l, _v, v2, v1, maior_melhor in metricas2:
                        bg, badge = _card_color(v2, v1, maior_melhor)
                        is_up = "▲" in badge
                        if maior_melhor:
                            badge_color = "#27AE60" if is_up else "#C0392B" if "▼" in badge else "#B7791F"
                        else:
                            badge_color = "#C0392B" if is_up else "#27AE60" if "▼" in badge else "#B7791F"
                        badge_html = f'<p style="font-size:0.68rem;font-weight:700;color:{badge_color};margin-top:2px;">{badge}</p>' if badge else ''
                        with _c:
                            st.markdown(f'<div class="kpi-card" style="background:{bg}!important;border-color:{bg}!important;"><p class="kpi-label" style="font-size:0.65rem;">{_l}</p><p class="kpi-value" style="font-size:0.85rem;">{_v}</p>{badge_html}</div>', unsafe_allow_html=True)

            st.markdown("<div style='height:48px'></div>", unsafe_allow_html=True)

            # ── Scatter comparativo ───────────────────────────────────────────
            st.markdown('<div class="sec-header-wrap"><p class="sec-header">Evolução de Preços — Comparativo</p></div>', unsafe_allow_html=True)

            fig_scatter = go.Figure()
            for df_kp, label_p, cor in [(df_p1, label1, COR_P1), (df_p2, label2, COR_P2)]:
                if df_kp.empty:
                    continue
                fig_scatter.add_trace(go.Scatter(
                    x=df_kp["Data Emissão"], y=df_kp["Preço (R$)"],
                    mode="markers", name=label_p,
                    marker=dict(color=cor, size=7, opacity=0.7),
                    hovertemplate=f"<b>{label_p}</b><br>Emissão: %{{x}}<br>Preço: R$ %{{y:,.2f}}<extra></extra>",
                ))
                df_md = df_kp.groupby("Data Emissão")["Preço (R$)"].mean().reset_index()
                fig_scatter.add_trace(go.Scatter(
                    x=df_md["Data Emissão"], y=df_md["Preço (R$)"],
                    mode="lines", name=f"Média {label_p}",
                    line=dict(color=cor, width=2, dash="dot"),
                    hovertemplate=f"Média {label_p}: R$ %{{y:,.2f}}<extra></extra>",
                ))

            plotly_layout(fig_scatter, height=380)
            st.plotly_chart(fig_scatter, use_container_width=True)

            st.markdown("<div style='height:48px'></div>", unsafe_allow_html=True)

            # ── Comparativo por Cia ───────────────────────────────────────────
            st.markdown('<div class="sec-header-wrap"><p class="sec-header">Comparativo por Cia Aérea</p></div>', unsafe_allow_html=True)

            def resumo_cia(df, label):
                if df.empty:
                    return pd.DataFrame()
                return (df.groupby("Cia", as_index=False)
                        .agg(**{
                            f"Médio {label}":   ("Preço (R$)", "mean"),
                            f"Mínimo {label}":  ("Preço (R$)", "min"),
                            f"Máximo {label}":  ("Preço (R$)", "max"),
                            f"Emissões {label}":("Preço (R$)", "count"),
                        }))

            df_r1 = resumo_cia(df_p1, label1)
            df_r2 = resumo_cia(df_p2, label2)

            if not df_r1.empty and not df_r2.empty:
                df_comp = df_r1.merge(df_r2, on="Cia", how="outer").fillna(0)
                df_comp["Variação (%)"] = df_comp.apply(
                    lambda r: round((r[f"Médio {label2}"] - r[f"Médio {label1}"]) / r[f"Médio {label1}"] * 100, 1)
                    if r[f"Médio {label1}"] > 0 else 0, axis=1
                )

                col_ch2, col_tb2 = st.columns([3, 2], gap="large")
                with col_ch2:
                    fig_comp = go.Figure()
                    for df_kp, label_p, cor in [(df_r1, label1, COR_P1), (df_r2, label2, COR_P2)]:
                        if df_kp.empty: continue
                        fig_comp.add_trace(go.Bar(
                            name=label_p, x=df_kp["Cia"],
                            y=df_kp[f"Médio {label_p}"].round(2),
                            marker_color=cor,
                            text=[brl(v) for v in df_kp[f"Médio {label_p}"]],
                            textposition="outside",
                            hovertemplate=f"<b>%{{x}}</b><br>{label_p}: R$ %{{y:,.2f}}<extra></extra>",
                        ))
                    fig_comp.update_layout(barmode="group")
                    fig_comp.update_yaxes(tickprefix="R$ ")
                    plotly_layout(fig_comp, height=340)
                    st.plotly_chart(fig_comp, use_container_width=True)

                with col_tb2:
                    df_comp_show = df_comp.copy()
                    for col_name in [f"Médio {label1}", f"Médio {label2}"]:
                        df_comp_show[col_name] = df_comp_show[col_name].apply(brl)
                    df_comp_show["Variação (%)"] = df_comp_show["Variação (%)"].apply(
                        lambda v: f"{'▲' if v > 0 else '▼' if v < 0 else '—'} {abs(v):.1f}%"
                    )
                    st.dataframe(
                        df_comp_show[["Cia", f"Médio {label1}", f"Médio {label2}", "Variação (%)"]],
                        use_container_width=True, hide_index=True, height=340,
                    )
            else:
                for df_kp, label_p in [(df_r1, label1), (df_r2, label2)]:
                    if not df_kp.empty:
                        st.markdown(f"**{label_p}**")
                        st.dataframe(_brl_df(df_kp), use_container_width=True, hide_index=True)

            st.markdown("<div style='height:48px'></div>", unsafe_allow_html=True)

            # ── Tabela detalhada ──────────────────────────────────────────────
            st.markdown('<div class="sec-header-wrap"><p class="sec-header">Emissões Detalhadas</p></div>', unsafe_allow_html=True)

            df_det = df_pricing.copy()
            df_det["Preço (R$)"] = df_det["Preço (R$)"].apply(brl)
            st.dataframe(
                df_det[["Período", "Data Emissão", "Data Voo", "Cia", "Trecho", "Saída", "Chegada", "Preço (R$)"]],
                use_container_width=True, hide_index=True,
                height=min(50 + len(df_det) * 35, 500),
            )

# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO: QUEM VOA O QUE?
# ══════════════════════════════════════════════════════════════════════════════
elif secao == "🛫  Quem voa o que?":

    st.markdown('<div class="sec-header-wrap"><p class="sec-header">Quem voa o que?</p></div>', unsafe_allow_html=True)

    # ── Carrega lista de cias ────────────────────────────────────────────────
    with st.spinner("Carregando companhias aéreas..."):
        try:
            lista_cias_qvq = q_lista_cias()
        except Exception as _e:
            st.error(f"Erro ao conectar ao BigQuery: {_e}")
            st.stop()

    _opcoes_qvq = {
        cia: f"{cia} — {IATA_NOMES.get(cia, 'Não identificada')}"
        for cia in lista_cias_qvq
    }

    _col_sel_qvq, _ = st.columns([2, 5])
    with _col_sel_qvq:
        st.markdown("<p style='font-size:0.75rem;font-weight:600;color:#5A6475;margin-bottom:4px;'>Companhia Aérea</p>", unsafe_allow_html=True)
        _cia_qvq = st.selectbox(
            "Companhia",
            options=[""] + list(_opcoes_qvq.keys()),
            format_func=lambda x: "Selecione uma companhia..." if x == "" else _opcoes_qvq[x],
            label_visibility="collapsed",
            key="qvq_cia_sel",
        )

    if not _cia_qvq:
        st.info("Selecione uma companhia aérea para ver quais clientes utilizaram no período.")
    else:
        _variantes_qvq = lista_cias_qvq[_cia_qvq]

        st.markdown(f"""
            <h3 style="margin:16px 0 4px 0; color:#0F172A; font-size:1.2rem; font-weight:800;">
                {_opcoes_qvq[_cia_qvq]}
            </h3>
        """, unsafe_allow_html=True)

        with st.spinner(f"Carregando clientes de {_cia_qvq}..."):
            try:
                df_qvq = q_quem_voa_o_que(_variantes_qvq, i_str, f_str)
            except Exception as _e:
                st.error(f"Erro ao consultar BigQuery: {_e}")
                st.stop()

        if df_qvq.empty:
            st.info("Nenhum cliente encontrou voos nessa companhia no período selecionado.")
        else:
            # ── KPIs rápidos ─────────────────────────────────────────────────
            _qvq_c1, _qvq_c2, _qvq_c3 = st.columns(3)
            with _qvq_c1:
                st.markdown(f"""
                    <div class="kpi-card">
                        <p class="kpi-label">Clientes</p>
                        <p class="kpi-value">{len(df_qvq):,}</p>
                    </div>
                """, unsafe_allow_html=True)
            with _qvq_c2:
                st.markdown(f"""
                    <div class="kpi-card orange">
                        <p class="kpi-label">Total de Trechos</p>
                        <p class="kpi-value">{int(df_qvq['Trechos'].sum()):,}</p>
                    </div>
                """, unsafe_allow_html=True)
            with _qvq_c3:
                st.markdown(f"""
                    <div class="kpi-card green">
                        <p class="kpi-label">GMV Total</p>
                        <p class="kpi-value">{brl(df_qvq['GMV'].sum())}</p>
                    </div>
                """, unsafe_allow_html=True)

            st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)

            # ── Tabela ───────────────────────────────────────────────────────
            _qvq_header_col, _qvq_btn_col = st.columns([6, 1])
            with _qvq_header_col:
                st.markdown('<div class="sec-header-wrap"><p class="sec-header">Clientes no período</p></div>', unsafe_allow_html=True)
            with _qvq_btn_col:
                _periodo_qvq = f"{i_str} a {f_str}"
                _cia_nome_qvq = _opcoes_qvq[_cia_qvq]
                _excel_bytes = _gerar_excel_top15(df_qvq, _cia_nome_qvq, _periodo_qvq)
                _fname_qvq = f"top15_{_cia_qvq}_{i_str}_{f_str}.xlsx".replace(" ", "_")
                st.download_button(
                    label="📥 Top 15",
                    data=_excel_bytes,
                    file_name=_fname_qvq,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    help="Exportar os 15 maiores clientes por GMV em Excel (pronto para e-mail)",
                )

            df_qvq_show = df_qvq.copy()
            df_qvq_show.insert(0, "Nº", range(1, len(df_qvq_show) + 1))
            df_qvq_show["GMV"] = df_qvq_show["GMV"].apply(brl)

            st.dataframe(
                df_qvq_show[["Nº", "Razão Social", "Nome Fantasia", "CNPJ", "GMV", "Trechos", "Trecho Mais Voado"]],
                use_container_width=True,
                hide_index=True,
                height=min(50 + len(df_qvq_show) * 35, 600),
                column_config={
                    "Nº":               st.column_config.NumberColumn("Nº", width="small"),
                    "Razão Social":     st.column_config.TextColumn("Razão Social"),
                    "Nome Fantasia":    st.column_config.TextColumn("Nome Fantasia"),
                    "CNPJ":             st.column_config.TextColumn("CNPJ"),
                    "GMV":              st.column_config.TextColumn("GMV"),
                    "Trechos":          st.column_config.NumberColumn("Trechos", width="small"),
                    "Trecho Mais Voado": st.column_config.TextColumn("Trecho Mais Voado"),
                },
            )

            # ── Totalizador ──────────────────────────────────────────────────
            _qvq_total = df_qvq["GMV"].sum()
            st.markdown(
                f"<div style='text-align:right;font-size:0.85rem;font-weight:700;"
                f"color:#1890FF;margin-top:8px;'>"
                f"GMV Total: {brl(_qvq_total)}"
                f"</div>",
                unsafe_allow_html=True,
            )

# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO: ANÁLISES COM IA
# ══════════════════════════════════════════════════════════════════════════════
elif secao == "🔍  Análises":

    # ── Cabeçalho ─────────────────────────────────────────────────────────────
    _ah_col, _ah_btn = st.columns([8, 1])
    with _ah_col:
        st.markdown('<div class="sec-header-wrap"><p class="sec-header">Análises com IA</p></div>', unsafe_allow_html=True)
    with _ah_btn:
        if st.button("🗑️", help="Limpar histórico", use_container_width=True):
            st.session_state.analises_history = []
            _history_save([])
            st.rerun()

    if _get_anthropic_client() is None:
        st.warning(
            "⚠️ **API key da Anthropic não configurada.**\n\n"
            "Adicione `ANTHROPIC_API_KEY` no arquivo `~/.streamlit/secrets.toml`:\n\n"
            "```toml\nANTHROPIC_API_KEY = \"sk-ant-...\"\n```\n\n"
            "Ou defina a variável de ambiente `ANTHROPIC_API_KEY` antes de iniciar o app."
        )
    else:
        # ── Inicializa histórico (carrega do disco na primeira vez) ───────────
        if "analises_history" not in st.session_state:
            st.session_state.analises_history = _history_load()

        # ── Input de pergunta ──────────────────────────────────────────────────
        nova_pergunta = st.chat_input("Digite sua pergunta sobre os dados do período selecionado...")

        if nova_pergunta:
            with st.spinner("🤖 Consultando dados..."):
                try:
                    resultado = _analises_query(nova_pergunta, i_str, f_str)
                    st.session_state.analises_history.insert(0, resultado)
                except Exception as _e:
                    resultado = {
                        "pergunta": nova_pergunta,
                        "sql": "",
                        "resumo": "",
                        "df": pd.DataFrame(),
                        "analise": f"❌ Erro ao processar a pergunta: {_e}",
                    }
                    st.session_state.analises_history.insert(0, resultado)
            _history_save(st.session_state.analises_history)
            st.rerun()

        # ── Histórico de análises ─────────────────────────────────────────────
        if not st.session_state.analises_history:
            st.markdown(
                "<div style='text-align:center;padding:48px 0;color:#888;'>"
                "💬 Nenhuma análise ainda. Digite uma pergunta acima para começar."
                "</div>",
                unsafe_allow_html=True,
            )
        else:
            for _i, _entry in enumerate(st.session_state.analises_history):
                _label = f"🙋 {_entry['pergunta']}"
                with st.expander(_label, expanded=False):
                    # Análise textual
                    st.markdown(_entry["analise"])
                    # Tabela de resultados
                    if not _entry["df"].empty:
                        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
                        st.dataframe(_brl_df(_entry["df"]), use_container_width=True, hide_index=True, height=min(50 + len(_entry["df"]) * 35, 400))
                    # SQL expansível
                    if _entry.get("sql"):
                        with st.expander("🔎 Ver SQL gerado"):
                            st.code(_entry["sql"], language="sql")

# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO: TENDÊNCIA
# ══════════════════════════════════════════════════════════════════════════════
elif secao == "📈  Tendência":
    from calendar import monthrange as _mr
    import numpy as np

    _hoje_t   = date.today()
    _mes_at   = _hoje_t.month
    _ano_at   = _hoje_t.year
    _mes_ant  = _mes_at - 1 if _mes_at > 1 else 12
    _ano_ant  = _ano_at if _mes_at > 1 else _ano_at - 1

    _MESES_PT = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                 "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
    _label_at  = f"{_MESES_PT[_mes_at-1]} {_ano_at}"
    _label_ant = f"{_MESES_PT[_mes_ant-1]} {_ano_ant}"
    _total_dias_at  = _mr(_ano_at, _mes_at)[1]
    _total_dias_ant = _mr(_ano_ant, _mes_ant)[1]

    with st.spinner("Carregando tendências..."):
        _df_at  = q_tendencia_diario(_mes_at, _ano_at)
        _df_ant = q_tendencia_diario(_mes_ant, _ano_ant)

    # ── Cálculos de projeção (dias úteis D-1 — exclui sáb/dom e dia atual) ──
    if not _df_at.empty:
        # Exclui o dia atual (D-1)
        _df_at_d1   = _df_at[_df_at["dia_num"] < _hoje_t.day]
        _gmv_so_far = float(_df_at_d1["gmv"].sum()) if not _df_at_d1.empty else float(_df_at["gmv"].sum())
        _gr_so_far  = float(_df_at["gross_revenue"].sum())
        _df_base    = _df_at_d1 if not _df_at_d1.empty else _df_at
        # Dias úteis com dados até D-1
        _uteis_passados_t = sum(
            1 for d in _df_base["dia_num"].unique()
            if date(_ano_at, _mes_at, int(d)).weekday() < 5
        )
        # Total de dias úteis no mês atual
        _uteis_mes_t = sum(
            1 for d in range(1, _total_dias_at + 1)
            if date(_ano_at, _mes_at, d).weekday() < 5
        )
        _dias_com_dados = _uteis_passados_t or len(_df_at)
        _proj_gmv       = (_gmv_so_far / _uteis_passados_t * _uteis_mes_t) if _uteis_passados_t > 0 else 0
        _proj_tr        = (_gr_so_far / float(_df_at["gmv"].sum()) * 100) if float(_df_at["gmv"].sum()) > 0 else 0
    else:
        _dias_com_dados = 0
        _gmv_so_far = _gr_so_far = _proj_gmv = _proj_tr = 0.0

    if not _df_ant.empty:
        _gmv_ant_total = float(_df_ant["gmv"].sum())
        _gr_ant_total  = float(_df_ant["gross_revenue"].sum())
        _tr_ant        = (_gr_ant_total / _gmv_ant_total * 100) if _gmv_ant_total > 0 else 0
    else:
        _gmv_ant_total = _gr_ant_total = _tr_ant = 0.0

    _var_gmv = ((_proj_gmv - _gmv_ant_total) / _gmv_ant_total * 100) if _gmv_ant_total > 0 else 0
    _var_tr  = _proj_tr - _tr_ant

    # ── Função auxiliar: montar série cumulativa + projeção ───────────────────
    def _build_series(df_atual, df_anterior, col, total_dias_at, total_dias_ant, proj_total):
        """Retorna (dias_at, cum_at, dias_ant, cum_ant, dias_proj, cum_proj)."""
        # Mês anterior — cumulativo completo
        if not df_anterior.empty:
            s_ant = df_anterior.groupby("dia_num")[col].sum().reindex(range(1, total_dias_ant+1), fill_value=0)
            cum_ant = s_ant.cumsum().tolist()
            dias_ant = list(range(1, total_dias_ant+1))
        else:
            dias_ant, cum_ant = [], []
        # Mês atual — cumulativo real
        if not df_atual.empty:
            s_at = df_atual.groupby("dia_num")[col].sum().reindex(range(1, total_dias_at+1), fill_value=0)
            cum_at_full = s_at.cumsum()
            last_day = int(df_atual["dia_num"].max())
            cum_at = cum_at_full[:last_day].tolist()
            dias_at = list(range(1, last_day+1))
            # Projeção: linha reta do último ponto ao valor projetado
            ultimo_val = cum_at[-1]
            dias_proj  = list(range(last_day, total_dias_at+1))
            cum_proj   = list(np.linspace(ultimo_val, proj_total, len(dias_proj)))
        else:
            dias_at, cum_at = [], []
            dias_proj, cum_proj = [], []
        return dias_at, cum_at, dias_ant, cum_ant, dias_proj, cum_proj

    # ── ══ BLOCO GMV ══ ───────────────────────────────────────────────────────
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">GMV — Tendência</p></div>', unsafe_allow_html=True)

    # Badge de sinal
    _badge_gmv_cor   = "#27AE60" if _var_gmv >= 0 else "#C0392B"
    _badge_gmv_emoji = "🟢" if _var_gmv >= 0 else "🔴"
    _badge_gmv_txt   = f"{'Acima' if _var_gmv >= 0 else 'Abaixo'} do mês anterior  ({'%+.1f' % _var_gmv}%)"
    st.markdown(
        f"<div style='background:{_badge_gmv_cor}18;border-left:4px solid {_badge_gmv_cor};"
        f"padding:10px 16px;border-radius:6px;margin-bottom:16px;font-weight:600;color:{_badge_gmv_cor};'>"
        f"{_badge_gmv_emoji} {_badge_gmv_txt}</div>",
        unsafe_allow_html=True,
    )

    # KPI cards GMV
    _kg1, _kg2, _kg3 = st.columns(3)
    with _kg1:
        st.markdown(f'<div class="kpi-card"><p class="kpi-label">GMV Projetado ({_label_at})</p><p class="kpi-value">{brl(_proj_gmv)}</p></div>', unsafe_allow_html=True)
    with _kg2:
        st.markdown(f'<div class="kpi-card"><p class="kpi-label">GMV Real até hoje</p><p class="kpi-value">{brl(_gmv_so_far)}</p></div>', unsafe_allow_html=True)
    with _kg3:
        _vc = "#27AE60" if _var_gmv >= 0 else "#C0392B"
        _vs = "▲" if _var_gmv >= 0 else "▼"
        st.markdown(f'<div class="kpi-card"><p class="kpi-label">vs {_label_ant}</p><p class="kpi-value" style="color:{_vc};">{_vs} {abs(_var_gmv):.1f}%</p><p style="font-size:0.8rem;color:#888;">{brl(_gmv_ant_total)}</p></div>', unsafe_allow_html=True)

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

    # Gráfico GMV
    _da, _ca, _dant, _cant, _dp, _cp = _build_series(_df_at, _df_ant, "gmv", _total_dias_at, _total_dias_ant, _proj_gmv)
    _fig_gmv = go.Figure()
    if _dant:
        _fig_gmv.add_trace(go.Scatter(x=_dant, y=_cant, mode="lines", name=_label_ant,
            line=dict(color="#94A3B8", width=2, dash="dash"),
            hovertemplate="Dia %{x}<br>Acum.: R$ %{y:,.0f}<extra></extra>"))
    if _da:
        _fig_gmv.add_trace(go.Scatter(x=_da, y=_ca, mode="lines", name=f"{_label_at} (real)",
            line=dict(color=ONFLY_BLUE, width=3),
            hovertemplate="Dia %{x}<br>Acum.: R$ %{y:,.0f}<extra></extra>"))
    if _dp:
        _fig_gmv.add_trace(go.Scatter(x=_dp, y=_cp, mode="lines", name="Projeção",
            line=dict(color=ONFLY_BLUE, width=2, dash="dot"),
            hovertemplate="Dia %{x}<br>Projeção: R$ %{y:,.0f}<extra></extra>"))
    plotly_layout(_fig_gmv, height=340)
    _fig_gmv.update_layout(xaxis_title="Dia do mês", yaxis_title="GMV Acumulado (R$)",
                           legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
    st.plotly_chart(_fig_gmv, use_container_width=True)

    st.markdown("<div style='height:48px'></div>", unsafe_allow_html=True)

    # ── ══ BLOCO TAKE RATE ══ ─────────────────────────────────────────────────
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">Take Rate — Tendência</p></div>', unsafe_allow_html=True)

    # Badge de sinal Take Rate
    _badge_tr_cor   = "#27AE60" if _var_tr >= 0 else "#C0392B"
    _badge_tr_emoji = "🟢" if _var_tr >= 0 else "🔴"
    _badge_tr_txt   = f"{'Acima' if _var_tr >= 0 else 'Abaixo'} do mês anterior  ({'%+.2f' % _var_tr} pp)"
    st.markdown(
        f"<div style='background:{_badge_tr_cor}18;border-left:4px solid {_badge_tr_cor};"
        f"padding:10px 16px;border-radius:6px;margin-bottom:16px;font-weight:600;color:{_badge_tr_cor};'>"
        f"{_badge_tr_emoji} {_badge_tr_txt}</div>",
        unsafe_allow_html=True,
    )

    # KPI cards Take Rate
    _kt1, _kt2, _kt3 = st.columns(3)
    with _kt1:
        st.markdown(f'<div class="kpi-card"><p class="kpi-label">Take Rate Projetado ({_label_at})</p><p class="kpi-value">{_proj_tr:.2f}%</p></div>', unsafe_allow_html=True)
    with _kt2:
        _tr_atual_real = (_gr_so_far / _gmv_so_far * 100) if _gmv_so_far > 0 else 0
        st.markdown(f'<div class="kpi-card"><p class="kpi-label">Take Rate Real até hoje</p><p class="kpi-value">{_tr_atual_real:.2f}%</p></div>', unsafe_allow_html=True)
    with _kt3:
        _vtc = "#27AE60" if _var_tr >= 0 else "#C0392B"
        _vts = "▲" if _var_tr >= 0 else "▼"
        st.markdown(f'<div class="kpi-card"><p class="kpi-label">vs {_label_ant}</p><p class="kpi-value" style="color:{_vtc};">{_vts} {abs(_var_tr):.2f} pp</p><p style="font-size:0.8rem;color:#888;">{_tr_ant:.2f}%</p></div>', unsafe_allow_html=True)

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

    # Gráfico Take Rate (cumulativo ponderado: gross_revenue / gmv acumulados)
    def _cum_tr(df, total_dias):
        if df.empty:
            return [], []
        s_gmv = df.groupby("dia_num")["gmv"].sum().reindex(range(1, total_dias+1), fill_value=0).cumsum()
        s_gr  = df.groupby("dia_num")["gross_revenue"].sum().reindex(range(1, total_dias+1), fill_value=0).cumsum()
        tr_cum = (s_gr / s_gmv.replace(0, float("nan")) * 100).tolist()
        return list(range(1, total_dias+1)), tr_cum

    _fig_tr = go.Figure()
    _dant_tr, _cant_tr = _cum_tr(_df_ant, _total_dias_ant)
    if _dant_tr:
        _fig_tr.add_trace(go.Scatter(x=_dant_tr, y=_cant_tr, mode="lines", name=_label_ant,
            line=dict(color="#94A3B8", width=2, dash="dash"),
            hovertemplate="Dia %{x}<br>TR Acum.: %{y:.2f}%<extra></extra>"))

    if not _df_at.empty:
        _last_day_tr = int(_df_at["dia_num"].max())
        _s_gmv_at = _df_at.groupby("dia_num")["gmv"].sum().reindex(range(1, _last_day_tr+1), fill_value=0).cumsum()
        _s_gr_at  = _df_at.groupby("dia_num")["gross_revenue"].sum().reindex(range(1, _last_day_tr+1), fill_value=0).cumsum()
        _tr_cum_at = (_s_gr_at / _s_gmv_at.replace(0, float("nan")) * 100).tolist()
        _dias_tr_at = list(range(1, _last_day_tr+1))
        _fig_tr.add_trace(go.Scatter(x=_dias_tr_at, y=_tr_cum_at, mode="lines", name=f"{_label_at} (real)",
            line=dict(color=ONFLY_BLUE, width=3),
            hovertemplate="Dia %{x}<br>TR Acum.: %{y:.2f}%<extra></extra>"))
        # Linha de projeção (horizontal até fim do mês no valor atual)
        _dias_proj_tr = list(range(_last_day_tr, _total_dias_at+1))
        _proj_tr_vals = [_proj_tr] * len(_dias_proj_tr)
        _fig_tr.add_trace(go.Scatter(x=_dias_proj_tr, y=_proj_tr_vals, mode="lines", name="Projeção",
            line=dict(color=ONFLY_BLUE, width=2, dash="dot"),
            hovertemplate="Dia %{x}<br>TR Projetado: %{y:.2f}%<extra></extra>"))

    plotly_layout(_fig_tr, height=340)
    _fig_tr.update_layout(xaxis_title="Dia do mês", yaxis_title="Take Rate Acumulado (%)",
                          legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
    st.plotly_chart(_fig_tr, use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
elif secao == "🔵  Azul":
    _az_dados = load_incentivo_azul()

    # ── Seletor de ano ──
    _az_col_ano, _ = st.columns([1, 5])
    with _az_col_ano:
        _az_anos = list(range(2024, hoje.year + 2))
        _az_ano  = st.selectbox("Ano", _az_anos,
                                index=_az_anos.index(hoje.year) if hoje.year in _az_anos else len(_az_anos)-1,
                                key="azul_ano")
    _az_ano_str = str(_az_ano)
    _az_dados_ano = _az_dados.get(_az_ano_str, {})

    # ══ helper: renderiza um bloco (Nacional ou Internacional) ══
    def _render_azul_bloco(bloco_key: str, titulo: str, pcts: list, labels: list):
        st.markdown(
            f'<div class="sec-header-wrap"><p class="sec-header">AZUL — {titulo}</p></div>',
            unsafe_allow_html=True,
        )
        niv_info = "  |  ".join(f"**{lbl}**" for lbl in labels)
        st.markdown(
            f"<p style='font-size:0.78rem;color:#64748B;margin-bottom:8px;'>Níveis: {niv_info}</p>",
            unsafe_allow_html=True,
        )
        st.caption("Preencha **Realizado** e os **thresholds de cada nível**. "
                   "% Incentivo e Pago atualizam automaticamente a cada alteração.")

        dados_bloco = _az_dados_ano.get(bloco_key, {})
        df_az = _azul_build_df(dados_bloco, labels)

        # ── Editor de entradas (sem colunas calculadas) ──
        col_cfg: dict = {
            "MÊS":       st.column_config.TextColumn("MÊS", disabled=True, width="small"),
            "Realizado": st.column_config.NumberColumn("Realizado", format="R$ %,.0f", min_value=0),
        }
        for lbl in labels:
            col_cfg[lbl] = st.column_config.NumberColumn(lbl, format="R$ %,.0f", min_value=0)

        col_order = ["MÊS", "Realizado"] + labels
        edited_az = st.data_editor(
            df_az[col_order],
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            column_config=col_cfg,
            key=f"azul_{bloco_key}_{_az_ano}",
        )

        # ── Resultados calculados em tempo real ──
        df_res = _azul_compute_results(edited_az, pcts, labels)
        _az_total_real = df_res["Realizado"].sum()
        _az_total_pago = df_res["Pago"].sum()

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        st.dataframe(
            df_res.style
                .format({"Realizado": lambda v: brl(v) if v else "—",
                         "Pago":      lambda v: brl(v) if v else "—"})
                .apply(lambda row: [
                    "background-color:#F0FDF4;font-weight:600" if row["Nível"] != "—" else ""
                ] * len(row), axis=1),
            use_container_width=True,
            hide_index=True,
        )

        # ── KPI totais ──
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        _az_tc1, _az_tc2, _az_tc3, _ = st.columns([1, 1, 1, 3])
        with _az_tc1:
            st.markdown(f'<div class="kpi-card"><p class="kpi-label">Total Realizado</p>'
                        f'<p class="kpi-value">{brl(_az_total_real)}</p></div>', unsafe_allow_html=True)
        with _az_tc2:
            st.markdown(f'<div class="kpi-card"><p class="kpi-label">Total Incentivo</p>'
                        f'<p class="kpi-value">{brl(_az_total_pago)}</p></div>', unsafe_allow_html=True)
        with _az_tc3:
            _az_pct_med = (_az_total_pago / _az_total_real * 100) if _az_total_real else 0
            st.markdown(f'<div class="kpi-card"><p class="kpi-label">% Médio Obtido</p>'
                        f'<p class="kpi-value">{_az_pct_med:.2f}%</p></div>', unsafe_allow_html=True)

        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

        if st.button(f"💾  Salvar {titulo}", type="primary", key=f"save_azul_{bloco_key}"):
            if _az_ano_str not in _az_dados:
                _az_dados[_az_ano_str] = {}
            bloco_salvo: dict = {}
            for _, r in edited_az.iterrows():
                mes  = r["MÊS"]
                real = float(r["Realizado"]) if r["Realizado"] is not None and not (isinstance(r["Realizado"], float) and pd.isna(r["Realizado"])) else None
                nivs = []
                for lbl in labels:
                    v = r[lbl]
                    nivs.append(float(v) if v is not None and not (isinstance(v, float) and pd.isna(v)) else None)
                bloco_salvo[mes] = {"realizado": real, "niveis": nivs}
            _az_dados[_az_ano_str][bloco_key] = bloco_salvo
            save_incentivo_azul(_az_dados)
            st.success(f"✅  {titulo} {_az_ano} salvo com sucesso!")
            st.rerun()

    _render_azul_bloco("nacional",       "Nacional",       AZUL_NAC_PCTS, AZUL_NAC_LABELS)
    st.markdown("<div style='height:32px'></div>", unsafe_allow_html=True)
    _render_azul_bloco("internacional",  "Internacional",  AZUL_INT_PCTS, AZUL_INT_LABELS)

# ══════════════════════════════════════════════════════════════════════════════
elif secao == "🔴  LATAM":
    _lt_dados = load_incentivo_latam()

    # ── Seletor de ano ──
    _lt_col_ano, _ = st.columns([1, 5])
    with _lt_col_ano:
        _lt_anos = list(range(2024, hoje.year + 2))
        _lt_ano  = st.selectbox("Ano", _lt_anos,
                                index=_lt_anos.index(hoje.year) if hoje.year in _lt_anos else len(_lt_anos)-1,
                                key="latam_ano")
    _lt_ano_str   = str(_lt_ano)
    _lt_dados_ano = _lt_dados.get(_lt_ano_str, {})

    def _fv_lt(v):
        return float(v) if v is not None and not (isinstance(v, float) and pd.isna(v)) else None

    # ══ BLOCO NACIONAL ════════════════════════════════════════════════════════
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">LATAM — Nacional</p></div>', unsafe_allow_html=True)
    st.markdown(
        f"<p style='font-size:0.78rem;color:#64748B;margin:4px 0 8px 0;'>"
        f"Parâmetros fixos: <b>Taxa Ref.</b> {LATAM_TAXA_REF*100:.1f}% &nbsp;|&nbsp; "
        f"<b>Fixo Corp</b> {LATAM_FIXO_CORP*100:.1f}% &nbsp;|&nbsp; "
        f"<b>Peso Receita</b> {int(LATAM_PESO_REC*100)}% &nbsp;|&nbsp; "
        f"<b>Peso Prm Rvn</b> {int(LATAM_PESO_PRM*100)}%</p>",
        unsafe_allow_html=True,
    )
    st.caption("Dados fornecidos pela LATAM mensalmente. Resultados atualizam automaticamente.")

    _lt_nac_rows = []
    for _mes in MESES_ABREV:
        d = _lt_dados_ano.get(_mes, {}).get("nacional", {})
        _lt_nac_rows.append({
            "MÊS":              _mes,
            "Meta Receita":     d.get("meta_receita"),
            "Real. MTD":        d.get("realizado_mtd"),
            "Real. Proj.":      d.get("realizado_proj"),
            "Baseline Prm Rvn": d.get("baseline_prm_rvn"),
            "Real. Prm Rvn":    d.get("realizado_prm_rvn"),
            "Taxa Real":        d.get("taxa_real"),
        })

    _lt_nac_cfg = {
        "MÊS":              st.column_config.TextColumn("MÊS", disabled=True, width="small"),
        "Meta Receita":     st.column_config.NumberColumn("Meta Receita",     format="R$ %,.0f", min_value=0),
        "Real. MTD":        st.column_config.NumberColumn("Real. MTD",        format="R$ %,.0f", min_value=0),
        "Real. Proj.":      st.column_config.NumberColumn("Real. Proj.",       format="R$ %,.0f", min_value=0),
        "Baseline Prm Rvn": st.column_config.NumberColumn("Baseline Prm Rvn", format="%.1f%%",   min_value=0, max_value=100),
        "Real. Prm Rvn":    st.column_config.NumberColumn("Real. Prm Rvn",    format="%.1f%%",   min_value=0, max_value=100),
        "Taxa Real":        st.column_config.NumberColumn("Taxa Real",         format="%.2f%%",   min_value=0, max_value=100),
    }
    _lt_nac_edited = st.data_editor(
        pd.DataFrame(_lt_nac_rows), use_container_width=True, hide_index=True,
        num_rows="fixed", column_config=_lt_nac_cfg, key=f"latam_nac_{_lt_ano}",
    )

    # Resultados Nacional
    _lt_nac_res, _lt_nac_total = [], 0.0
    for _, _r in _lt_nac_edited.iterrows():
        _calc = _latam_calcular(
            _r["Meta Receita"], _r["Real. Proj."],
            (_r["Baseline Prm Rvn"] / 100) if _r["Baseline Prm Rvn"] else None,
            (_r["Real. Prm Rvn"]    / 100) if _r["Real. Prm Rvn"]    else None,
            (_r["Taxa Real"]        / 100) if _r["Taxa Real"]        else None,
        )
        if _calc:
            _lt_nac_total += _calc["rs_incent"]
            _lt_nac_res.append({"MÊS": _r["MÊS"],
                "% Cumpr. Receita": f"{_calc['pct_rec']*100:.1f}%",
                "% Cumpr. Prm Rvn": f"{_calc['pct_prm']*100:.1f}%",
                "% Cumpr. Final":   f"{_calc['pct_final']*100:.1f}%",
                "Fixo Corp": f"{LATAM_FIXO_CORP*100:.1f}%",
                "Variável":  f"{_calc['variavel']*100:.2f}%",
                "% Incentivo": f"{_calc['pct_incent']*100:.2f}%",
                "R$ Incentivo": _calc["rs_incent"]})
        else:
            _lt_nac_res.append({"MÊS": _r["MÊS"],
                "% Cumpr. Receita": "—", "% Cumpr. Prm Rvn": "—",
                "% Cumpr. Final": "—", "Fixo Corp": "—",
                "Variável": "—", "% Incentivo": "—", "R$ Incentivo": None})

    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
    st.dataframe(pd.DataFrame(_lt_nac_res).style.format(
        {"R$ Incentivo": lambda v: brl(v) if v else "—"}),
        use_container_width=True, hide_index=True)

    _lnk1, _lnk2, _lnk3, _ = st.columns([1, 1, 1, 3])
    _lt_nac_meses = [r for r in _lt_nac_res if r["R$ Incentivo"]]
    _lt_nac_pct_med = (sum(float(r["% Incentivo"].replace("%","")) for r in _lt_nac_meses) / len(_lt_nac_meses)) if _lt_nac_meses else 0
    with _lnk1:
        st.markdown(f'<div class="kpi-card"><p class="kpi-label">Total Incentivo Nac.</p><p class="kpi-value">{brl(_lt_nac_total)}</p></div>', unsafe_allow_html=True)
    with _lnk2:
        st.markdown(f'<div class="kpi-card"><p class="kpi-label">% Médio Nac.</p><p class="kpi-value">{_lt_nac_pct_med:.2f}%</p></div>', unsafe_allow_html=True)
    with _lnk3:
        st.markdown(f'<div class="kpi-card"><p class="kpi-label">Meses Preenchidos</p><p class="kpi-value">{len(_lt_nac_meses)} / 12</p></div>', unsafe_allow_html=True)

    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    if st.button("💾  Salvar Nacional", type="primary", key="save_latam_nac"):
        if _lt_ano_str not in _lt_dados: _lt_dados[_lt_ano_str] = {}
        for _, _r in _lt_nac_edited.iterrows():
            _mes = _r["MÊS"]
            if _mes not in _lt_dados[_lt_ano_str]: _lt_dados[_lt_ano_str][_mes] = {}
            _lt_dados[_lt_ano_str][_mes]["nacional"] = {
                "meta_receita":     _fv_lt(_r["Meta Receita"]),
                "realizado_mtd":    _fv_lt(_r["Real. MTD"]),
                "realizado_proj":   _fv_lt(_r["Real. Proj."]),
                "baseline_prm_rvn": _fv_lt(_r["Baseline Prm Rvn"]),
                "realizado_prm_rvn":_fv_lt(_r["Real. Prm Rvn"]),
                "taxa_real":        _fv_lt(_r["Taxa Real"]),
            }
        save_incentivo_latam(_lt_dados)
        st.success("✅  Nacional salvo!")
        st.rerun()

    # ══ BLOCO INTERNACIONAL ═══════════════════════════════════════════════════
    st.markdown("<div style='height:32px'></div>", unsafe_allow_html=True)
    st.markdown('<div class="sec-header-wrap"><p class="sec-header">LATAM — Internacional</p></div>', unsafe_allow_html=True)
    st.markdown(
        f"<p style='font-size:0.78rem;color:#64748B;margin:4px 0 8px 0;'>"
        f"Parâmetros fixos: <b>Fixo Corp</b> {LATAM_FIXO_INTER*100:.1f}% + variável (interpolação linear na faixa). "
        f"Ranges mudam mensalmente.</p>",
        unsafe_allow_html=True,
    )
    st.caption("Selecione o mês, informe o Realizado e os ranges fornecidos pela LATAM. Resultado atualiza automaticamente.")

    # ── Seletor de mês ──
    _lt_int_col_mes, _ = st.columns([1, 5])
    with _lt_int_col_mes:
        _lt_int_mes = st.selectbox("Mês", MESES_ABREV,
                                   index=min(hoje.month - 1, 11),
                                   key="latam_inter_mes")

    _lt_int_dados_mes = _lt_dados_ano.get(_lt_int_mes, {}).get("internacional", {})

    # ── helper: renderiza sub-bloco (eu_ca ou outros) ──
    def _render_inter_sub(sub_key: str, titulo: str, n_ranges: int, key_sfx: str):
        st.markdown(f"**{titulo}**", unsafe_allow_html=False)
        sub = _lt_int_dados_mes.get(sub_key, {})
        real_saved = sub.get("realizado")
        ranges_saved = sub.get("ranges", [{}] * n_ranges)
        while len(ranges_saved) < n_ranges:
            ranges_saved.append({})

        # Realizado
        real_val = st.number_input(
            f"Realizado {titulo} (R$)", value=float(real_saved) if real_saved else 0.0,
            min_value=0.0, step=1000.0, format="%.0f",
            key=f"latam_inter_real_{sub_key}_{_lt_int_mes}_{_lt_ano}",
        )

        # Ranges
        range_rows = []
        for i, r in enumerate(ranges_saved):
            range_rows.append({
                "Range":     f"Range {i+1}",
                "Mínimo":    r.get("min"),
                "Máximo":    r.get("max"),
                "Taxa Mín%": r.get("taxa_min"),
                "Taxa Máx%": r.get("taxa_max"),
            })
        rng_cfg = {
            "Range":     st.column_config.TextColumn("Range", disabled=True, width="small"),
            "Mínimo":    st.column_config.NumberColumn("Mínimo",    format="R$ %,.0f", min_value=0),
            "Máximo":    st.column_config.NumberColumn("Máximo",    format="R$ %,.0f", min_value=0),
            "Taxa Mín%": st.column_config.NumberColumn("Taxa Mín%", format="%.2f%%",   min_value=0, max_value=100),
            "Taxa Máx%": st.column_config.NumberColumn("Taxa Máx%", format="%.2f%%",   min_value=0, max_value=100),
        }
        edited_rng = st.data_editor(
            pd.DataFrame(range_rows), use_container_width=True, hide_index=True,
            num_rows="fixed", column_config=rng_cfg,
            key=f"latam_inter_rng_{sub_key}_{_lt_int_mes}_{_lt_ano}",
        )

        # Calcula em tempo real
        ranges_list = []
        for _, rr in edited_rng.iterrows():
            ranges_list.append({
                "min":      rr["Mínimo"],   "max":      rr["Máximo"],
                "taxa_min": rr["Taxa Mín%"], "taxa_max": rr["Taxa Máx%"],
            })
        calc = _latam_inter_calcular(real_val if real_val > 0 else None, ranges_list)

        # Resultado inline
        if calc:
            _rc1, _rc2, _rc3 = st.columns(3)
            with _rc1:
                st.markdown(f'<div class="kpi-card"><p class="kpi-label">Variável</p>'
                            f'<p class="kpi-value">{calc["variavel_pct"]:.2f}%</p></div>', unsafe_allow_html=True)
            with _rc2:
                st.markdown(f'<div class="kpi-card"><p class="kpi-label">% Incentivo</p>'
                            f'<p class="kpi-value">{calc["pct_incent"]:.2f}%</p>'
                            f'<p style="font-size:0.75rem;color:#888;">Fixo {LATAM_FIXO_INTER*100:.1f}% + Var. {calc["variavel_pct"]:.2f}%</p></div>',
                            unsafe_allow_html=True)
            with _rc3:
                st.markdown(f'<div class="kpi-card"><p class="kpi-label">R$ Incentivo</p>'
                            f'<p class="kpi-value">{brl(calc["rs_incent"])}</p></div>', unsafe_allow_html=True)
        else:
            st.caption("Preencha o Realizado e os Ranges para ver o resultado.")

        return real_val, ranges_list, calc

    col_euca, col_outros = st.columns(2)
    with col_euca:
        _lt_real_euca, _lt_rng_euca, _lt_calc_euca = _render_inter_sub("eu_ca",  "INT EU+CA", 7, "euca")
    with col_outros:
        _lt_real_out,  _lt_rng_out,  _lt_calc_out  = _render_inter_sub("outros", "INT Outros", 3, "outros")

    # ── KPIs totais Internacional ──
    _lt_int_total = (_lt_calc_euca["rs_incent"] if _lt_calc_euca else 0) + \
                    (_lt_calc_out["rs_incent"]  if _lt_calc_out  else 0)
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    _lik1, _lik2, _ = st.columns([1, 1, 4])
    with _lik1:
        st.markdown(f'<div class="kpi-card"><p class="kpi-label">Total Internacional</p>'
                    f'<p class="kpi-value">{brl(_lt_int_total)}</p></div>', unsafe_allow_html=True)
    with _lik2:
        _lt_grand_total = _lt_nac_total + _lt_int_total
        st.markdown(f'<div class="kpi-card"><p class="kpi-label">Total LATAM (Nac + Inter)</p>'
                    f'<p class="kpi-value">{brl(_lt_grand_total)}</p></div>', unsafe_allow_html=True)

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
    if st.button("💾  Salvar Internacional", type="primary", key="save_latam_inter"):
        if _lt_ano_str not in _lt_dados: _lt_dados[_lt_ano_str] = {}
        if _lt_int_mes not in _lt_dados[_lt_ano_str]: _lt_dados[_lt_ano_str][_lt_int_mes] = {}
        _lt_dados[_lt_ano_str][_lt_int_mes]["internacional"] = {
            "eu_ca":  {"realizado": float(_lt_real_euca) if _lt_real_euca else None,
                       "ranges":   _lt_rng_euca},
            "outros": {"realizado": float(_lt_real_out)  if _lt_real_out  else None,
                       "ranges":   _lt_rng_out},
        }
        save_incentivo_latam(_lt_dados)
        st.success(f"✅  Internacional {_lt_int_mes}/{_lt_ano} salvo!")
        st.rerun()

# ── 📝 Anotações Importantes ────────────────────────────────────────────────
elif secao == "📝  Anotações":
    import uuid as _uuid
    from datetime import datetime as _dt

    # ── Cabeçalho ─────────────────────────────────────────────────────────────
    _an_hcol, _an_ncol = st.columns([5, 1])
    with _an_hcol:
        st.markdown('<div class="sec-header-wrap"><p class="sec-header">📝 Anotações Importantes</p></div>',
                    unsafe_allow_html=True)
    with _an_ncol:
        st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)
        if st.button("➕  Nova Nota", use_container_width=True, key="an_nova"):
            st.session_state["an_editing"]  = True
            st.session_state["an_edit_id"]  = None
            st.rerun()

    # ── Session state ──────────────────────────────────────────────────────────
    if "an_editing" not in st.session_state:
        st.session_state["an_editing"] = False
    if "an_edit_id" not in st.session_state:
        st.session_state["an_edit_id"] = None

    anotacoes = load_anotacoes()

    # ── Formulário (nova ou edição) ────────────────────────────────────────────
    if st.session_state["an_editing"]:
        _edit_id = st.session_state["an_edit_id"]
        _nota_exist = next((n for n in anotacoes if n["id"] == _edit_id), {}) if _edit_id else {}

        with st.container(border=True):
            _fhcol, _fxcol = st.columns([8, 1])
            with _fhcol:
                st.markdown(
                    f"<p style='font-size:0.85rem;font-weight:700;color:#0F172A;margin:0 0 12px 0;'>"
                    f"{'✏️  Editar nota' if _nota_exist else '✏️  Nova nota'}</p>",
                    unsafe_allow_html=True,
                )
            with _fxcol:
                if st.button("✖️", key="an_fechar", use_container_width=True):
                    st.session_state["an_editing"] = False
                    st.session_state["an_edit_id"] = None
                    st.rerun()

            _titulo_val  = _nota_exist.get("titulo", "")
            _conteudo_val = _nota_exist.get("conteudo", "")

            _novo_titulo   = st.text_input("Título", value=_titulo_val, placeholder="Ex: Acordo com LATAM 2026...", key="an_titulo")
            _novo_conteudo = st.text_area("Conteúdo", value=_conteudo_val, placeholder="Escreva sua anotação aqui...", height=200, key="an_conteudo")

            _sc1, _sc2 = st.columns([2, 8])
            with _sc1:
                if st.button("💾  Salvar", type="primary", use_container_width=True, key="an_salvar"):
                    if not _novo_titulo.strip():
                        st.warning("O título não pode ser vazio.")
                    else:
                        agora = _dt.now().strftime("%Y-%m-%d %H:%M")
                        if _nota_exist:
                            for _n in anotacoes:
                                if _n["id"] == _edit_id:
                                    _n["titulo"]       = _novo_titulo.strip()
                                    _n["conteudo"]     = _novo_conteudo.strip()
                                    _n["atualizado_em"] = agora
                                    break
                        else:
                            anotacoes.insert(0, {
                                "id":           str(_uuid.uuid4()),
                                "titulo":       _novo_titulo.strip(),
                                "conteudo":     _novo_conteudo.strip(),
                                "criado_em":    agora,
                                "atualizado_em": agora,
                            })
                        save_anotacoes(anotacoes)
                        st.session_state["an_editing"] = False
                        st.session_state["an_edit_id"] = None
                        st.success("✅  Nota salva!")
                        st.rerun()

    # ── Lista de anotações ─────────────────────────────────────────────────────
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    if not anotacoes:
        st.info("Nenhuma anotação ainda. Clique em **➕ Nova Nota** para começar.")
    else:
        for _nota in anotacoes:
            _label_exp = (
                f"**{_nota['titulo']}**  "
                f"<span style='font-size:0.68rem;color:#94A3B8;'>"
                f"— {_nota.get('atualizado_em','')}</span>"
            )
            with st.expander(_nota["titulo"], expanded=False):
                st.markdown(
                    f"<p style='font-size:0.68rem;color:#94A3B8;margin:0 0 10px 0;'>"
                    f"Criado em {_nota.get('criado_em','')}  ·  Atualizado em {_nota.get('atualizado_em','')}</p>",
                    unsafe_allow_html=True,
                )
                _conteudo_exib = _nota.get("conteudo", "").strip()
                if _conteudo_exib:
                    st.markdown(
                        f"<p style='font-size:0.9rem;color:#334155;white-space:pre-wrap;margin:0 0 14px 0;'>"
                        f"{_conteudo_exib}</p>",
                        unsafe_allow_html=True,
                    )
                _nbtn1, _nbtn2, _ = st.columns([1, 1, 6])
                with _nbtn1:
                    if st.button("✏️  Editar", key=f"an_edit_{_nota['id']}", use_container_width=True):
                        st.session_state["an_editing"] = True
                        st.session_state["an_edit_id"] = _nota["id"]
                        st.rerun()
                with _nbtn2:
                    if st.button("🗑️  Excluir", key=f"an_del_{_nota['id']}", use_container_width=True):
                        anotacoes = [n for n in anotacoes if n["id"] != _nota["id"]]
                        save_anotacoes(anotacoes)
                        st.rerun()

# ── 🏭 Consolidadores ────────────────────────────────────────────────────────
elif secao == "🏭  Consolidadores":

    st.markdown('<div class="sec-header-wrap"><p class="sec-header">🏭 Consolidadores</p></div>',
                unsafe_allow_html=True)

    # Carrega lista de consolidadores antes das abas (usada nas duas)
    with st.spinner("Carregando consolidadores..."):
        try:
            df_cons = q_consolidadores_lista(i_str, f_str)
        except Exception as _e:
            st.error(f"Erro ao carregar dados: {_e}")
            st.stop()

    if df_cons.empty:
        st.info("Nenhuma emissão encontrada no período selecionado.")
        st.stop()

    _tab_op, _tab_evo = st.tabs(["Operação", "Evolução"])

    # ══════════════════════════════════════════════════════════════════════════
    # ABA: OPERAÇÃO
    # ══════════════════════════════════════════════════════════════════════════
    with _tab_op:

        # KPIs totais
        _ct1, _ct2, _ct3 = st.columns(3)
        with _ct1:
            st.markdown(f'<div class="kpi-card"><p class="kpi-label">GMV Total</p>'
                        f'<p class="kpi-value">{brl(df_cons["GMV"].sum())}</p></div>',
                        unsafe_allow_html=True)
        with _ct2:
            st.markdown(f'<div class="kpi-card"><p class="kpi-label">Total de Reservas</p>'
                        f'<p class="kpi-value">{int(df_cons["Reservas"].sum()):,}</p></div>',
                        unsafe_allow_html=True)
        with _ct3:
            st.markdown(f'<div class="kpi-card"><p class="kpi-label">Consolidadores</p>'
                        f'<p class="kpi-value">{len(df_cons)}</p></div>',
                        unsafe_allow_html=True)

        st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

        # Tabela ranking
        _df_cons_fmt = df_cons.copy()
        _df_cons_fmt.insert(0, "Nº", range(1, len(_df_cons_fmt) + 1))
        _df_cons_fmt["GMV"]          = _df_cons_fmt["GMV"].apply(brl)
        _df_cons_fmt["Ticket Médio"] = _df_cons_fmt["Ticket Médio"].apply(brl)
        st.dataframe(_df_cons_fmt, use_container_width=True, hide_index=True, height=280)

        # Gráfico de barras
        import plotly.express as px
        _df_bar = df_cons.head(15).copy()
        _fig_bar = px.bar(
            _df_bar, x="GMV", y="Consolidador", orientation="h",
            text=_df_bar["GMV"].apply(brl),
            color_discrete_sequence=["#0EA5E9"],
        )
        _fig_bar.update_traces(textposition="outside")
        plotly_layout(_fig_bar, height=max(300, len(_df_bar) * 32))
        _fig_bar.update_layout(yaxis=dict(autorange="reversed"), xaxis_title="", yaxis_title="")
        st.plotly_chart(_fig_bar, use_container_width=True)

        # ── Detalhe por consolidador ───────────────────────────────────────────
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        st.markdown('<div class="sec-header-wrap"><p class="sec-header">Detalhe por Consolidador</p></div>',
                    unsafe_allow_html=True)

        _cons_opcoes = df_cons["Consolidador"].tolist()
        _cons_sel = st.multiselect(
            "Selecione um ou mais consolidadores (vazio = todos)",
            options=_cons_opcoes,
            placeholder="Todos os consolidadores...",
            key="cons_sel_det",
        )
        # None = todos; lista = filtro múltiplo
        _cons_filtro = _cons_sel if _cons_sel else None

        _spinner_label = ", ".join(_cons_sel) if _cons_sel else "todos os consolidadores"
        df_cias_c = None
        if not _cons_sel:
            st.info("☝️ Selecione um ou mais consolidadores acima para carregar o detalhamento de cias, clientes, rotas e voos emitidos.")
        else:
            with st.spinner(f"Carregando detalhes de {_spinner_label}..."):
                try:
                    df_cias_c     = q_consolidador_cias(_cons_filtro, i_str, f_str)
                    df_clientes_c = q_consolidador_clientes(_cons_filtro, i_str, f_str)
                    df_rotas_c    = q_consolidador_rotas(_cons_filtro, i_str, f_str)
                    df_voos_c     = q_consolidador_voos(_cons_filtro, i_str, f_str)
                except Exception as _e:
                    st.error(f"Erro: {_e}")

        if df_cias_c is not None:
            # KPIs — soma dos selecionados (ou total geral)
            if _cons_sel:
                _df_kpi = df_cons[df_cons["Consolidador"].isin(_cons_sel)]
            else:
                _df_kpi = df_cons
            _kpi_gmv      = _df_kpi["GMV"].sum()
            _kpi_reservas = int(_df_kpi["Reservas"].sum())
            _kpi_tm       = _kpi_gmv / _kpi_reservas if _kpi_reservas else 0

            _ck1, _ck2, _ck3 = st.columns(3)
            with _ck1:
                st.markdown(f'<div class="kpi-card"><p class="kpi-label">GMV</p>'
                            f'<p class="kpi-value">{brl(_kpi_gmv)}</p></div>',
                            unsafe_allow_html=True)
            with _ck2:
                st.markdown(f'<div class="kpi-card"><p class="kpi-label">Reservas</p>'
                            f'<p class="kpi-value">{_kpi_reservas:,}</p></div>',
                            unsafe_allow_html=True)
            with _ck3:
                st.markdown(f'<div class="kpi-card"><p class="kpi-label">Ticket Médio</p>'
                            f'<p class="kpi-value">{brl(_kpi_tm)}</p></div>',
                            unsafe_allow_html=True)

            st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

            # Três colunas: Cias | Clientes | Rotas
            _col_cias, _col_cli, _col_rot = st.columns(3)

            with _col_cias:
                st.markdown("<p style='font-size:0.72rem;font-weight:700;text-transform:uppercase;"
                            "letter-spacing:0.07em;color:#8C9BAB;margin-bottom:8px;'>✈️ Cias Aéreas</p>",
                            unsafe_allow_html=True)
                if df_cias_c.empty:
                    st.caption("Sem dados.")
                else:
                    _df_cias_fmt = df_cias_c.copy()
                    _df_cias_fmt["GMV"]          = _df_cias_fmt["GMV"].apply(brl)
                    _df_cias_fmt["Ticket Médio"] = _df_cias_fmt["Ticket Médio"].apply(brl)
                    st.dataframe(_df_cias_fmt, use_container_width=True, hide_index=True,
                                 height=min(400, (len(_df_cias_fmt) + 1) * 35 + 10))

            with _col_cli:
                st.markdown("<p style='font-size:0.72rem;font-weight:700;text-transform:uppercase;"
                            "letter-spacing:0.07em;color:#8C9BAB;margin-bottom:8px;'>🏢 Clientes</p>",
                            unsafe_allow_html=True)
                if df_clientes_c.empty:
                    st.caption("Sem dados.")
                else:
                    _df_cli_fmt = df_clientes_c.copy()
                    _df_cli_fmt["GMV"] = _df_cli_fmt["GMV"].apply(brl)
                    st.dataframe(_df_cli_fmt, use_container_width=True, hide_index=True,
                                 height=min(400, (len(_df_cli_fmt) + 1) * 35 + 10))

            with _col_rot:
                st.markdown("<p style='font-size:0.72rem;font-weight:700;text-transform:uppercase;"
                            "letter-spacing:0.07em;color:#8C9BAB;margin-bottom:8px;'>🗺️ Rotas</p>",
                            unsafe_allow_html=True)
                if df_rotas_c.empty:
                    st.caption("Sem dados.")
                else:
                    _df_rot_fmt = df_rotas_c.copy()
                    _df_rot_fmt["GMV"] = _df_rot_fmt["GMV"].apply(brl)
                    st.dataframe(_df_rot_fmt, use_container_width=True, hide_index=True,
                                 height=min(400, (len(_df_rot_fmt) + 1) * 35 + 10))

            # ── Voos emitidos ──────────────────────────────────────────────────────
            st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)
            st.markdown('<div class="sec-header-wrap"><p class="sec-header">✈️ Voos Emitidos</p></div>',
                        unsafe_allow_html=True)

            if df_voos_c.empty:
                st.info("Nenhum voo encontrado no período.")
            else:
                # Filtros rápidos inline
                _fv1, _fv2, _fv3 = st.columns([2, 2, 2])
                with _fv1:
                    _cias_opcoes = ["Todas"] + sorted(df_voos_c["Cia"].dropna().unique().tolist())
                    _cia_filtro = st.selectbox("Filtrar por Cia", _cias_opcoes, key="cons_voos_cia")
                with _fv2:
                    _cli_opcoes = ["Todos"] + sorted(df_voos_c["Cliente"].dropna().unique().tolist())
                    _cli_filtro = st.selectbox("Filtrar por Cliente", _cli_opcoes, key="cons_voos_cli")
                with _fv3:
                    _prot_filtro = st.text_input("Protocolo", placeholder="Digite o protocolo...", key="cons_voos_prot")

                if _prot_filtro.strip():
                    # Busca direta no BigQuery — ignora período e consolidador selecionado
                    with st.spinner(f"Buscando protocolo {_prot_filtro.strip()}..."):
                        _df_voos_fil = q_busca_protocolo(_prot_filtro.strip())
                    if _df_voos_fil.empty:
                        st.warning(f"Nenhum registro encontrado para o protocolo **{_prot_filtro.strip()}**.")
                    else:
                        st.caption(f"🔍 Busca global por protocolo — resultado fora do filtro de período")
                else:
                    _df_voos_fil = df_voos_c.copy()
                    if _cia_filtro != "Todas":
                        _df_voos_fil = _df_voos_fil[_df_voos_fil["Cia"] == _cia_filtro]
                    if _cli_filtro != "Todos":
                        _df_voos_fil = _df_voos_fil[_df_voos_fil["Cliente"] == _cli_filtro]

                # Botão exportar Excel (resultado completo filtrado)
                _filtros_label_parts = []
                if _cia_filtro != "Todas":
                    _filtros_label_parts.append(_cia_filtro)
                if _cli_filtro != "Todos":
                    _filtros_label_parts.append(_cli_filtro)
                if _prot_filtro.strip():
                    _filtros_label_parts.append(f"Protocolo {_prot_filtro.strip()}")
                _filtros_label = " · ".join(_filtros_label_parts) if _filtros_label_parts else "Todos"
                _periodo_label = f"{i_str} a {f_str}"
                _xlsx_bytes = _gerar_excel_voos(_df_voos_fil, _filtros_label, _periodo_label)
                st.download_button(
                    label=f"📥 Exportar Excel ({len(_df_voos_fil)} registros)",
                    data=_xlsx_bytes,
                    file_name=f"voos_consolidadores_{i_str}_{f_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="cons_voos_export",
                )

                # Paginação
                _PAGE_SIZE = 50
                _total_rows = len(_df_voos_fil)
                _total_pages = max(1, -(-_total_rows // _PAGE_SIZE))  # ceil division

                # Reinicia página ao mudar filtros
                _filtros_key = (_cia_filtro, _cli_filtro, _prot_filtro.strip())
                if st.session_state.get("cons_voos_filtros_prev") != _filtros_key:
                    st.session_state["cons_voos_page"] = 0
                    st.session_state["cons_voos_filtros_prev"] = _filtros_key
                _page = st.session_state.get("cons_voos_page", 0)

                _start = _page * _PAGE_SIZE
                _end   = min(_start + _PAGE_SIZE, _total_rows)
                _df_voos_page = _df_voos_fil.iloc[_start:_end].copy()

                _gmv_total_voos = _df_voos_fil["GMV"].sum()
                _df_voos_page["GMV"] = _df_voos_page["GMV"].apply(brl)

                st.dataframe(
                    _df_voos_page,
                    use_container_width=True,
                    hide_index=True,
                    height=min(520, (_end - _start + 1) * 35 + 10),
                )

                # Rodapé: GMV + paginação
                _rf_info, _rf_nav = st.columns([3, 2])
                with _rf_info:
                    st.markdown(
                        f"<div style='font-size:0.88rem;font-weight:700;color:#0F172A;padding:6px 0 0 0;'>"
                        f"Total GMV: <span style='color:#0EA5E9;font-size:1rem;'>{brl(_gmv_total_voos)}</span>"
                        f"&nbsp;&nbsp;·&nbsp;&nbsp;{_total_rows} voo{'s' if _total_rows != 1 else ''}"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
                with _rf_nav:
                    _pn1, _pn2, _pn3 = st.columns([1, 2, 1])
                    with _pn1:
                        if st.button("◀", key="cons_voos_prev", disabled=(_page == 0)):
                            st.session_state["cons_voos_page"] = _page - 1
                            st.rerun()
                    with _pn2:
                        st.markdown(
                            f"<div style='text-align:center;font-size:0.82rem;color:#64748B;padding-top:6px;'>"
                            f"Pág. {_page + 1} de {_total_pages}"
                            f"&nbsp;·&nbsp;{_start + 1}–{_end}"
                            f"</div>",
                            unsafe_allow_html=True,
                        )
                    with _pn3:
                        if st.button("▶", key="cons_voos_next", disabled=(_page >= _total_pages - 1)):
                            st.session_state["cons_voos_page"] = _page + 1
                            st.rerun()

    # ══════════════════════════════════════════════════════════════════════════
    # ABA: EVOLUÇÃO
    # ══════════════════════════════════════════════════════════════════════════
    with _tab_evo:

        _evo_opcoes = df_cons["Consolidador"].tolist()
        _evo_sel = st.multiselect(
            "Selecione consolidadores (vazio = todos)",
            options=_evo_opcoes,
            placeholder="Todos os consolidadores...",
            key="cons_evo_sel",
        )
        _evo_filtro = tuple(_evo_sel) if _evo_sel else None

        _evo_label = ", ".join(_evo_sel) if _evo_sel else "todos os consolidadores"

        df_evo = None
        with st.spinner(f"⏳ Consultando BigQuery — evolução de **{_evo_label}**..."):
            try:
                df_evo = q_consolidador_evolucao_mensal(_evo_filtro, i_str, f_str)
            except Exception as _e:
                st.error(f"Erro ao consultar BigQuery: {_e}")

        if df_evo is None or df_evo.empty:
            st.info("Nenhum dado encontrado para o período e filtro selecionados.")
        else:
            # KPIs
            _evo_c1, _evo_c2, _evo_c3 = st.columns(3)
            with _evo_c1:
                st.markdown(f'<div class="kpi-card"><p class="kpi-label">GMV Total no Período</p>'
                            f'<p class="kpi-value">{brl(df_evo["GMV"].sum())}</p></div>',
                            unsafe_allow_html=True)
            with _evo_c2:
                st.markdown(f'<div class="kpi-card orange"><p class="kpi-label">Meses com Emissão</p>'
                            f'<p class="kpi-value">{len(df_evo)}</p></div>',
                            unsafe_allow_html=True)
            with _evo_c3:
                _evo_media = df_evo["GMV"].mean()
                st.markdown(f'<div class="kpi-card green"><p class="kpi-label">Média Mensal</p>'
                            f'<p class="kpi-value">{brl(_evo_media)}</p></div>',
                            unsafe_allow_html=True)

            st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)

            # Gráfico de barras verticais — GMV por mês
            import base64 as _b64
            _logo_path = os.path.join(os.path.dirname(__file__), "assets", "logo.svg")
            _logo_uri = ""
            if os.path.exists(_logo_path):
                with open(_logo_path, "rb") as _lf:
                    _logo_uri = "data:image/svg+xml;base64," + _b64.b64encode(_lf.read()).decode()

            _titulo_evo = f"GMV Mensal · {_evo_label}"
            _fig_evo = go.Figure()
            _fig_evo.add_trace(go.Bar(
                x=df_evo["Mês"],
                y=df_evo["GMV"],
                marker_color=ONFLY_BLUE,
                text=[brl(v) for v in df_evo["GMV"]],
                textposition="outside",
                hovertemplate="<b>%{x}</b><br>GMV: R$ %{y:,.2f}<extra></extra>",
            ))
            _fig_evo.update_layout(
                title=dict(
                    text="Emissões realizadas em Consolidadores",
                    font=dict(size=18, color="#0F172A"),
                    x=0.02, xanchor="left",
                    pad=dict(l=8),
                ),
                xaxis_title="",
                yaxis_title="GMV (R$)",
                yaxis=dict(tickprefix="R$ ", tickformat=",.0f"),
                bargap=0.35,
                annotations=[dict(
                    text=_titulo_evo,
                    xref="paper", yref="paper",
                    x=0.5, y=-0.12,
                    xanchor="center", yanchor="top",
                    showarrow=False,
                    font=dict(size=11, color="#8C9BAB"),
                )],
            )
            plotly_layout(_fig_evo, height=440)
            # Margem: topo para título + logo, base para legenda rodapé
            _fig_evo.update_layout(margin=dict(l=16, r=8, t=72, b=48))

            # Logo Onfly — canto superior direito
            if _logo_uri:
                _fig_evo.add_layout_image(dict(
                    source=_logo_uri,
                    xref="paper", yref="paper",
                    x=1.0, y=1.18,
                    sizex=0.18, sizey=0.18,
                    xanchor="right", yanchor="top",
                    layer="above",
                ))

            # Botão de download da imagem (para WhatsApp)
            _col_chart, _col_btn = st.columns([8, 1])
            with _col_chart:
                st.plotly_chart(_fig_evo, use_container_width=True)
            with _col_btn:
                try:
                    import plotly.io as _pio
                    # Cópia para exportação: labels dentro das barras, margens generosas
                    import copy as _copy
                    _fig_export = _copy.deepcopy(_fig_evo)
                    _fig_export.update_traces(
                        textposition="inside",
                        insidetextanchor="middle",
                        textfont=dict(color="white", size=11),
                    )
                    _fig_export.update_layout(margin=dict(l=32, r=32, t=88, b=72))
                    _fig_export.update_yaxes(showticklabels=False, title_text="")
                    _img_bytes = _pio.to_image(_fig_export, format="png", width=1100, height=580, scale=2)
                    _img_fname = f"consolidadores_{i_str}_{f_str}.png".replace(" ", "_")
                    st.markdown("<div style='height:120px'></div>", unsafe_allow_html=True)
                    st.download_button(
                        label="📸",
                        data=_img_bytes,
                        file_name=_img_fname,
                        mime="image/png",
                        use_container_width=True,
                        help="Baixar imagem do gráfico para enviar pelo WhatsApp",
                    )
                except Exception:
                    pass

            # Tabela auxiliar
            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
            _df_evo_show = df_evo[["Mês", "GMV"]].copy()
            _df_evo_show["GMV"] = _df_evo_show["GMV"].apply(brl)
            st.dataframe(
                _df_evo_show,
                use_container_width=True,
                hide_index=True,
                height=min(50 + len(_df_evo_show) * 35, 400),
            )

# ══════════════════════════════════════════════════════════════════════════════
# 🤝  CRM Aéreo
# ══════════════════════════════════════════════════════════════════════════════
elif secao == "🤝  CRM Aéreo":
    import json as _json
    import streamlit.components.v1 as _components

    # ── Banner de separação ────────────────────────────────────────────────
    st.markdown("""
        <div style="display:flex;align-items:center;gap:12px;padding:10px 16px;
                    background:#F0F7FF;border-left:4px solid #1890FF;
                    border-radius:0 8px 8px 0;margin-bottom:20px;">
            <span style="font-size:20px;">🤝</span>
            <div>
                <div style="font-size:14px;font-weight:700;color:#1890FF;">CRM Aéreo — Sourcing</div>
                <div style="font-size:11px;color:#6C757D;margin-top:1px;">
                    Painel externo integrado · BigQuery ao vivo · Perfil/Visitas/Mercado salvos localmente
                </div>
            </div>
        </div>
    """, unsafe_allow_html=True)

    # ── Seletor de cia (sidebar) ───────────────────────────────────────────
    with st.spinner("Carregando cias aéreas..."):
        _crm_airlines = q_crm_airlines()

    _airline_opts = [r["a"] for r in _crm_airlines]
    _crm_cia = st.sidebar.selectbox(
        "Cia Aérea (CRM)",
        _airline_opts,
        index=None,
        placeholder="Selecione...",
        label_visibility="visible",
        key="crm_cia_sel",
    )

    if not _crm_cia:
        st.info("Selecione uma cia aérea no menu lateral para carregar o CRM.")
    else:
        _from = inicio.strftime("%Y-%m-%d")
        _to   = fim.strftime("%Y-%m-%d")

        with st.spinner(f"Carregando dados de {_crm_cia} ({_from} → {_to})..."):
            _crm_trans = q_crm_transacional(_crm_cia, _from, _to)
            _crm_gmail = q_crm_gmail(_crm_cia)  # None = não configurado, [] = sem resultados

        _crm_data = {
            "selectedAirline": _crm_cia,
            "airlines":        _crm_airlines,
            "fromDate":        _from,
            "toDate":          _to,
            "transacional":    _crm_trans,
            "gmail":           _crm_gmail,  # None → OAuth não configurado
        }

        _tmpl_path = os.path.join(os.path.dirname(__file__), "assets", "crm_aereo_template.html")
        _html = open(_tmpl_path, encoding="utf-8").read()
        _html = _html.replace(
            "/* __CRM_DATA_PLACEHOLDER__ */",
            f"window.__CRM__ = {_json.dumps(_crm_data, ensure_ascii=False)};",
        )

        _components.html(_html, height=1350, scrolling=True)

elif secao == "🔍  Buscas & Conversão":

    st.markdown('<div class="sec-header-wrap"><p class="sec-header">🔍 Buscas & Conversão</p></div>',
                unsafe_allow_html=True)

    with st.spinner("Carregando dados de buscas..."):
        df_conv_mes  = q_buscas_conversao_mensal(i_str, f_str)
        df_dest_int  = q_buscas_destinos(i_str, f_str, internacional=True, limit=20)
        df_dest_dom  = q_buscas_destinos(i_str, f_str, internacional=False, limit=20)
        df_emp_conv  = q_buscas_conversao_empresas(i_str, f_str, min_buscas=300)

    # ── KPIs ──
    total_buscas   = int(df_conv_mes["Buscas"].sum())   if not df_conv_mes.empty else 0
    total_emissoes = int(df_conv_mes["Emissões"].sum()) if not df_conv_mes.empty else 0
    conv_geral     = round(total_emissoes * 100 / total_buscas, 2) if total_buscas else 0

    c1, c2, c3 = st.columns(3)
    c1.metric("Total de Buscas", f"{total_buscas:,}".replace(",", "."))
    c2.metric("Total de Emissões", f"{total_emissoes:,}".replace(",", "."))
    c3.metric("Conversão Geral", f"{conv_geral}%")

    st.markdown("---")

    # ── Gráfico conversão mensal ──
    if not df_conv_mes.empty:
        st.markdown("#### Buscas e Conversão por Mês")
        import plotly.graph_objects as go
        fig_conv = go.Figure()
        fig_conv.add_bar(x=df_conv_mes["Mês"], y=df_conv_mes["Buscas"],
                         name="Buscas", marker_color=ONFLY_BLUE, opacity=0.7)
        fig_conv.add_bar(x=df_conv_mes["Mês"], y=df_conv_mes["Emissões"],
                         name="Emissões", marker_color=ONFLY_GREEN, opacity=0.9)
        fig_conv.add_scatter(x=df_conv_mes["Mês"], y=df_conv_mes["Conversão (%)"],
                             name="Conversão (%)", mode="lines+markers",
                             line=dict(color=ONFLY_ORANGE, width=2),
                             yaxis="y2")
        fig_conv.update_layout(
            barmode="group", yaxis2=dict(overlaying="y", side="right", title="Conversão (%)"),
            height=380, margin=dict(t=20, b=20),
            legend=dict(orientation="h", y=-0.15),
        )
        st.plotly_chart(fig_conv, use_container_width=True)

    st.markdown("---")

    # ── Destinos ──
    col_int, col_dom = st.columns(2)

    with col_int:
        st.markdown("#### 🌍 Top 20 Destinos Internacionais")
        if not df_dest_int.empty:
            st.dataframe(
                df_dest_int[["Destino", "País", "Buscas", "% do Total"]].style.format(
                    {"Buscas": lambda x: f"{int(x):,}".replace(",", "."), "% do Total": lambda x: f"{x:.1f}%"}
                ),
                use_container_width=True, hide_index=True, height=600,
            )

    with col_dom:
        st.markdown("#### 🇧🇷 Top 20 Destinos Domésticos")
        if not df_dest_dom.empty:
            st.dataframe(
                df_dest_dom[["Destino", "Buscas", "% do Total"]].style.format(
                    {"Buscas": lambda x: f"{int(x):,}".replace(",", "."), "% do Total": lambda x: f"{x:.1f}%"}
                ),
                use_container_width=True, hide_index=True, height=600,
            )

    st.markdown("---")

    # ── Empresas com menor conversão ──
    st.markdown("#### ⚠️ Empresas com Maior Potencial de Ativação (menor conversão)")
    st.caption("Empresas com mínimo 300 buscas no período, ordenadas da menor para maior conversão.")
    if not df_emp_conv.empty:
        df_show = df_emp_conv[df_emp_conv["Empresa"].str.strip().str.upper() != df_emp_conv["ID"].astype(str)].copy()
        df_show["GMV"] = df_show["GMV"].apply(lambda x: f"R$ {x:,.2f}".replace(",","X").replace(".",",").replace("X","."))
        df_show["Buscas"] = df_show["Buscas"].apply(lambda x: f"{x:,}".replace(",","."))
        df_show["Emissões"] = df_show["Emissões"].apply(lambda x: f"{x:,}".replace(",","."))
        df_show["Conversão (%)"] = df_show["Conversão (%)"].apply(lambda x: f"{x:.1f}%")
        st.dataframe(
            df_show[["Empresa", "Buscas", "Emissões", "Conversão (%)", "GMV"]],
            use_container_width=True, hide_index=True, height=700,
        )
